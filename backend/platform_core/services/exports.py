import csv
import hashlib
import io
from decimal import Decimal

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle

from ..models import ExportEvent


def _formats(tenant):
    settings=(getattr(tenant,'settings',None) or {}) if tenant else {}
    return str(settings.get('export_date_format') or '%Y-%m-%d'),str(settings.get('export_datetime_format') or '%Y-%m-%d %H:%M:%S%z')


def _formula_safe(value):
    text='' if value is None else str(value)
    return "'"+text if text[:1] in {'=','+','-','@'} else text


def _value(obj,field,tenant=None):
    v=getattr(obj,field,None);date_fmt,datetime_fmt=_formats(tenant)
    if isinstance(v,Decimal):return format(v,'f')
    if hasattr(v,'hour') and hasattr(v,'strftime'):return v.strftime(datetime_fmt)
    if hasattr(v,'strftime'):return v.strftime(date_fmt)
    return _formula_safe(v)


def _metadata(tenant,export_type,filters):
    return [
        ['Export',export_type],['Tenant',getattr(tenant,'name','')],['Generated',timezone.now().isoformat()],
        ['Locale',getattr(tenant,'locale','')],['Default currency',str((getattr(tenant,'settings',{}) or {}).get('default_currency') or 'PHP').upper()],
        ['Filters',str(filters or {})],
    ]


def export_queryset(request,qs,fields,export_type,fmt='csv',approval=None,filters=None):
    rows=list(qs[:50000]);fmt=fmt.lower();tenant=getattr(request,'tenant',None);filters=dict(filters if filters is not None else request.GET.items());meta=_metadata(tenant,export_type,filters)
    if fmt=='csv':
        out=io.StringIO();w=csv.writer(out,lineterminator='\n');[w.writerow(x) for x in meta];w.writerow([]);w.writerow(fields);[w.writerow([_value(o,f,tenant) for f in fields]) for o in rows];blob=out.getvalue().encode('utf-8-sig');mime='text/csv'
    elif fmt=='xlsx':
        wb=Workbook();summary=wb.active;summary.title='Metadata';[summary.append(x) for x in meta];detail=wb.create_sheet('Detail');detail.append(fields);[detail.append([_value(o,f,tenant) for f in fields]) for o in rows];detail.freeze_panes='A2';detail.auto_filter.ref=f'A1:{get_column_letter(max(1,len(fields)))}{max(1,len(rows)+1)}';fill=PatternFill('solid',fgColor='102A43')
        for c in detail[1]:c.font=Font(color='FFFFFF',bold=True);c.fill=fill
        for ws in (summary,detail):
            for col in ws.columns:ws.column_dimensions[get_column_letter(col[0].column)].width=min(45,max(12,max(len(str(c.value or '')) for c in col)+2))
        buf=io.BytesIO();wb.save(buf);blob=buf.getvalue();mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    elif fmt=='pdf':
        buf=io.BytesIO();styles=getSampleStyleSheet();doc=SimpleDocTemplate(buf,pagesize=landscape(A4),rightMargin=15*mm,leftMargin=15*mm,topMargin=15*mm,bottomMargin=15*mm,title=f'{export_type.title()} export');story=[Paragraph(getattr(tenant,'name',''),styles['Heading2']),Paragraph(f'{export_type.title()} Export',styles['Title'])];story.extend(Paragraph(f'<b>{k}</b>: {_formula_safe(v)}',styles['BodyText']) for k,v in meta[2:]);data=[fields]+[[_value(o,f,tenant) for f in fields] for o in rows]
        if rows:
            table=Table(data,repeatRows=1,colWidths=[doc.width/max(1,len(fields))]*len(fields));table.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#102a43')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('GRID',(0,0),(-1,-1),.25,colors.HexColor('#d5dade')),('VALIGN',(0,0),(-1,-1),'TOP'),('FONTSIZE',(0,0),(-1,-1),7)]));story.append(table)
        else:story.append(Paragraph('No data for selected filters',styles['Heading3']))
        doc.build(story);blob=buf.getvalue();mime='application/pdf'
    else:return HttpResponse('unsupported export format',status=400)
    digest=hashlib.sha256(blob).hexdigest();response=HttpResponse(blob,content_type=mime);response['Content-Disposition']=f'attachment; filename="{export_type}-{timezone.now():%Y%m%d}.{fmt}"';response['X-Export-SHA256']=digest;response['Cache-Control']='private, no-store'
    if tenant and getattr(request,'user',None) and request.user.is_authenticated:
        event=ExportEvent.objects.create(tenant=tenant,actor=request.user,export_type=export_type,format=fmt,row_count=len(rows),filters=filters,approved_by=getattr(approval,'approved_by',None),approval_request=approval,result='completed',sha256=digest)
        from .audit import append_audit
        append_audit(request=request,action='export.completed',object_type='ExportEvent',object_id=event.id,metadata={'export_type':export_type,'format':fmt,'filters':filters,'record_count':len(rows),'approver_id':getattr(approval,'approved_by_id',None),'result':'completed','sha256':digest})
    return response
