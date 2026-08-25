from __future__ import annotations

import csv
import hashlib
import io
import re
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal
from pathlib import PurePosixPath
from typing import Any

from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db.models import Count, Q, Sum
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, LETTER, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.platypus import KeepTogether, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from ..models import (
    AuditEvent, Attendance, BillingEvent, Candidate, CandidateMentorAssignment,
    CandidateTask, CommunicationCampaign, DeliveryLog, DuesAssessment, DuesCycle,
    Expense, ExportApprovalRequest, ExportEvent, GeneratedReport, Meeting, Member,
    OfficerTerm, ParticipationRecord, Payment, ReportDeliveryLog, SecurityAlert,
    Subscription, SystemJob, Tenant, Visitor, WebhookEvent,
)
from ..domain.report_contracts import validate_report_output

FORMATS = {'pdf', 'csv', 'xlsx'}

# Appendix C commercial report catalogue. Aliases preserve backwards compatibility with
# the earlier UI while all public labels and new saved/scheduled reports use stable IDs.
REPORT_CATALOG: dict[str, dict[str, Any]] = {
    'R-001': {'name':'Monthly Lodge Management Pack','family':'Executive','formats':['pdf','xlsx'],'schedulable':True,'key':'management_pack'},
    'R-002': {'name':'Annual Lodge Summary','family':'Executive','formats':['pdf','xlsx'],'schedulable':True,'key':'annual_summary'},
    'R-003': {'name':'Lodge Health Dashboard Export','family':'Executive','formats':['pdf','xlsx'],'schedulable':False,'key':'health_dashboard'},
    'R-010': {'name':'Membership Register','family':'Membership','formats':['pdf','xlsx','csv'],'schedulable':False,'key':'membership_register'},
    'R-011': {'name':'Membership Movement','family':'Membership','formats':['pdf','xlsx'],'schedulable':True,'key':'membership_movement'},
    'R-012': {'name':'Membership Trend & Retention','family':'Membership','formats':['pdf','xlsx'],'schedulable':True,'key':'membership_trend'},
    'R-013': {'name':'Member Profile / History Export','family':'Membership','formats':['pdf'],'schedulable':False,'key':'member_profile'},
    'R-014': {'name':'Member Contact / Directory','family':'Membership','formats':['pdf','xlsx','csv'],'schedulable':False,'key':'member_directory'},
    'R-020': {'name':'Current Officer Roster','family':'Governance','formats':['pdf','xlsx'],'schedulable':False,'key':'officer_roster'},
    'R-021': {'name':'Officer Term History','family':'Governance','formats':['pdf','xlsx'],'schedulable':False,'key':'officer_history'},
    'R-030': {'name':'Candidate Pipeline','family':'Candidates','formats':['pdf','xlsx'],'schedulable':True,'key':'candidate_pipeline'},
    'R-031': {'name':'Candidate Time-in-Stage','family':'Candidates','formats':['pdf','xlsx'],'schedulable':False,'key':'candidate_time_stage'},
    'R-032': {'name':'Candidate Milestones & Overdue Tasks','family':'Candidates','formats':['pdf','xlsx'],'schedulable':True,'key':'candidate_tasks'},
    'R-033': {'name':'Mentor Workload','family':'Candidates','formats':['pdf','xlsx'],'schedulable':False,'key':'mentor_workload'},
    'R-040': {'name':'Meeting Attendance Report','family':'Meetings','formats':['pdf','xlsx','csv'],'schedulable':True,'key':'attendance_report'},
    'R-041': {'name':'Attendance Trend','family':'Meetings','formats':['pdf','xlsx'],'schedulable':True,'key':'attendance_trend'},
    'R-042': {'name':'Member Participation Report','family':'Meetings','formats':['pdf','xlsx'],'schedulable':False,'key':'participation'},
    'R-043': {'name':'Visitor Register','family':'Meetings','formats':['pdf','xlsx'],'schedulable':False,'key':'visitor_register'},
    'R-044': {'name':'Meeting Pack Index','family':'Meetings','formats':['pdf'],'schedulable':True,'key':'meeting_pack_index'},
    'R-050': {'name':'Dues Cycle Summary','family':'Finance','formats':['pdf','xlsx'],'schedulable':True,'key':'dues_cycle_summary'},
    'R-051': {'name':'Arrears Aging','family':'Finance','formats':['pdf','xlsx','csv'],'schedulable':True,'key':'arrears_aging'},
    'R-052': {'name':'Collection Performance','family':'Finance','formats':['pdf','xlsx'],'schedulable':True,'key':'collection_performance'},
    'R-053': {'name':'Payment Register','family':'Finance','formats':['pdf','xlsx','csv'],'schedulable':False,'key':'payment_register'},
    'R-054': {'name':'Member Statement','family':'Finance','formats':['pdf'],'schedulable':False,'key':'member_statement'},
    'R-055': {'name':'Receipt Register','family':'Finance','formats':['pdf'],'schedulable':False,'key':'receipt_register'},
    'R-056': {'name':'Expense Report','family':'Finance','formats':['pdf','xlsx','csv'],'schedulable':True,'key':'expense_report'},
    'R-057': {'name':'Finance Summary / Treasurer Pack','family':'Finance','formats':['pdf','xlsx'],'schedulable':True,'key':'finance_pack'},
    'R-060': {'name':'Campaign Delivery Report','family':'Communications','formats':['pdf','xlsx','csv'],'schedulable':False,'key':'campaign_delivery'},
    'R-061': {'name':'Communication Activity Summary','family':'Communications','formats':['pdf','xlsx'],'schedulable':True,'key':'communication_summary'},
    'R-070': {'name':'Audit Activity Report','family':'Compliance','formats':['pdf','xlsx','csv'],'schedulable':False,'key':'audit_activity','sensitive':True},
    'R-071': {'name':'Export History Report','family':'Compliance','formats':['pdf','xlsx','csv'],'schedulable':False,'key':'export_history','sensitive':True},
    'R-072': {'name':'Sensitive Approval Report','family':'Compliance','formats':['pdf','xlsx'],'schedulable':False,'key':'sensitive_approval','sensitive':True},
    'R-073': {'name':'Administrative Change Report','family':'Compliance','formats':['pdf','xlsx'],'schedulable':False,'key':'admin_changes','sensitive':True},
    'R-080': {'name':'Tenant Health Report','family':'Platform','formats':['pdf','xlsx'],'schedulable':False,'key':'tenant_health','platform':True},
    'R-081': {'name':'Subscription / Billing Operations Report','family':'Platform','formats':['pdf','xlsx'],'schedulable':False,'key':'billing_operations','platform':True},
    'R-082': {'name':'Webhook Exception Report','family':'Platform','formats':['pdf','xlsx'],'schedulable':False,'key':'webhook_exceptions','platform':True},
    'R-083': {'name':'Background Job Failure Report','family':'Platform','formats':['pdf','xlsx'],'schedulable':False,'key':'job_failures','platform':True},
    'R-084': {'name':'Email Delivery Exception Report','family':'Platform','formats':['pdf','xlsx'],'schedulable':False,'key':'email_exceptions','platform':True},
}

ALIASES = {
    'membership':'R-010','members':'R-010','attendance':'R-040','meetings':'R-040',
    'candidates':'R-030','finance':'R-057','compliance':'R-070','audit':'R-070',
    'management_pack':'R-001',
}
REPORT_TYPES = set(REPORT_CATALOG) | set(ALIASES)


def canonical_report_id(report_type: str) -> str:
    value=str(report_type or '').strip()
    value=ALIASES.get(value, value.upper())
    if value not in REPORT_CATALOG:
        raise ValueError('unsupported_report_type')
    return value


def report_catalog(include_platform=False):
    rows=[]
    for report_id, definition in REPORT_CATALOG.items():
        if definition.get('platform') and not include_platform:
            continue
        rows.append({'id':report_id, **definition})
    return rows


def _date_filter(qs, field, filters):
    start=(filters or {}).get('start_date'); end=(filters or {}).get('end_date')
    if start: qs=qs.filter(**{f'{field}__date__gte' if field.endswith('_at') else f'{field}__gte': start})
    if end: qs=qs.filter(**{f'{field}__date__lte' if field.endswith('_at') else f'{field}__lte': end})
    return qs


def _safe_text(value):
    """Prevent spreadsheet formula injection while keeping visible user data."""
    if value is None: return ''
    text=str(value)
    return "'"+text if text[:1] in {'=','+','-','@'} else text


def _money(value): return format(Decimal(value or 0), '.2f')


def _outstanding(a):
    return max(Decimal('0'),Decimal(a.amount)+Decimal(a.late_fee)+Decimal(a.adjustment_amount)-Decimal(a.waived_amount)-Decimal(a.paid_amount))


def _parse_iso_at(value):
    if not value: return None
    try:
        parsed=datetime.fromisoformat(str(value).replace('Z','+00:00'))
        return timezone.make_aware(parsed,timezone.get_current_timezone()) if timezone.is_naive(parsed) else parsed
    except Exception: return None


def _report_data(tenant, report_type, filters=None):
    """Authorized report query engine used by preview, export, async and schedule paths."""
    filters=filters or {}; report_id=canonical_report_id(report_type); key=REPORT_CATALOG[report_id]['key']; today=timezone.localdate()
    columns=[]; rows=[]; kpis=[]; trend=[]

    if key in {'membership_register','member_directory'}:
        qs=Member.objects.filter(tenant=tenant).order_by('last_name','first_name')
        if filters.get('status'): qs=qs.filter(status=filters['status'])
        if filters.get('tag'): qs=qs.filter(tags__contains=[filters['tag']])
        include_contact=key=='member_directory';custom_fields=[]
        if not include_contact:
            try:custom_fields=list(tenant.profile.custom_fields or [])
            except Exception:custom_fields=[]
        columns=['Member ID','Name','Status','Degree','Joined'] + (['Email','Phone'] if include_contact else ['Tags / groups','Custom fields'])
        for m in qs[:50000]:
            base=[str(m.id),f'{m.first_name} {m.last_name}',m.status,m.degree,m.joined_on.isoformat() if m.joined_on else '']
            custom='; '.join(f"{field.get('label') or field.get('key')}: {(m.custom_data or {}).get(field.get('key'),'')}" for field in custom_fields if isinstance(field,dict) and field.get('key'))
            rows.append(base+([m.email,m.phone] if include_contact else [', '.join(m.tags or []),custom]))
        kpis=[('Records',qs.count()),('Active',qs.filter(status='active').count()),('MM',qs.filter(degree='MM').count())]
    elif key=='membership_movement':
        qs=AuditEvent.objects.filter(tenant=tenant,object_type='Member').order_by('-occurred_at'); qs=_date_filter(qs,'occurred_at',filters)
        columns=['Date','Member ID','Action','Previous status','New status','Actor']
        for e in qs[:50000]:
            old=(e.before or {}).get('status',''); new=(e.after or {}).get('status','')
            if e.action.endswith('created') or old!=new:
                rows.append([e.occurred_at.date().isoformat(),e.object_id,e.action,old,new,e.actor.get_username() if e.actor else 'system'])
        kpis=[('Movements',len(rows))]
    elif key=='membership_trend':
        qs=Member.objects.filter(tenant=tenant);columns=['Month','Opening','Joined','Exited','Closing','Growth','Churn %','Retention %'];joined=defaultdict(int);exited=defaultdict(int)
        for value in qs.exclude(joined_on=None).values_list('joined_on',flat=True):joined[value.strftime('%Y-%m')]+=1
        movement=AuditEvent.objects.filter(tenant=tenant,object_type='Member').order_by('occurred_at');movement=_date_filter(movement,'occurred_at',filters)
        for event in movement:
            old=(event.before or {}).get('status');new=(event.after or {}).get('status')
            if old=='active' and new and new!='active':exited[event.occurred_at.strftime('%Y-%m')]+=1
            elif old and old!='active' and new=='active':joined[event.occurred_at.strftime('%Y-%m')]+=1
        running=0
        for month in sorted(set(joined)|set(exited)):
            opening=running;joins=joined[month];leaves=exited[month];running=max(0,opening+joins-leaves);churn=round(100*leaves/max(opening,1),1);retention=round(100-churn,1);rows.append([month,opening,joins,leaves,running,joins-leaves,churn,retention]);trend.append({'label':month,'value':running})
        kpis=[('Active',qs.filter(status='active').count()),('Inactive / exited',qs.exclude(status='active').count()),('Retention %',round(100*qs.filter(status='active').count()/max(qs.count(),1),1))]
    elif key=='member_profile':
        member_id=filters.get('member_id')
        m=Member.objects.filter(tenant=tenant,pk=member_id).first() if member_id else None
        if not m: raise ValueError('member_id_required_or_not_found')
        columns=['Field','Value']; rows=[['Member ID',m.id],['Name',f'{m.first_name} {m.last_name}'],['Status',m.status],['Degree',m.degree],['Joined',m.joined_on or ''],['Email',m.email],['Phone',m.phone],['Tags',', '.join(m.tags or [])]]
        for e in AuditEvent.objects.filter(tenant=tenant,object_type='Member',object_id=str(m.id)).order_by('-occurred_at')[:200]: rows.append([f'History · {e.occurred_at:%Y-%m-%d %H:%M}',e.action])
    elif key in {'officer_roster','officer_history'}:
        qs=OfficerTerm.objects.filter(tenant=tenant).select_related('member').order_by('office','-starts_on')
        if key=='officer_roster': qs=qs.filter(is_current=True)
        qs=_date_filter(qs,'starts_on',filters)
        columns=['Office','Officer','Start','End','Current']; rows=[[o.office,f'{o.member.first_name} {o.member.last_name}',o.starts_on.isoformat(),o.ends_on.isoformat() if o.ends_on else '', 'Yes' if o.is_current else 'No'] for o in qs[:50000]]
        kpis=[('Officer terms',qs.count())]
    elif key=='candidate_pipeline':
        qs=Candidate.objects.filter(tenant=tenant).select_related('member').order_by('stage_order','member__last_name')
        if filters.get('stage'): qs=qs.filter(stage=filters['stage'])
        columns=['Candidate','Stage','Status','Days in current stage','Target degree']
        for c in qs[:50000]:
            history=c.stage_history or []; at=_parse_iso_at(history[-1].get('at')) if history else None
            days=max(0,(timezone.now()-at).days) if at and timezone.is_aware(at) else ''
            rows.append([f'{c.member.first_name} {c.member.last_name}',c.stage,c.status,days,c.target_degree_date.isoformat() if c.target_degree_date else ''])
        by_stage=list(qs.values('stage').annotate(count=Count('id')).order_by('stage'))
        trend=[{'label':x['stage'],'value':x['count']} for x in by_stage]; kpis=[('Candidates',qs.count()),('Stages represented',len(by_stage))]
    elif key=='candidate_time_stage':
        qs=Candidate.objects.filter(tenant=tenant).select_related('member').order_by('member__last_name');thresholds=((tenant.settings or {}).get('candidate_stage_threshold_days') or {});default_threshold=max(1,int(thresholds.get('default',60) or 60));columns=['Candidate','Stage','Entered','Exited','Days','Threshold','Exception']
        for candidate in qs[:50000]:
            history=[item for item in (candidate.stage_history or []) if isinstance(item,dict) and item.get('to') and _parse_iso_at(item.get('at'))]
            if not history:history=[{'to':candidate.stage,'at':candidate.created_at.isoformat()}]
            for index,item in enumerate(history):
                entered=_parse_iso_at(item['at']);exited=_parse_iso_at(history[index+1]['at']) if index+1<len(history) else timezone.now();days=max(0,(exited-entered).days) if entered and exited else 0;threshold=max(1,int(thresholds.get(item['to'],default_threshold) or default_threshold));rows.append([f'{candidate.member.first_name} {candidate.member.last_name}',item['to'],entered.isoformat() if entered else '',exited.isoformat() if index+1<len(history) and exited else '',days,threshold,'Overdue' if days>threshold else ''])
        kpis=[('Stage periods',len(rows)),('Exceptions',sum(1 for row in rows if row[-1]))]
    elif key=='candidate_tasks':
        qs=CandidateTask.objects.filter(tenant=tenant).select_related('candidate__member','assignee').order_by('due_on')
        if filters.get('status'): qs=qs.filter(status=filters['status'])
        columns=['Candidate','Task','Status','Due','Overdue','Assignee / mentor']; rows=[]
        for t in qs[:50000]: rows.append([f'{t.candidate.member.first_name} {t.candidate.member.last_name}',t.title,t.status,t.due_on.isoformat() if t.due_on else '', 'Yes' if t.due_on and t.due_on<today and not t.completed_at else 'No',f'{t.assignee.first_name} {t.assignee.last_name}' if t.assignee else ''])
        kpis=[('Open tasks',qs.exclude(status__in=['completed','cancelled']).count()),('Overdue',qs.filter(due_on__lt=today,completed_at__isnull=True).count())]
    elif key=='mentor_workload':
        qs=CandidateMentorAssignment.objects.filter(tenant=tenant,is_active=True).select_related('mentor','candidate')
        columns=['Mentor','Active candidates','Open tasks','Overdue tasks']; grouped={}
        for a in qs:
            g=grouped.setdefault(a.mentor_id,{'mentor':f'{a.mentor.first_name} {a.mentor.last_name}','candidates':set(),'open':0,'overdue':0});g['candidates'].add(a.candidate_id);g['open']+=a.candidate.tasks.exclude(status__in=['completed','cancelled']).count();g['overdue']+=a.candidate.tasks.filter(due_on__lt=today,completed_at__isnull=True).count()
        rows=[[g['mentor'],len(g['candidates']),g['open'],g['overdue']] for g in grouped.values()]; rows.sort(key=lambda x:(-x[1],x[0])); trend=[{'label':r[0],'value':r[1]} for r in rows[:12]]
    elif key in {'attendance_report','attendance_trend'}:
        qs=Attendance.objects.filter(tenant=tenant).select_related('meeting','member').order_by('-meeting__starts_at','member__last_name');qs=_date_filter(qs,'meeting__starts_at',filters)
        if filters.get('status'): qs=qs.filter(status=filters['status'])
        if key=='attendance_report':
            columns=['Meeting','Date','Member','Status','Check-in'];rows=[[a.meeting.title,a.meeting.starts_at.date().isoformat(),f'{a.member.first_name} {a.member.last_name}',a.status,a.checked_in_at.isoformat() if a.checked_in_at else ''] for a in qs[:50000]]
        else:
            columns=['Meeting','Date','Present','Recorded','Attendance %']; grouped=qs.values('meeting_id','meeting__title','meeting__starts_at').annotate(total=Count('id'))
            for g in grouped:
                present=qs.filter(meeting_id=g['meeting_id'],status='present').count(); rate=round(100*present/max(g['total'],1),1);rows.append([g['meeting__title'],g['meeting__starts_at'].date().isoformat(),present,g['total'],rate]);trend.append({'label':g['meeting__starts_at'].date().isoformat(),'value':rate})
        kpis=[('Attendance rows',qs.count()),('Present',qs.filter(status='present').count()),('Excused',qs.filter(status='excused').count())]
    elif key=='participation':
        members=Member.objects.filter(tenant=tenant).order_by('last_name','first_name');columns=['Member','Meetings attended','Attendance records','Participation records','Participation points'];total_points=0
        for member in members[:50000]:
            attendance=Attendance.objects.filter(tenant=tenant,member=member);participation=ParticipationRecord.objects.filter(tenant=tenant,member=member);points=participation.aggregate(value=Sum('points'))['value'] or 0;total_points+=points;rows.append([f'{member.first_name} {member.last_name}',attendance.filter(status='present').count(),attendance.count(),participation.count(),points])
        kpis=[('Members',len(rows)),('Participation points',total_points)]
    elif key=='visitor_register':
        qs=Visitor.objects.filter(tenant=tenant).select_related('meeting').order_by('-meeting__starts_at'); qs=_date_filter(qs,'meeting__starts_at',filters); columns=['Meeting','Date','Visitor','Home lodge','Lodge no.','Jurisdiction','Notes'];rows=[[x.meeting.title,x.meeting.starts_at.date().isoformat(),x.name,x.home_lodge,x.lodge_number,x.jurisdiction,x.notes] for x in qs[:50000]];kpis=[('Visitors',qs.count())]
    elif key=='meeting_pack_index':
        meeting_id=filters.get('meeting_id');meeting=Meeting.objects.filter(tenant=tenant,pk=meeting_id).first() if meeting_id else Meeting.objects.filter(tenant=tenant).order_by('-starts_at').first()
        if not meeting:raise ValueError('meeting_id_required_or_not_found')
        columns=['Section','Item','Status','Reference'];rows=[['Meeting','Title',meeting.status,meeting.title],['Meeting','Date and location','configured',f'{meeting.starts_at.isoformat()} · {meeting.location}']]
        for item in meeting.agenda_items.order_by('sort_order','created_at'):rows.append(['Agenda',item.title,item.status if hasattr(item,'status') else 'included',str(item.id)])
        rows.extend([['Attendance','Recorded members','included',str(meeting.attendance.count())],['Visitors','Recorded visitors','included',str(meeting.visitors.count())]])
        for minute in meeting.minute_versions.filter(state__in=['approved','locked']).order_by('-version')[:1]:rows.append(['Prior minutes',f'Version {minute.version}',minute.state,str(minute.id)])
        for attachment in meeting.pack_attachments.filter(include_in_pack=True).select_related('document').order_by('sort_order','created_at'):rows.append(['Attachment',attachment.label or attachment.document.title,'included',str(attachment.document_id)])
        kpis=[('Pack sections',len(rows)),('Agenda items',meeting.agenda_items.count()),('Attendance rows',meeting.attendance.count())]
    elif key=='dues_cycle_summary':
        cycles=DuesCycle.objects.filter(tenant=tenant).order_by('-period_start');cycles=_date_filter(cycles,'period_start',filters);columns=['Cycle','Period','Currency','Assessments','Base amount','Adjustments','Late fees','Paid','Waived','Outstanding','Collection %'];rows=[]
        for c in cycles[:5000]:
            asses=list(c.assessments.all());base=sum((Decimal(a.amount) for a in asses),Decimal('0'));adjustments=sum((Decimal(a.adjustment_amount) for a in asses),Decimal('0'));late_fees=sum((Decimal(a.late_fee) for a in asses),Decimal('0'));assessed=base+adjustments+late_fees;paid=sum((Decimal(a.paid_amount) for a in asses),Decimal('0'));waived=sum((Decimal(a.waived_amount) for a in asses),Decimal('0'));out=sum((_outstanding(a) for a in asses),Decimal('0'));rate=round(100*paid/max(assessed-waived,Decimal('0.01')),1);rows.append([c.name,f'{c.period_start} to {c.period_end}',c.currency,len(asses),_money(base),_money(adjustments),_money(late_fees),_money(paid),_money(waived),_money(out),rate]);trend.append({'label':c.name,'value':rate})
        kpis=[('Cycles',cycles.count()),('Assessments',sum(row[3] for row in rows))]
    elif key=='collection_performance':
        cycles=DuesCycle.objects.filter(tenant=tenant).order_by('period_start');cycles=_date_filter(cycles,'period_start',filters);columns=['Cycle','Period','Currency','Assessed','Collected','Collection %'];totals=defaultdict(lambda:{'assessed':Decimal('0'),'collected':Decimal('0')})
        for cycle in cycles[:5000]:
            assessments=list(cycle.assessments.all());assessed=sum((Decimal(a.amount)+Decimal(a.adjustment_amount)+Decimal(a.late_fee)-Decimal(a.waived_amount) for a in assessments),Decimal('0'));collected=sum((Decimal(a.paid_amount) for a in assessments),Decimal('0'));rate=round(100*collected/max(assessed,Decimal('0.01')),1);rows.append([cycle.name,f'{cycle.period_start} to {cycle.period_end}',cycle.currency,_money(assessed),_money(collected),rate]);totals[cycle.currency]['assessed']+=assessed;totals[cycle.currency]['collected']+=collected;trend.append({'label':cycle.name,'value':rate})
        kpis=[]
        for currency,total in sorted(totals.items()):kpis.extend([(f'Assessed {currency}',_money(total['assessed'])),(f'Collected {currency}',_money(total['collected']))])
    elif key=='arrears_aging':
        qs=DuesAssessment.objects.filter(tenant=tenant).select_related('member','cycle').order_by('cycle__due_on','member__last_name');columns=['Member','Cycle','Due date','Currency','Outstanding','Aging bucket','Status'];totals=defaultdict(Decimal)
        for a in qs[:50000]:
            out=_outstanding(a)
            if out<=0: continue
            age=max(0,(today-a.cycle.due_on).days); bucket='Current' if age<=0 else '1-30' if age<=30 else '31-60' if age<=60 else '61-90' if age<=90 else '90+'; rows.append([f'{a.member.first_name} {a.member.last_name}',a.cycle.name,a.cycle.due_on.isoformat(),a.currency,_money(out),bucket,a.status]);totals[a.currency]+=out
        kpis=[(f'Outstanding {c}',_money(v)) for c,v in sorted(totals.items())]
    elif key=='payment_register':
        qs=Payment.objects.filter(tenant=tenant).select_related('member','assessment__cycle').order_by('-paid_at');qs=_date_filter(qs,'paid_at',filters);columns=['Receipt','Paid at','Member','Assessment','Amount','Currency','Method','Provider','Reference','Reversed'];rows=[[p.receipt_number,p.paid_at.isoformat(),f'{p.member.first_name} {p.member.last_name}' if p.member else '',p.assessment.cycle.name if p.assessment else '',_money(p.amount),p.currency,p.method,p.provider,p.provider_reference,'Yes' if p.is_reversed else 'No'] for p in qs[:50000]];kpis=[('Payments',qs.filter(is_reversed=False).count()),('Reversed',qs.filter(is_reversed=True).count())]
    elif key=='receipt_register':
        payment_id=filters.get('payment_id');payment=Payment.objects.filter(tenant=tenant,pk=payment_id).select_related('member','assessment__cycle').first() if payment_id else None
        if not payment:raise ValueError('payment_id_required_or_not_found')
        columns=['Field','Value'];rows=[['Receipt number',payment.receipt_number],['Paid at',payment.paid_at.isoformat()],['Member',f'{payment.member.first_name} {payment.member.last_name}' if payment.member else ''],['Assessment',payment.assessment.cycle.name if payment.assessment else ''],['Amount',f'{payment.currency} {_money(payment.amount)}'],['Method',payment.method],['Provider',payment.provider],['Reference',payment.provider_reference],['Status','Reversed' if payment.is_reversed else 'Paid']];kpis=[('Receipt',payment.receipt_number),('Amount',f'{payment.currency} {_money(payment.amount)}')]
    elif key=='member_statement':
        member_id=filters.get('member_id');m=Member.objects.filter(tenant=tenant,pk=member_id).first() if member_id else None
        if not m: raise ValueError('member_id_required_or_not_found')
        asses=DuesAssessment.objects.filter(tenant=tenant,member=m).select_related('cycle').order_by('cycle__period_start');columns=['Date','Type','Reference','Currency','Charge','Payment','Waiver / adjustment','Balance'];running=defaultdict(Decimal)
        for a in asses:
            charge=Decimal(a.amount)+Decimal(a.late_fee);relief=Decimal(a.waived_amount)-Decimal(a.adjustment_amount);running[a.currency]+=charge-relief-Decimal(a.paid_amount);rows.append([a.cycle.period_start,'Assessment',a.cycle.name,a.currency,_money(charge),_money(a.paid_amount),_money(relief),_money(running[a.currency])])
        kpis=[(f'Balance {c}',_money(v)) for c,v in sorted(running.items())]
    elif key=='expense_report':
        qs=Expense.objects.filter(tenant=tenant).order_by('-incurred_on');qs=_date_filter(qs,'incurred_on',filters);columns=['Date','Category','Payee','Description','Amount','Currency','Approval','Attachment'];rows=[[x.incurred_on,x.category,x.vendor,x.description,_money(x.amount),x.currency,x.approval_status,'Yes' if x.attachment_document_id or x.receipt_path else 'No'] for x in qs[:50000]];kpis=[('Expenses',qs.count()),('Approved',qs.filter(approval_status='approved').count())]
    elif key=='finance_pack':
        assessments=_date_filter(DuesAssessment.objects.filter(tenant=tenant),'cycle__period_start',filters);payments=_date_filter(Payment.objects.filter(tenant=tenant,is_reversed=False),'paid_at',filters);expenses=_date_filter(Expense.objects.filter(tenant=tenant,approval_status='approved'),'incurred_on',filters);totals=defaultdict(lambda:defaultdict(Decimal))
        for item in assessments:totals[item.currency]['assessed']+=Decimal(item.amount)+Decimal(item.late_fee)+Decimal(item.adjustment_amount);totals[item.currency]['waived']+=Decimal(item.waived_amount);totals[item.currency]['outstanding']+=_outstanding(item)
        for item in payments:totals[item.currency]['collected']+=Decimal(item.amount);totals[item.currency][f'payment:{item.method}']+=Decimal(item.amount)
        for item in expenses:totals[item.currency]['expenses']+=Decimal(item.amount)
        columns=['Section','Metric','Currency','Amount'];rows=[]
        for currency,total in sorted(totals.items()):
            for key_name,label in [('assessed','Dues assessed'),('collected','Collections'),('waived','Waivers'),('outstanding','Arrears outstanding'),('expenses','Approved expenses')]:rows.append(['Summary',label,currency,_money(total[key_name])])
            rows.append(['Summary','Net cash after expenses',currency,_money(total['collected']-total['expenses'])])
            for metric,value in sorted(total.items()):
                if metric.startswith('payment:'):rows.append(['Payment mix',metric.split(':',1)[1],currency,_money(value)])
        kpis=[(f'Collections {currency}',_money(total['collected'])) for currency,total in sorted(totals.items())]
    elif key=='campaign_delivery':
        qs=DeliveryLog.objects.filter(tenant=tenant).select_related('campaign').order_by('-created_at');columns=['Campaign','Recipient','Status','Attempts','Provider reference','Delivered'];rows=[[x.campaign.name,x.recipient,x.status,x.attempts,x.provider_id,x.delivered_at.isoformat() if x.delivered_at else ''] for x in qs[:50000]];kpis=[('Delivered',qs.filter(status='delivered').count()),('Failed',qs.filter(status='failed').count())]
    elif key=='communication_summary':
        qs=CommunicationCampaign.objects.filter(tenant=tenant).order_by('-created_at');qs=_date_filter(qs,'created_at',filters);columns=['Campaign','Created','Status','Recipients','Delivered','Failed'];rows=[]
        for c in qs[:10000]: rows.append([c.name,c.created_at.isoformat(),c.status,c.deliveries.count(),c.deliveries.filter(status='delivered').count(),c.deliveries.filter(status='failed').count()])
        kpis=[('Campaigns',qs.count()),('Sent',qs.filter(status='sent').count()),('Failed',qs.filter(status='failed').count())]
    elif key in {'audit_activity','admin_changes'}:
        qs=AuditEvent.objects.filter(tenant=tenant).select_related('actor').order_by('-occurred_at');qs=_date_filter(qs,'occurred_at',filters)
        if filters.get('actor_id'): qs=qs.filter(actor_id=filters['actor_id'])
        if filters.get('action'): qs=qs.filter(action__startswith=filters['action'])
        if filters.get('object_type'): qs=qs.filter(object_type=filters['object_type'])
        if key=='admin_changes': qs=qs.filter(action__regex=r'(settings|role|credential|feature|security|platform)')
        columns=['Time','Actor','Action','Object type','Object ID','Request ID','Hash'];rows=[[x.occurred_at.isoformat(),x.actor.get_username() if x.actor else 'system',x.action,x.object_type,x.object_id,x.request_id,x.event_hash[:12]] for x in qs[:50000]];kpis=[('Audit events',qs.count())]
    elif key=='export_history':
        qs=ExportEvent.objects.filter(tenant=tenant).select_related('actor','approved_by').order_by('-created_at');columns=['Time','Requester','Export','Format','Records','Approver','Result','Hash'];rows=[[x.created_at.isoformat(),x.actor.get_username() if x.actor else 'system',x.export_type,x.format,x.row_count,x.approved_by.get_username() if x.approved_by else '',x.result,x.sha256[:12]] for x in qs[:50000]];kpis=[('Exports',qs.count()),('Failed',qs.exclude(result='completed').count())]
    elif key=='sensitive_approval':
        qs=ExportApprovalRequest.objects.filter(tenant=tenant).select_related('requester','approved_by').order_by('-created_at');columns=['Requested','Requester','Export','Format','Status','Approver','Reason','Expires'];rows=[[x.created_at.isoformat(),x.requester.get_username(),x.export_type,x.format,x.status,x.approved_by.get_username() if x.approved_by else '',x.reason,x.expires_at.isoformat() if x.expires_at else ''] for x in qs[:50000]];kpis=[('Pending',qs.filter(status='pending').count()),('Approved',qs.filter(status='approved').count())]
    elif key=='health_dashboard':
        members=Member.objects.filter(tenant=tenant); meetings=Meeting.objects.filter(tenant=tenant);jobs=SystemJob.objects.filter(tenant=tenant);alerts=SecurityAlert.objects.filter(tenant=tenant,acknowledged_at__isnull=True);columns=['Metric','Value'];rows=[['Active members',members.filter(status='active').count()],['Upcoming meetings',meetings.filter(starts_at__gte=timezone.now()).count()],['Failed jobs',jobs.filter(status='failed').count()],['Open security alerts',alerts.count()],['Generated reports',GeneratedReport.objects.filter(tenant=tenant).count()]];kpis=[tuple(x) for x in rows]
    elif key=='management_pack':
        # Executive pack exposes reconciled section KPIs from the same canonical sources.
        members=Member.objects.filter(tenant=tenant); meetings=Meeting.objects.filter(tenant=tenant); candidates=Candidate.objects.filter(tenant=tenant); assessments=DuesAssessment.objects.filter(tenant=tenant); payments=Payment.objects.filter(tenant=tenant,is_reversed=False); expenses=Expense.objects.filter(tenant=tenant,approval_status='approved')
        outstanding_by_currency=defaultdict(Decimal); paid_by_currency=defaultdict(Decimal); exp_by_currency=defaultdict(Decimal)
        for a in assessments: outstanding_by_currency[a.currency]+=_outstanding(a)
        for p in payments: paid_by_currency[p.currency]+=p.amount
        for e in expenses: exp_by_currency[e.currency]+=e.amount
        columns=['Section','Metric','Value'];rows=[['Membership','Active members',members.filter(status='active').count()],['Membership','Total members',members.count()],['Attendance','Meetings recorded',meetings.count()],['Candidates','Active candidates',candidates.filter(status='active').count()],['Compliance','Audit events',AuditEvent.objects.filter(tenant=tenant).count()],['Compliance','Exports',ExportEvent.objects.filter(tenant=tenant).count()]]
        for c,v in sorted(paid_by_currency.items()):rows.append(['Finance',f'Collected ({c})',_money(v)])
        for c,v in sorted(outstanding_by_currency.items()):rows.append(['Finance',f'Outstanding ({c})',_money(v)])
        for c,v in sorted(exp_by_currency.items()):rows.append(['Finance',f'Approved expenses ({c})',_money(v)])
        kpis=[(r[1],r[2]) for r in rows[:6]]
    elif key=='annual_summary':
        columns=['Year','Section','Metric','Value'];years=set()
        for value in Member.objects.filter(tenant=tenant).exclude(joined_on=None).values_list('joined_on',flat=True):years.add(value.year)
        for value in Meeting.objects.filter(tenant=tenant).values_list('starts_at',flat=True):years.add(value.year)
        if filters.get('year'):
            try:years={int(filters['year'])}
            except (TypeError,ValueError):raise ValueError('invalid_year')
        if not years:years={today.year}
        for year in sorted(years):
            joined=Member.objects.filter(tenant=tenant,joined_on__year=year).count();meetings=Meeting.objects.filter(tenant=tenant,starts_at__year=year);attendance=Attendance.objects.filter(tenant=tenant,meeting__starts_at__year=year);present=attendance.filter(status='present').count();candidate_count=Candidate.objects.filter(tenant=tenant,created_at__year=year).count();officer_terms=OfficerTerm.objects.filter(tenant=tenant,starts_on__year=year).count();rows.extend([[year,'Membership','Members joined',joined],[year,'Attendance','Meetings',meetings.count()],[year,'Attendance','Attendance rate %',round(100*present/max(attendance.count(),1),1)],[year,'Candidates','Candidates added',candidate_count],[year,'Governance','Officer terms started',officer_terms]])
            paid=Payment.objects.filter(tenant=tenant,is_reversed=False,paid_at__year=year).values('currency').annotate(value=Sum('amount'))
            for value in paid:rows.append([year,'Finance',f"Collected {value['currency']}",_money(value['value'])])
        kpis=[('Years',len(years)),('Latest year',max(years))]
    elif key in {'tenant_health','billing_operations','webhook_exceptions','job_failures','email_exceptions'}:
        # Platform report IDs are intentionally tenant-bounded when invoked from the tenant engine.
        if key=='tenant_health':
            columns=['Metric','Value'];rows=[['Tenant',tenant.name],['Status','Active' if tenant.is_active else 'Suspended'],['Plan',tenant.plan],['Members',tenant.members.count()],['Failed jobs',tenant.jobs.filter(status='failed').count()]]
        elif key=='billing_operations':
            sub=Subscription.objects.filter(tenant=tenant).first();columns=['Metric','Value'];rows=[['Plan',sub.plan if sub else tenant.plan],['Status',sub.status if sub else 'not_configured'],['Provider',sub.provider if sub else ''],['Period end',sub.current_period_end.isoformat() if sub and sub.current_period_end else '']]
        elif key=='webhook_exceptions':
            qs=WebhookEvent.objects.filter(tenant=tenant).exclude(status='processed').order_by('-created_at');columns=['Provider','Event','Type','Mode','Status','Attempts','Error'];rows=[[x.provider,x.event_id,x.event_type,x.mode,x.status,x.attempts,x.quarantine_reason or x.last_error] for x in qs[:50000]]
        elif key=='job_failures':
            qs=SystemJob.objects.filter(tenant=tenant,status='failed').order_by('-created_at');columns=['Job','Created','Attempts','Error'];rows=[[x.kind,x.created_at.isoformat(),x.attempts,x.last_error] for x in qs[:50000]]
        else:
            qs=DeliveryLog.objects.filter(tenant=tenant,status__in=['failed','retry_pending']).select_related('campaign').order_by('-created_at');columns=['Campaign','Recipient','Status','Attempts','Error'];rows=[[x.campaign.name,x.recipient,x.status,x.attempts,x.error] for x in qs[:50000]]
    else:
        raise ValueError('unsupported_report_type')

    validate_report_output(report_id,columns,rows)
    return {'report_id':report_id,'definition':REPORT_CATALOG[report_id],'columns':columns,'rows':rows,'kpis':kpis,'trend':trend,'filters':filters}


def report_rows(tenant, report_type, filters=None):
    data=_report_data(tenant,report_type,filters); return data['columns'],data['rows']


def report_preview(tenant, report_type, filters=None, limit=100):
    data=_report_data(tenant,report_type,filters);data['total_rows']=len(data['rows']);data['rows']=data['rows'][:max(1,min(int(limit),500))];return data


def estimate_report_rows(tenant, report_type, filters=None):
    # Accurate enough for queue selection without trusting a client-provided count.
    return len(_report_data(tenant,report_type,filters)['rows'])


class _PlatformReportContext:
    """Presentation context for control-plane reports without a tenant owner."""

    name = 'LodgeFlow Platform'
    slug = 'platform'
    id = 'platform'
    settings = {'reports': {'classification': 'Platform Confidential', 'page_size': 'A4'}}


PLATFORM_REPORT_CONTEXT = _PlatformReportContext()


def _platform_report_data(report_type, filters=None):
    filters=filters or {};report_id=canonical_report_id(report_type);definition=REPORT_CATALOG[report_id]
    if not definition.get('platform'):
        raise ValueError('tenant_report_requires_tenant_context')
    key=definition['key'];columns=[];rows=[];kpis=[];trend=[]
    if key=='tenant_health':
        qs=Tenant.objects.order_by('name').annotate(member_count=Count('members',distinct=True),failed_job_count=Count('jobs',filter=Q(jobs__status='failed'),distinct=True))
        if filters.get('status')=='active':qs=qs.filter(is_active=True)
        if filters.get('status')=='suspended':qs=qs.filter(is_active=False)
        columns=['Tenant ID','Tenant','Lodge number','Status','Plan','Active users','Members','Features','Failed jobs','Stored bytes']
        for x in qs[:50000]:
            generated_bytes=sum(int((item.metadata or {}).get('size_bytes') or 0) for item in x.generated_reports.only('metadata'));document_bytes=x.documents.aggregate(value=Sum('file_size'))['value'] or 0;features=', '.join(x.feature_flags.filter(enabled=True).order_by('key').values_list('key',flat=True));rows.append([str(x.id),x.name,x.lodge_number,'active' if x.is_active else 'suspended',x.plan,x.user_memberships.filter(is_active=True).count(),x.member_count,features,x.failed_job_count,int(document_bytes)+generated_bytes])
        kpis=[('Tenants',qs.count()),('Active',qs.filter(is_active=True).count()),('Suspended',qs.filter(is_active=False).count())]
    elif key=='billing_operations':
        qs=Subscription.objects.select_related('tenant').order_by('tenant__name');qs=_date_filter(qs,'created_at',filters)
        if filters.get('status'):qs=qs.filter(status=filters['status'])
        columns=['Tenant','Record type','Provider','Reference','Plan / event','Status','Amount','Currency','Occurred']
        rows=[[x.tenant.name,'subscription',x.provider,x.provider_subscription_id,x.plan,x.status,'','',x.last_synced_at.isoformat() if x.last_synced_at else x.created_at.isoformat()] for x in qs[:50000]]
        events=BillingEvent.objects.select_related('tenant').order_by('-created_at');events=_date_filter(events,'created_at',filters)
        for x in events[:50000]:rows.append([x.tenant.name if x.tenant else 'Platform','billing event',x.provider,x.provider_reference,x.event_type,x.status,_money(x.amount) if x.amount is not None else '',x.currency,x.created_at.isoformat()])
        kpis=[('Subscriptions',qs.count()),('Active',qs.filter(status__in=['active','trialing']).count()),('Past due',qs.filter(status='past_due').count()),('Billing events',events.count())]
    elif key=='webhook_exceptions':
        qs=WebhookEvent.objects.select_related('tenant').exclude(status='processed').order_by('-created_at');qs=_date_filter(qs,'created_at',filters)
        if filters.get('provider'):qs=qs.filter(provider=filters['provider'])
        columns=['Tenant','Provider','Event','Type','Mode','Status','Attempts','Error']
        rows=[[x.tenant.name if x.tenant else 'Unresolved',x.provider,x.event_id,x.event_type,x.mode,x.status,x.attempts,x.quarantine_reason or x.last_error] for x in qs[:50000]]
        kpis=[('Exceptions',qs.count()),('Quarantined',qs.filter(status='quarantined').count()),('Failed',qs.filter(status='failed').count())]
    elif key=='job_failures':
        qs=SystemJob.objects.select_related('tenant').filter(status='failed').order_by('-created_at');qs=_date_filter(qs,'created_at',filters)
        columns=['Tenant','Job','Created','Age minutes','Attempts','Terminal','Finished','Error']
        rows=[[x.tenant.name if x.tenant else 'Platform',x.kind,x.created_at.isoformat(),max(0,int((timezone.now()-x.created_at).total_seconds()//60)),x.attempts,'Yes' if x.finished_at else 'No',x.finished_at.isoformat() if x.finished_at else '',x.last_error] for x in qs[:50000]]
        kpis=[('Failed jobs',qs.count())]
    elif key=='email_exceptions':
        qs=DeliveryLog.objects.select_related('tenant','campaign').filter(status__in=['failed','retry_pending']).order_by('-created_at');qs=_date_filter(qs,'created_at',filters)
        columns=['Tenant','Campaign','Recipient','Status','Attempts','Last attempt','Error']
        rows=[[x.tenant.name,x.campaign.name,x.recipient,x.status,x.attempts,x.last_attempt_at.isoformat() if x.last_attempt_at else '',x.error] for x in qs[:50000]]
        kpis=[('Exceptions',qs.count()),('Failed',qs.filter(status='failed').count()),('Retry pending',qs.filter(status='retry_pending').count())]
    else:
        raise ValueError('unsupported_platform_report_type')
    validate_report_output(report_id,columns,rows)
    return {'report_id':report_id,'definition':definition,'columns':columns,'rows':rows,'kpis':kpis,'trend':trend,'filters':filters}


def platform_report_preview(report_type, filters=None, limit=100):
    data=_platform_report_data(report_type,filters);data['total_rows']=len(data['rows']);data['rows']=data['rows'][:max(1,min(int(limit),500))];return data


def _period_label(filters):
    filters=filters or {}; start=filters.get('start_date'); end=filters.get('end_date')
    if start or end:return f'{start or "Beginning"} to {end or "Present"}'
    return 'All available authorized records'


def _page_size(tenant, wide):
    cfg=((tenant.settings or {}).get('reports') or {}); base=LETTER if str(cfg.get('page_size','A4')).upper()=='LETTER' else A4
    return landscape(base) if wide else base


class NumberedCanvas(canvas.Canvas):
    def __init__(self,*args,**kwargs): super().__init__(*args,**kwargs);self._saved=[]
    def showPage(self): self._saved.append(dict(self.__dict__));self._startPage()
    def save(self):
        count=len(self._saved)+1;self._saved.append(dict(self.__dict__))
        for state in self._saved:
            self.__dict__.update(state);self.draw_page_number(count);super().showPage()
        super().save()
    def draw_page_number(self,count):
        self.saveState();self.setFont('Helvetica',7);self.setFillColor(colors.HexColor('#5b6570'));self.drawRightString(self._pagesize[0]-15*mm,8*mm,f'Page {self._pageNumber} of {count}');self.restoreState()


def _pdf_bytes(tenant, report_id, definition, columns, rows, filters, kpis):
    cfg=((tenant.settings or {}).get('reports') or {});classification=str(cfg.get('classification','Internal'))[:40];wide=len(columns)>5;pagesize=_page_size(tenant,wide);buf=io.BytesIO();styles=getSampleStyleSheet()
    title=definition['name'];doc=SimpleDocTemplate(buf,pagesize=pagesize,rightMargin=15*mm,leftMargin=15*mm,topMargin=23*mm,bottomMargin=18*mm,title=title,author='Quill of the Craft',subject=f'{report_id} · {_period_label(filters)}')
    small=ParagraphStyle('small',parent=styles['BodyText'],fontSize=7.4,leading=9.2,textColor=colors.HexColor('#1d2a33'))
    meta=ParagraphStyle('meta',parent=styles['BodyText'],fontSize=8,leading=10,textColor=colors.HexColor('#52616b'))
    cover=ParagraphStyle('cover',parent=styles['Title'],fontSize=25,leading=30,textColor=colors.HexColor('#102a43'),alignment=TA_CENTER,spaceAfter=8)
    story=[]
    if report_id in {'R-001','R-002','R-057'}:
        story += [Spacer(1,18*mm),Paragraph(tenant.name,cover),Paragraph(title,styles['Title']),Paragraph(f'{report_id} · {_period_label(filters)}',meta),Spacer(1,8*mm)]
        if kpis:
            kpi_data=[[Paragraph(str(k),small),Paragraph(str(v),small)] for k,v in kpis]
            kt=Table(kpi_data,colWidths=[doc.width*.65,doc.width*.35]);kt.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),colors.HexColor('#f5f3ee')),('BOX',(0,0),(-1,-1),.4,colors.HexColor('#d8d2c3')),('INNERGRID',(0,0),(-1,-1),.2,colors.HexColor('#e4dfd3')),('FONTNAME',(1,0),(1,-1),'Helvetica-Bold'),('ALIGN',(1,0),(1,-1),'RIGHT'),('PADDING',(0,0),(-1,-1),7)]));story += [Paragraph('Executive summary',styles['Heading2']),kt,Spacer(1,8*mm)]
    else:
        story += [Paragraph(tenant.name,styles['Heading2']),Paragraph(title,styles['Title']),Paragraph(f'{report_id} · Reporting period: {_period_label(filters)} · Generated: {timezone.now().isoformat()} · Classification: {classification}',meta),Spacer(1,5*mm)]
        if kpis:
            kpi_text=' · '.join(f'<b>{k}</b>: {v}' for k,v in kpis[:8]);story += [KeepTogether([Paragraph(kpi_text,meta),Spacer(1,4*mm)])]
    if not rows:
        story.append(Paragraph('No data for selected filters',styles['Heading3']))
    else:
        table_data=[[Paragraph(str(c),small) for c in columns]]+[[Paragraph(_safe_text(v) or '—',small) for v in row] for row in rows]
        widths=[doc.width/max(1,len(columns))]*len(columns);table=Table(table_data,colWidths=widths,repeatRows=1,hAlign='LEFT')
        table.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#102a43')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('VALIGN',(0,0),(-1,-1),'TOP'),('GRID',(0,0),(-1,-1),.25,colors.HexColor('#d4d9dd')),('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#f6f7f7')]),('LEFTPADDING',(0,0),(-1,-1),4),('RIGHTPADDING',(0,0),(-1,-1),4),('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4)]));story.append(table)
    def header_footer(canv,doc_):
        canv.saveState();canv.setStrokeColor(colors.HexColor('#c7a64a'));canv.line(15*mm,pagesize[1]-14*mm,pagesize[0]-15*mm,pagesize[1]-14*mm);canv.setFont('Helvetica-Bold',7);canv.setFillColor(colors.HexColor('#102a43'));canv.drawString(15*mm,pagesize[1]-11*mm,f'{tenant.name} · {report_id} · {title}');canv.setFont('Helvetica',7);canv.setFillColor(colors.HexColor('#59636b'));canv.drawString(15*mm,8*mm,f'{classification} · Generated {timezone.now():%Y-%m-%d %H:%M %Z} · {_period_label(filters)}');canv.restoreState()
    doc.build(story,onFirstPage=header_footer,onLaterPages=header_footer,canvasmaker=NumberedCanvas);return buf.getvalue()


def _csv_bytes(tenant, report_id, definition, columns, rows, filters):
    out=io.StringIO();writer=csv.writer(out,lineterminator='\n');writer.writerow(['Report ID',report_id]);writer.writerow(['Report',definition['name']]);writer.writerow(['Tenant',tenant.name]);writer.writerow(['Reporting period',_period_label(filters)]);writer.writerow(['Filters',str(filters)]);writer.writerow([]);writer.writerow(columns);writer.writerows([[_safe_text(v) for v in row] for row in rows]);return out.getvalue().encode('utf-8-sig')


def _xlsx_bytes(tenant, report_id, definition, columns, rows, filters, kpis):
    wb=Workbook();summary=wb.active;summary.title='Summary';summary.append(['Quill of the Craft',definition['name']]);summary.append(['Report ID',report_id]);summary.append(['Tenant',tenant.name]);summary.append(['Reporting period',_period_label(filters)]);summary.append(['Generated',timezone.now().isoformat()]);summary.append(['Filters',str(filters)])
    if kpis:
        summary.append([]);summary.append(['KPI','Value'])
        for k,v in kpis:summary.append([_safe_text(k),_safe_text(v)])
    detail=wb.create_sheet('Detail');detail.append(columns)
    for row in rows:detail.append([_safe_text(v) for v in row])
    detail.freeze_panes='A2';detail.auto_filter.ref=f'A1:{get_column_letter(max(1,len(columns)))}{max(1,len(rows)+1)}'
    header_fill=PatternFill('solid',fgColor='102A43');header_font=Font(color='FFFFFF',bold=True)
    for c in detail[1]:c.fill=header_fill;c.font=header_font;c.alignment=Alignment(vertical='center')
    for ws in (summary,detail):
        for col in ws.columns:
            letter=get_column_letter(col[0].column);ws.column_dimensions[letter].width=min(48,max(12,max(len(str(c.value or '')) for c in col)+2))
    buf=io.BytesIO();wb.save(buf);return buf.getvalue()


def render_report_bytes(tenant, report_type, fmt='pdf', filters=None):
    fmt=str(fmt).lower();filters=filters or {};report_id=canonical_report_id(report_type);definition=REPORT_CATALOG[report_id]
    if fmt not in FORMATS or fmt not in definition['formats']:raise ValueError('unsupported_report_format')
    data=_report_data(tenant,report_id,filters);columns,rows=data['columns'],data['rows']
    if fmt=='pdf':blob=_pdf_bytes(tenant,report_id,definition,columns,rows,filters,data['kpis']);mime='application/pdf'
    elif fmt=='csv':blob=_csv_bytes(tenant,report_id,definition,columns,rows,filters);mime='text/csv'
    else:blob=_xlsx_bytes(tenant,report_id,definition,columns,rows,filters,data['kpis']);mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return blob,mime,fmt,len(rows)


def render_platform_report_bytes(report_type, fmt='pdf', filters=None):
    fmt=str(fmt).lower();filters=filters or {};report_id=canonical_report_id(report_type);definition=REPORT_CATALOG[report_id]
    if not definition.get('platform'):raise ValueError('tenant_report_requires_tenant_context')
    if fmt not in FORMATS or fmt not in definition['formats']:raise ValueError('unsupported_report_format')
    data=_platform_report_data(report_id,filters);columns,rows=data['columns'],data['rows'];context=PLATFORM_REPORT_CONTEXT
    if fmt=='pdf':blob=_pdf_bytes(context,report_id,definition,columns,rows,filters,data['kpis']);mime='application/pdf'
    elif fmt=='csv':blob=_csv_bytes(context,report_id,definition,columns,rows,filters);mime='text/csv'
    else:blob=_xlsx_bytes(context,report_id,definition,columns,rows,filters,data['kpis']);mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return blob,mime,fmt,len(rows)


def safe_report_filename(tenant, report_type, fmt, when=None):
    report_id=canonical_report_id(report_type).lower();stamp=(when or timezone.now()).strftime('%Y%m%d');slug=re.sub(r'[^a-z0-9-]+','-',str(tenant.slug).lower()).strip('-')[:60] or 'tenant';return f'{slug}-{report_id}-{stamp}.{fmt}'


def generate_report(tenant, report_type, fmt='pdf', filters=None, user=None, existing=None):
    report_id=canonical_report_id(report_type);data,mime,ext,row_count=render_report_bytes(tenant,report_id,fmt,filters);digest=hashlib.sha256(data).hexdigest();filename=safe_report_filename(tenant,report_id,ext);name=str(PurePosixPath(f'tenants/{tenant.id}/generated/reports/{timezone.now():%Y/%m}/{digest[:12]}-{filename}'));path=default_storage.save(name,ContentFile(data));meta={'mime_type':mime,'size_bytes':len(data),'report_id':report_id,'report_name':REPORT_CATALOG[report_id]['name'],'filename':filename,'classification':str(((tenant.settings or {}).get('reports') or {}).get('classification','Internal'))[:40],'data_as_of':timezone.now().isoformat()}
    if existing:
        obj=existing;obj.report_type=report_id;obj.format=fmt;obj.filters=filters or {};obj.reporting_period=_period_label(filters);obj.file_path=path;obj.sha256=digest;obj.row_count=row_count;obj.generated_by=user or obj.generated_by;obj.metadata=meta;obj.status='completed';obj.progress=100;obj.error='';obj.save()
    else:
        obj=GeneratedReport.objects.create(tenant=tenant,report_type=report_id,format=fmt,filters=filters or {},reporting_period=_period_label(filters),file_path=path,sha256=digest,row_count=row_count,generated_by=user,metadata=meta,status='completed',progress=100)
    return obj,data,mime
