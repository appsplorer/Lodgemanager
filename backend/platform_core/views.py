from __future__ import annotations
import csv, hashlib, io, json, secrets, uuid
from decimal import Decimal
from datetime import date, timedelta
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.tokens import default_token_generator
from django.contrib.auth.password_validation import validate_password
from django.core.mail import send_mail
from django.core.files.storage import default_storage
from django.core.exceptions import ValidationError
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes
from django.conf import settings as django_settings
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from .services.platform_permissions import platform_permission_required, user_has_platform_permission
from django.db import transaction
from django.db.models import Sum, Count, Q
from django.http import JsonResponse, HttpResponse, FileResponse
from django.middleware.csrf import get_token
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from .models import *
from .services.audit import append_audit, verify_audit_chain
from .services.permissions import permission_required
from .services.exports import export_queryset
from .services.encryption import encrypt_json, decrypt_json
from .domain.permissions import has_permission, ensure_grantable_permissions, effective_custom_permissions
from .domain.report_policy import platform_permission_for_report
from .domain.retention import normalize_retention_policy
from .domain.workflows import transition_minutes, transition_candidate_pipeline
from .domain.finance import assessment_status, amount_due, receipt_number
from .domain.tenant_config import normalize_member_custom_fields, normalize_candidate_stages
from .domain.public_config import normalize_analytics_config
from .domain.analytics import pct, lodge_energy_score, candidate_conversion_funnel
from .domain.ab_testing import assign_variant
from .domain.data_quality import completeness, likely_duplicate
from .services.payments import StripeAdapter, PayMongoAdapter, XenditAdapter
from .services.rate_limit import allow as rate_limit_allow
from .domain.totp import generate_secret, verify_totp
from .validation import validate_security_settings, validate_redirect_rule, validate_report_recipients, validate_cms_blocks, validate_navigation_items, validate_lead_form_fields, validate_lead_submission
from .services.communications import campaign_members
from .services.security import has_recent_step_up, security_alert
from .services.reports import generate_report, report_rows, report_catalog, report_preview, platform_report_preview, render_platform_report_bytes, estimate_report_rows, canonical_report_id, safe_report_filename, PLATFORM_REPORT_CONTEXT, REPORT_CATALOG, REPORT_TYPES, FORMATS
from .services.meetings import parse_rrule, sync_meeting_agenda_json, generate_recurring_occurrences
from .services.account_tokens import AccountTokenError, consume_account_token, issue_account_token
from .services.safe_urls import UnsafeURLError, validate_public_url
from .services.member_operations import bulk_fingerprint, duplicate_reasons
from .domain.media_assets import build_media_metadata, public_download_allowed
from .domain.job_policy import retryable_job

JSON_FIELDS={'tags','custom_data','agenda','business_items','communications','approved_bills','segment','permissions','meeting_schedule','signature_blocks','custom_fields','retention_policy','blocks','og','schema','config','merge_fields','recipients','filters','settings'}
DATE_FIELDS={'joined_on','initiated_on','passed_on','raised_on','date_of_birth','starts_on','ends_on','target_degree_date','proficiency_date','due_on','period_start','period_end','incurred_on'}
DATETIME_FIELDS={'starts_at','ends_at','scheduled_for'}

def _body(request):
    if request.content_type and 'application/json' in request.content_type:
        try: return json.loads(request.body or b'{}')
        except json.JSONDecodeError: raise ValueError('Invalid JSON')
    return request.POST.dict()

def _serialize(obj, fields=None):
    fields=fields or [f.name for f in obj._meta.fields if f.name not in ('ciphertext',)]
    out={}
    for name in fields:
        val=getattr(obj,name,None)
        if hasattr(val,'pk'): val=val.pk
        if isinstance(val,(Decimal,)): val=str(val)
        elif hasattr(val,'isoformat'): val=val.isoformat()
        elif isinstance(val,uuid.UUID): val=str(val)
        out[name]=val
    return out

def _apply(obj,data,allowed,tenant=None):
    for key in allowed:
        if key not in data: continue
        field=obj._meta.get_field(key); val=data[key]
        if field.is_relation and (getattr(field,'many_to_one',False) or getattr(field,'one_to_one',False)):
            if val and tenant is not None:
                related=field.remote_field.model
                if any(f.name=='tenant' for f in related._meta.fields):
                    if not related.objects.filter(pk=val,tenant=tenant).exists():
                        raise ValueError(f'invalid_or_cross_tenant_relation:{key}')
            setattr(obj,f'{key}_id',val or None)
        else: setattr(obj,key,val)
    return obj

def _candidate_queryset_for_request(request):
    qs=Candidate.objects.filter(tenant=request.tenant)
    if getattr(request,'tenant_membership',None) and request.tenant_membership.role=='mentor':
        # Mentor/Education users are object-scoped to candidates assigned to the member record
        # associated with their login email.  A missing linkage safely yields no candidates.
        email=(request.user.email or '').strip()
        if not email:
            return qs.none()
        mentor_ids=Member.objects.filter(tenant=request.tenant,email__iexact=email).values_list('id',flat=True)
        return qs.filter(Q(mentor_id__in=mentor_ids)|Q(mentor_assignments__mentor_id__in=mentor_ids,mentor_assignments__is_active=True)).distinct()
    return qs

def _candidate_for_request(request,pk):
    return _candidate_queryset_for_request(request).filter(pk=pk).first()

def _offline_version_conflict(request,obj):
    base=(request.headers.get('X-LodgeFlow-Base-Updated-At') or '').strip()
    current=getattr(obj,'updated_at',None)
    if not base or current is None:
        return None
    current_value=current.isoformat()
    if base!=current_value:
        return JsonResponse({'error':'offline_conflict','detail':'This record changed on the server after the offline copy was created. Refresh and review before applying your edit.','base_updated_at':base,'server_updated_at':current_value},status=409)
    return None

def _member_duplicate_matches(tenant,candidate,exclude_id=None,limit=10):
    qs=Member.objects.filter(tenant=tenant)
    if exclude_id:qs=qs.exclude(pk=exclude_id)
    candidates=qs.filter(Q(email__iexact=getattr(candidate,'email',''))|Q(phone=getattr(candidate,'phone',''))|Q(first_name__iexact=getattr(candidate,'first_name',''),last_name__iexact=getattr(candidate,'last_name',''))).order_by('last_name','first_name')[:100]
    matches=[]
    for existing in candidates:
        reasons=duplicate_reasons(candidate,existing)
        if reasons:matches.append({'id':str(existing.id),'name':f'{existing.first_name} {existing.last_name}'.strip(),'email':existing.email,'reasons':reasons})
        if len(matches)>=limit:break
    return matches

def _crud_list_create(request,model,allowed,permission_module,filters=None,order='-created_at'):
    if request.method=='GET':
        if not has_permission(request.tenant_membership.role,f'{permission_module}.view',effective_custom_permissions(request.tenant_membership)): return JsonResponse({'error':'forbidden'},status=403)
        qs=model.objects.filter(tenant=request.tenant)
        for key in filters or []:
            if request.GET.get(key): qs=qs.filter(**{key:request.GET[key]})
        q=request.GET.get('q')
        if q and model is Member: qs=qs.filter(Q(first_name__icontains=q)|Q(last_name__icontains=q)|Q(email__icontains=q)|Q(phone__icontains=q))
        try: limit=max(1,min(int(request.GET.get('limit',100)),500)); offset=max(int(request.GET.get('offset',0)),0)
        except (TypeError,ValueError): return JsonResponse({'error':'invalid_pagination'},status=400)
        return JsonResponse({'count':qs.count(),'results':[_serialize(x) for x in qs.order_by(order)[offset:offset+limit]]})
    if not has_permission(request.tenant_membership.role,f'{permission_module}.create',effective_custom_permissions(request.tenant_membership)): return JsonResponse({'error':'forbidden'},status=403)
    try: data=_body(request)
    except ValueError as e: return JsonResponse({'error':str(e)},status=400)
    obj=model(tenant=request.tenant)
    try:
        _apply(obj,data,allowed,request.tenant); obj.full_clean()
        duplicate_override=''
        if model is Member:
            matches=_member_duplicate_matches(request.tenant,obj)
            duplicate_override=str(data.get('duplicate_override_reason') or '').strip()
            if matches and len(duplicate_override)<5:return JsonResponse({'error':'possible_duplicate','matches':matches,'override_required':True},status=409)
        obj.save()
    except (ValueError,TypeError,ValidationError) as e: return JsonResponse({'error':'validation_error','detail':str(e)},status=400)
    append_audit(request=request,action=f'{permission_module}.created',object_type=model.__name__,object_id=obj.pk,after=_serialize(obj),metadata=({'duplicate_override_reason':duplicate_override} if model is Member and duplicate_override else {}))
    return JsonResponse(_serialize(obj),status=201)

def _crud_detail(request,model,pk,allowed,permission_module):
    try: obj=model.objects.get(pk=pk,tenant=request.tenant)
    except model.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
    action={'GET':'view','PATCH':'edit','PUT':'edit','DELETE':'delete'}[request.method]
    if not has_permission(request.tenant_membership.role,f'{permission_module}.{action}',effective_custom_permissions(request.tenant_membership)): return JsonResponse({'error':'forbidden'},status=403)
    if request.method=='GET': return JsonResponse(_serialize(obj))
    conflict=_offline_version_conflict(request,obj)
    if conflict: return conflict
    if request.method=='DELETE':
        before=_serialize(obj); obj.delete(); append_audit(request=request,action=f'{permission_module}.deleted',object_type=model.__name__,object_id=pk,before=before); return HttpResponse(status=204)
    before=_serialize(obj)
    try: data=_body(request)
    except ValueError as e: return JsonResponse({'error':str(e)},status=400)
    try:
        _apply(obj,data,allowed,request.tenant); obj.full_clean()
        duplicate_override=''
        if model is Member:
            matches=_member_duplicate_matches(request.tenant,obj,exclude_id=obj.id)
            duplicate_override=str(data.get('duplicate_override_reason') or '').strip()
            if matches and len(duplicate_override)<5:return JsonResponse({'error':'possible_duplicate','matches':matches,'override_required':True},status=409)
        obj.save()
    except (ValueError,TypeError,ValidationError) as e: return JsonResponse({'error':'validation_error','detail':str(e)},status=400)
    append_audit(request=request,action=f'{permission_module}.updated',object_type=model.__name__,object_id=obj.pk,before=before,after=_serialize(obj),metadata=({'duplicate_override_reason':duplicate_override} if model is Member and duplicate_override else {}))
    return JsonResponse(_serialize(obj))

@require_http_methods(['GET'])
def health(request):
    return JsonResponse({'status':'ok','service':'lodgeflow-api','time':timezone.now().isoformat()})

@require_http_methods(['GET'])
def readiness(request):
    checks={'database':False,'cache':False,'storage':False,'malware_scanner':False}
    try:
        from django.db import connection
        with connection.cursor() as cur: cur.execute('SELECT 1'); cur.fetchone()
        checks['database']=True
    except Exception: pass
    try:
        from django.core.cache import cache
        token=str(uuid.uuid4()); cache.set('lf-readiness',token,10); checks['cache']=cache.get('lf-readiness')==token
    except Exception: pass
    try:
        from django.core.files.base import ContentFile
        from django.core.files.storage import default_storage
        probe=f'_health/readiness-{uuid.uuid4()}.txt'; stored=default_storage.save(probe,ContentFile(b'lodgeflow-ready'))
        try: checks['storage']=default_storage.exists(stored)
        finally:
            try: default_storage.delete(stored)
            except Exception: pass
    except Exception: pass
    try:
        from .services.file_security import clamav_readiness
        checks['malware_scanner']=clamav_readiness()
    except Exception: pass
    ok=all(checks.values())
    return JsonResponse({'status':'ready' if ok else 'not_ready','checks':checks,'time':timezone.now().isoformat()},status=200 if ok else 503)

@require_http_methods(['GET'])
def csrf(request): return JsonResponse({'csrfToken':get_token(request)})

@require_http_methods(['POST'])
def login_view(request):
    client_key=getattr(request,'ip_hash','') or request.META.get('REMOTE_ADDR','unknown')
    if not rate_limit_allow(f'login:{client_key}',limit=8,window=300,critical=True):
        return JsonResponse({'error':'too_many_attempts'},status=429)
    try: data=_body(request)
    except ValueError as e: return JsonResponse({'error':str(e)},status=400)
    identity=str(data.get('username') or data.get('email') or '').strip()
    username=identity
    if '@' in identity:
        matches=list(get_user_model().objects.filter(email__iexact=identity,is_active=True).values_list('username',flat=True)[:2])
        if len(matches)==1: username=matches[0]
    user=authenticate(request,username=username,password=data.get('password',''))
    if not user or not user.is_active:
        append_audit(request=request,action='auth.login_failed',metadata={'success':False})
        return JsonResponse({'error':'invalid_credentials'},status=401)
    memberships=TenantMembership.objects.filter(user=user,is_active=True,tenant__is_active=True).select_related('tenant','custom_role')
    mfa_required=memberships.filter(mfa_required=True).exists() or any(bool(((m.tenant.settings or {}).get('security') or {}).get('mfa_required')) for m in memberships) or user.platform_roles.filter(is_active=True).exists() or user.is_superuser
    confirmed=MFADevice.objects.filter(user=user,is_confirmed=True).exists()
    if mfa_required:
        request.session['pending_mfa_user_id']=user.id
        request.session['pending_mfa_authenticated_at']=timezone.now().isoformat()
        append_audit(request=request,actor=user,action='auth.password_verified',metadata={'mfa_required':True})
        return JsonResponse({'mfa_required':True,'mfa_enrollment_required':not confirmed})
    request.session.pop('mfa_verified_at',None); login(request,user); request.session['password_verified_at']=timezone.now().isoformat(); append_audit(request=request,actor=user,action='auth.login',metadata={'success':True,'mfa':False})
    return JsonResponse({'user':{'id':user.id,'username':user.get_username(),'email':user.email},'tenants':[{'id':str(m.tenant_id),'name':m.tenant.name,'role':m.role} for m in memberships]})

@require_http_methods(['POST'])
def mfa_setup(request):
    user_id=request.session.get('pending_mfa_user_id')
    if not user_id: return JsonResponse({'error':'password_verification_required'},status=401)
    authenticated_at=request.session.get('pending_mfa_authenticated_at')
    try: verified_at=timezone.datetime.fromisoformat(authenticated_at) if authenticated_at else None
    except (TypeError,ValueError): verified_at=None
    if not verified_at or timezone.now()-verified_at>timezone.timedelta(minutes=10):
        request.session.pop('pending_mfa_user_id',None); request.session.pop('pending_mfa_authenticated_at',None)
        return JsonResponse({'error':'password_verification_expired'},status=401)
    user=get_user_model().objects.get(pk=user_id)
    if MFADevice.objects.filter(user=user,is_confirmed=True).exists():
        return JsonResponse({'error':'mfa_already_enrolled','detail':'An administrator must reset MFA before a new authenticator can be enrolled.'},status=409)
    device=MFADevice.objects.filter(user=user,is_confirmed=False).order_by('-created_at').first()
    if device:
        secret=decrypt_json(device.secret_ciphertext)['secret']
    else:
        secret=generate_secret(); device=MFADevice.objects.create(user=user,name='Authenticator',secret_ciphertext=encrypt_json({'secret':secret}),is_confirmed=False)
    issuer='LodgeFlow'; label=f'{issuer}:{user.get_username()}'
    uri=f'otpauth://totp/{label}?secret={secret}&issuer={issuer}&algorithm=SHA1&digits=6&period=30'
    return JsonResponse({'device_id':str(device.id),'secret':secret,'otpauth_uri':uri})

@require_http_methods(['POST'])
def mfa_verify(request):
    user_id=request.session.get('pending_mfa_user_id')
    if not user_id: return JsonResponse({'error':'password_verification_required'},status=401)
    authenticated_at=request.session.get('pending_mfa_authenticated_at')
    try: verified_at=timezone.datetime.fromisoformat(authenticated_at) if authenticated_at else None
    except (TypeError,ValueError): verified_at=None
    if not verified_at or timezone.now()-verified_at>timezone.timedelta(minutes=10):
        request.session.pop('pending_mfa_user_id',None); request.session.pop('pending_mfa_authenticated_at',None)
        return JsonResponse({'error':'password_verification_expired'},status=401)
    if not rate_limit_allow(f'mfa:{user_id}',limit=10,window=300,critical=True): return JsonResponse({'error':'too_many_attempts'},status=429)
    user=get_user_model().objects.get(pk=user_id); data=_body(request); code=str(data.get('code',''))
    device_id=data.get('device_id'); devices=MFADevice.objects.filter(user=user)
    if device_id: devices=devices.filter(pk=device_id)
    device=devices.order_by('-is_confirmed','-created_at').first()
    if not device: return JsonResponse({'error':'mfa_device_not_found'},status=404)
    secret=decrypt_json(device.secret_ciphertext)['secret']; ok,counter=verify_totp(secret,code,last_counter=device.last_used_counter)
    if not ok: return JsonResponse({'error':'invalid_mfa_code'},status=401)
    device.last_used_counter=counter
    if not device.is_confirmed: device.is_confirmed=True; device.confirmed_at=timezone.now()
    device.save(update_fields=['last_used_counter','is_confirmed','confirmed_at','updated_at'])
    login(request,user); request.session['mfa_verified_at']=timezone.now().isoformat(); request.session['password_verified_at']=timezone.now().isoformat(); request.session.pop('pending_mfa_user_id',None); request.session.pop('pending_mfa_authenticated_at',None)
    append_audit(request=request,actor=user,action='auth.login',metadata={'success':True,'mfa':True})
    memberships=TenantMembership.objects.filter(user=user,is_active=True,tenant__is_active=True).select_related('tenant','custom_role')
    return JsonResponse({'user':{'id':user.id,'username':user.get_username(),'email':user.email},'tenants':[{'id':str(m.tenant_id),'name':m.tenant.name,'role':m.role} for m in memberships]})

@require_http_methods(['POST'])
@login_required
def step_up_auth(request):
    client_key=getattr(request,'ip_hash','') or request.META.get('REMOTE_ADDR','unknown')
    if not rate_limit_allow(f'step-up:{request.user.id}:{client_key}',limit=8,window=300,critical=True):
        return JsonResponse({'error':'too_many_attempts'},status=429)
    data=_body(request); password=str(data.get('password') or '')
    if not password or not request.user.check_password(password):
        append_audit(request=request,actor=request.user,action='auth.step_up_failed',metadata={'factor':'password'})
        return JsonResponse({'error':'invalid_credentials'},status=401)
    confirmed=MFADevice.objects.filter(user=request.user,is_confirmed=True).order_by('-created_at').first()
    if confirmed:
        code=str(data.get('code') or '')
        if not code:return JsonResponse({'error':'mfa_code_required'},status=400)
        secret=decrypt_json(confirmed.secret_ciphertext)['secret']; ok,counter=verify_totp(secret,code,last_counter=confirmed.last_used_counter)
        if not ok:
            append_audit(request=request,actor=request.user,action='auth.step_up_failed',metadata={'factor':'mfa'})
            return JsonResponse({'error':'invalid_mfa_code'},status=401)
        confirmed.last_used_counter=counter; confirmed.save(update_fields=['last_used_counter','updated_at'])
        request.session['mfa_verified_at']=timezone.now().isoformat()
    request.session['password_verified_at']=timezone.now().isoformat()
    append_audit(request=request,actor=request.user,action='auth.step_up',metadata={'mfa':bool(confirmed)})
    return JsonResponse({'ok':True,'mfa_used':bool(confirmed),'valid_for_minutes':10})

@require_http_methods(['POST'])
@login_required
def logout_view(request): append_audit(request=request,action='auth.logout'); logout(request); return JsonResponse({'ok':True})

@require_http_methods(['GET'])
@login_required
def me(request):
    memberships=TenantMembership.objects.filter(user=request.user,is_active=True,tenant__is_active=True).select_related('tenant','custom_role')
    return JsonResponse({'user':{'id':request.user.id,'username':request.user.get_username(),'email':request.user.email,'is_staff':request.user.is_staff,'is_superuser':request.user.is_superuser},'tenants':[{'id':str(m.tenant_id),'name':m.tenant.name,'role':m.role,'permissions':effective_custom_permissions(m)} for m in memberships]})

@require_http_methods(['GET'])
def workspace_bootstrap(request):
    """Runtime contract for tenant identity, effective theme, feature flags and shell controls."""
    if not getattr(request,'tenant',None) or not getattr(request,'tenant_membership',None):
        return JsonResponse({'error':'tenant_required'},status=400)
    default_theme={'preset':'heritage','brand_name':'LodgeFlow','tagline':'Private Lodge Office','logo_text':'LF','primary':'#0c2830','primary_alt':'#123841','accent':'#c49a54','success':'#1f6f5f','danger':'#a53c3c','surface':'#ffffff','background':'#f6f8f7','text':'#142329','muted':'#66757b','line':'#dde4e5','radius':18,'density':'comfortable','sidebar':'dark'}
    global_setting=SiteSetting.objects.filter(key='workspace_theme').first(); global_theme=(global_setting.value or {}) if global_setting else {}
    tenant_settings=dict(request.tenant.settings or {}); branding=dict(tenant_settings.get('branding') or {}); theme={**default_theme,**global_theme,**branding}
    flags={}; global_flags={x.key:x for x in FeatureFlag.objects.filter(tenant=None)}; tenant_flags={x.key:x for x in FeatureFlag.objects.filter(tenant=request.tenant)}
    for key in sorted(set(global_flags)|set(tenant_flags)):
        f=tenant_flags.get(key) or global_flags.get(key); flags[key]={'enabled':bool(f.enabled),'config':f.config or {}}
    memberships=TenantMembership.objects.filter(user=request.user,is_active=True,tenant__is_active=True).select_related('tenant','custom_role'); profile,_=LodgeProfile.objects.get_or_create(tenant=request.tenant)
    now=timezone.now(); announcements=PlatformAnnouncement.objects.filter(is_active=True).filter(Q(starts_at=None)|Q(starts_at__lte=now)).filter(Q(ends_at=None)|Q(ends_at__gte=now)).order_by('-created_at')[:3]
    return JsonResponse({'tenant':{'id':str(request.tenant.id),'name':request.tenant.name,'slug':request.tenant.slug,'lodge_number':request.tenant.lodge_number,'jurisdiction':request.tenant.jurisdiction,'plan':request.tenant.plan,'locale':request.tenant.locale,'timezone':request.tenant.timezone},'profile':{'address':profile.address,'meeting_schedule':profile.meeting_schedule},'user':{'id':request.user.id,'username':request.user.get_username(),'email':request.user.email,'role':request.tenant_membership.role,'is_staff':request.user.is_staff,'is_superuser':request.user.is_superuser},'tenants':[{'id':str(m.tenant_id),'name':m.tenant.name,'lodge_number':m.tenant.lodge_number,'role':m.role} for m in memberships],'theme':theme,'features':flags,'announcements':[_serialize(x) for x in announcements],'capabilities':{'can_admin_settings':has_permission(request.tenant_membership.role,'settings.admin',effective_custom_permissions(request.tenant_membership)),'can_export':has_permission(request.tenant_membership.role,'reports.export',effective_custom_permissions(request.tenant_membership))},'impersonation':({'active':True,'id':str(request.impersonation.id),'actor':request.impersonation.actor.get_username(),'target':request.user.get_username(),'expires_at':request.impersonation.expires_at.isoformat()} if getattr(request,'impersonation',None) else {'active':False}),'system':{'status':'operational','generated_at':timezone.now().isoformat()}})

@require_http_methods(['GET'])
@permission_required('dashboard.view')
def dashboard(request):
    members=Member.objects.filter(tenant=request.tenant); active=members.filter(status='active').count(); total=members.count()
    meetings=Meeting.objects.filter(tenant=request.tenant); past=meetings.filter(starts_at__lte=timezone.now()).order_by('-starts_at')[:12]
    att_counts=[m.attendance.filter(status='present').count() for m in past]
    avg_att=round(sum(att_counts)/len(att_counts),1) if att_counts else 0
    assessments=DuesAssessment.objects.filter(tenant=request.tenant); billed=sum((a.amount+a.late_fee-a.waived_amount for a in assessments),Decimal('0')); paid=sum((a.paid_amount for a in assessments),Decimal('0'))
    dues_rate=pct(float(paid),float(billed)) if billed else 100.0
    candidates=Candidate.objects.filter(tenant=request.tenant,status='active'); stalled=sum(1 for c in candidates if c.target_degree_date and c.target_degree_date < date.today())
    current_officers=OfficerTerm.objects.filter(tenant=request.tenant,is_current=True); with_successor=current_officers.exclude(successor=None).count(); officer_eng=pct(with_successor,current_officers.count()) if current_officers.exists() else 100.0
    attendance_rate=pct(avg_att,max(active,1)); candidate_momentum=max(0,100-pct(stalled,max(candidates.count(),1)))
    energy=lodge_energy_score(min(attendance_rate,100),min(dues_rate,100),min(candidate_momentum,100),min(officer_eng,100))
    outstanding=sum((amount_due(a.amount,a.paid_amount,a.waived_amount,a.late_fee,a.adjustment_amount) for a in assessments),Decimal('0'))
    recent=[_serialize(x,['occurred_at','action','object_type','object_id']) for x in AuditEvent.objects.filter(tenant=request.tenant)[:12]]
    trend=[{'meeting':m.title,'date':m.starts_at.date().isoformat(),'present':m.attendance.filter(status='present').count()} for m in reversed(list(past[:6]))]
    return JsonResponse({'kpis':{'members_total':total,'members_active':active,'average_attendance':avg_att,'dues_collection_rate':dues_rate,'outstanding_dues':str(outstanding),'candidates_active':candidates.count(),'candidates_overdue':stalled,'lodge_energy_score':energy},'attendance_trend':trend,'recent_activity':recent})

@require_http_methods(['GET'])
def global_search(request):
    q=str(request.GET.get('q') or '').strip()
    if len(q)<2:return JsonResponse({'error':'query_min_2_chars'},status=400)
    try: limit=max(1,min(int(request.GET.get('limit',10)),25))
    except ValueError:return JsonResponse({'error':'invalid_pagination'},status=400)
    results=[]
    role=request.tenant_membership.role; custom=effective_custom_permissions(request.tenant_membership)
    if has_permission(role,'members.view',custom):
        for x in Member.objects.filter(tenant=request.tenant).filter(Q(first_name__icontains=q)|Q(last_name__icontains=q)|Q(email__icontains=q)|Q(phone__icontains=q)).order_by('last_name')[:limit]:
            results.append({'type':'member','id':str(x.id),'title':f'{x.first_name} {x.last_name}','subtitle':x.email,'path':f'/members/{x.id}'})
    if has_permission(role,'candidates.view',custom):
        for x in _candidate_queryset_for_request(request).select_related('member').filter(Q(member__first_name__icontains=q)|Q(member__last_name__icontains=q)|Q(stage__icontains=q)).order_by('-updated_at')[:limit]:
            results.append({'type':'candidate','id':str(x.id),'title':f'{x.member.first_name} {x.member.last_name}','subtitle':x.stage,'path':f'/candidates/{x.id}'})
    if has_permission(role,'meetings.view',custom):
        for x in Meeting.objects.filter(tenant=request.tenant).filter(Q(title__icontains=q)|Q(location__icontains=q)).order_by('-starts_at')[:limit]:
            results.append({'type':'meeting','id':str(x.id),'title':x.title,'subtitle':x.starts_at.isoformat(),'path':f'/meetings/{x.id}'})
    if has_permission(role,'documents.view',custom):
        for x in Document.objects.filter(tenant=request.tenant).filter(Q(title__icontains=q)|Q(folder__icontains=q)|Q(content__icontains=q)).order_by('-updated_at')[:limit]:
            if _document_allowed(request,x):results.append({'type':'document','id':str(x.id),'title':x.title,'subtitle':x.folder,'path':f'/documents/{x.id}'})
    return JsonResponse({'query':q,'results':results[:limit*4]})

@require_http_methods(['GET'])
@permission_required('members.view')
def member_activity(request,pk):
    member=Member.objects.filter(pk=pk,tenant=request.tenant).first()
    if not member:return JsonResponse({'error':'not_found'},status=404)
    events=[]
    for x in AuditEvent.objects.filter(tenant=request.tenant,object_id=str(member.id)).order_by('-occurred_at')[:100]:events.append({'kind':'audit','at':x.occurred_at.isoformat(),'summary':x.action,'id':str(x.id)})
    for x in member.attendance.select_related('meeting').order_by('-meeting__starts_at')[:100]:events.append({'kind':'attendance','at':x.meeting.starts_at.isoformat(),'summary':f'{x.meeting.title}: {x.status}','id':str(x.id)})
    if has_permission(request.tenant_membership.role,'dues.view',effective_custom_permissions(request.tenant_membership)):
        for x in member.dues_assessments.select_related('cycle').order_by('-created_at')[:100]:events.append({'kind':'dues','at':x.created_at.isoformat(),'summary':f'{x.cycle.name}: {x.status} · {x.amount}','id':str(x.id)})
        for x in member.payments.order_by('-paid_at')[:100]:events.append({'kind':'payment','at':x.paid_at.isoformat(),'summary':f'{x.currency} {x.amount} · {x.receipt_number}','id':str(x.id)})
    if has_permission(request.tenant_membership.role,'communications.view',effective_custom_permissions(request.tenant_membership)):
        for x in DeliveryLog.objects.filter(tenant=request.tenant,member=member).select_related('campaign').order_by('-created_at')[:100]:events.append({'kind':'communication','at':x.created_at.isoformat(),'summary':f'{x.campaign.name}: {x.status}','id':str(x.id)})
    events.sort(key=lambda x:x['at'],reverse=True)
    return JsonResponse({'member_id':str(member.id),'results':events[:200]})

# Entity CRUD endpoints
MEMBER_FIELDS=['first_name','last_name','email','phone','address','status','degree','joined_on','initiated_on','passed_on','raised_on','date_of_birth','tags','notes','custom_data','communication_preferences','is_dues_exempt','custom_dues_amount']
OFFICER_FIELDS=['member','office','starts_on','ends_on','is_current','successor','notes']
CANDIDATE_FIELDS=['member','stage','stage_order','mentor','target_degree_date','proficiency_date','referred_by','status','stage_history']
MEETING_FIELDS=['title','meeting_type','starts_at','ends_at','location','agenda','business_items','communications','approved_bills','status','recurrence_rule','education_minutes','business_minutes','fellowship_minutes']
DUESCYCLE_FIELDS=['name','period_start','period_end','due_on','standard_amount','late_fee','currency','status','proration_enabled','is_open']
ASSESSMENT_FIELDS=['member','cycle','amount','late_fee','adjustment_amount','currency','waived_amount','paid_amount','status']
PAYMENT_FIELDS=['member','assessment','amount','currency','method','provider','provider_reference','receipt_number','paid_at','notes','is_reversed','reversed_at','reversal_reason']
EXPENSE_FIELDS=['meeting','category','category_definition','description','amount','currency','incurred_on','vendor','receipt_path','attachment_document','is_charity']
DOCUMENT_FIELDS=['title','folder','tags','content','permissions','template_key']
CAMPAIGN_FIELDS=['name','channel','subject','body','segment','template_key','administrative_notice','scheduled_for']

def members(request): return _crud_list_create(request,Member,MEMBER_FIELDS,'members',['status','degree'])
def member_detail(request,pk): return _crud_detail(request,Member,pk,MEMBER_FIELDS,'members')
def _validate_officer_term(tenant,obj,exclude_id=None):
    if obj.ends_on and obj.ends_on<obj.starts_on: raise ValidationError({'ends_on':'End date cannot precede start date.'})
    definitions=OfficerRoleDefinition.objects.filter(tenant=tenant,is_active=True)
    if definitions.exists() and not definitions.filter(Q(key=obj.office)|Q(name=obj.office)).exists(): raise ValidationError({'office':'Choose an active tenant officer role.'})
    end=obj.ends_on or date.max
    qs=OfficerTerm.objects.filter(tenant=tenant,office=obj.office,member=obj.member)
    if exclude_id:qs=qs.exclude(pk=exclude_id)
    for other in qs:
        other_end=other.ends_on or date.max
        if obj.starts_on<=other_end and other.starts_on<=end: raise ValidationError({'starts_on':'This member already has an overlapping term for the same office.'})

@require_http_methods(['GET','POST'])
def officers(request):
    if request.method=='GET':return _crud_list_create(request,OfficerTerm,OFFICER_FIELDS,'officers',['office','is_current'])
    if not has_permission(request.tenant_membership.role,'officers.create',effective_custom_permissions(request.tenant_membership)) and not has_permission(request.tenant_membership.role,'officers.edit',effective_custom_permissions(request.tenant_membership)):return JsonResponse({'error':'forbidden'},status=403)
    data=_body(request);obj=OfficerTerm(tenant=request.tenant)
    try:_apply(obj,data,OFFICER_FIELDS,request.tenant);_validate_officer_term(request.tenant,obj);obj.full_clean();obj.save()
    except (ValueError,TypeError,ValidationError) as exc:return JsonResponse({'error':'validation_error','detail':getattr(exc,'message_dict',str(exc))},status=400)
    append_audit(request=request,action='officers.created',object_type='OfficerTerm',object_id=obj.id,after=_serialize(obj));return JsonResponse(_serialize(obj),status=201)

@require_http_methods(['GET','PATCH','PUT','DELETE'])
def officer_detail(request,pk):
    obj=OfficerTerm.objects.filter(pk=pk,tenant=request.tenant).first()
    if not obj:return JsonResponse({'error':'not_found'},status=404)
    action={'GET':'view','PATCH':'edit','PUT':'edit','DELETE':'delete'}[request.method]
    if not has_permission(request.tenant_membership.role,f'officers.{action}',effective_custom_permissions(request.tenant_membership)):return JsonResponse({'error':'forbidden'},status=403)
    if request.method=='GET':return JsonResponse(_serialize(obj))
    if request.method=='DELETE':before=_serialize(obj);obj.delete();append_audit(request=request,action='officers.deleted',object_type='OfficerTerm',object_id=pk,before=before);return HttpResponse(status=204)
    before=_serialize(obj);data=_body(request)
    try:_apply(obj,data,OFFICER_FIELDS,request.tenant);_validate_officer_term(request.tenant,obj,obj.id);obj.full_clean();obj.save()
    except (ValueError,TypeError,ValidationError) as exc:return JsonResponse({'error':'validation_error','detail':getattr(exc,'message_dict',str(exc))},status=400)
    append_audit(request=request,action='officers.updated',object_type='OfficerTerm',object_id=obj.id,before=before,after=_serialize(obj));return JsonResponse(_serialize(obj))

@require_http_methods(['GET','POST'])
@permission_required('officers.view')
def officer_roles(request):
    if request.method=='GET':return JsonResponse({'results':[_serialize(x) for x in OfficerRoleDefinition.objects.filter(tenant=request.tenant).order_by('sort_order','name')]})
    if not has_permission(request.tenant_membership.role,'officers.admin',effective_custom_permissions(request.tenant_membership)) and request.tenant_membership.role not in {'owner','admin'}:return JsonResponse({'error':'forbidden'},status=403)
    data=_body(request);key=str(data.get('key') or '').strip().lower().replace(' ','-');name=str(data.get('name') or '').strip()
    if not key or not name:return JsonResponse({'error':'key_and_name_required'},status=400)
    obj,_=OfficerRoleDefinition.objects.update_or_create(tenant=request.tenant,key=key,defaults={'name':name,'sort_order':int(data.get('sort_order',0)),'is_active':bool(data.get('is_active',True)),'permissions':data.get('permissions') or []})
    append_audit(request=request,action='officers.role.saved',object_type='OfficerRoleDefinition',object_id=obj.id);return JsonResponse(_serialize(obj),status=201)
@require_http_methods(['GET','POST'])
def candidates(request):
    if request.method=='GET':
        if not has_permission(request.tenant_membership.role,'candidates.view',effective_custom_permissions(request.tenant_membership)): return JsonResponse({'error':'forbidden'},status=403)
        qs=_candidate_queryset_for_request(request).prefetch_related('mentor_assignments')
        for key in ['stage','status']:
            if request.GET.get(key): qs=qs.filter(**{key:request.GET.get(key)})
        try: limit=min(max(int(request.GET.get('limit','100')),1),500)
        except ValueError:return JsonResponse({'error':'invalid_pagination'},status=400)
        return JsonResponse({'results':[dict(_serialize(x),mentor_ids=[str(m.mentor_id) for m in x.mentor_assignments.all() if m.is_active]) for x in qs.order_by('-created_at')[:limit]]})
    if not has_permission(request.tenant_membership.role,'candidates.create',effective_custom_permissions(request.tenant_membership)): return JsonResponse({'error':'forbidden'},status=403)
    try: data=_body(request)
    except ValueError as exc: return JsonResponse({'error':str(exc)},status=400)
    stages=(request.tenant.settings or {}).get('candidate_stages') or ['enquiry','application','investigation','ballot','accepted','EA','FC','MM']; stage=str(data.get('stage') or stages[0])
    if stage not in stages: return JsonResponse({'error':'unknown_candidate_stage'},status=400)
    obj=Candidate(tenant=request.tenant,stage=stage,stage_order=stages.index(stage),stage_history=[{'from':None,'to':stage,'at':timezone.now().isoformat(),'by':request.user.id}])
    try:
        _apply(obj,data,['member','mentor','target_degree_date','proficiency_date','referred_by','status'],request.tenant); obj.full_clean(); obj.save()
        mentor_ids=list(data.get('mentor_ids') or ([] if not obj.mentor_id else [str(obj.mentor_id)]))
        for mentor_id in mentor_ids:
            mentor=Member.objects.filter(pk=mentor_id,tenant=request.tenant).first()
            if not mentor: raise ValueError('invalid_or_cross_tenant_mentor')
            CandidateMentorAssignment.objects.get_or_create(tenant=request.tenant,candidate=obj,mentor=mentor,role='mentor')
    except (ValueError,TypeError,ValidationError) as exc:
        if obj.pk: obj.delete()
        return JsonResponse({'error':'validation_error','detail':str(exc)},status=400)
    append_audit(request=request,action='candidates.created',object_type='Candidate',object_id=obj.pk,after=_serialize(obj)); return JsonResponse(dict(_serialize(obj),mentor_ids=mentor_ids),status=201)

@require_http_methods(['GET','PATCH','PUT','DELETE'])
def candidate_detail(request,pk):
    obj=_candidate_for_request(request,pk)
    if not obj: return JsonResponse({'error':'not_found'},status=404)
    action={'GET':'view','PATCH':'edit','PUT':'edit','DELETE':'delete'}[request.method]
    if not has_permission(request.tenant_membership.role,f'candidates.{action}',effective_custom_permissions(request.tenant_membership)): return JsonResponse({'error':'forbidden'},status=403)
    if request.method=='GET': return JsonResponse(dict(_serialize(obj),mentor_ids=list(obj.mentor_assignments.filter(is_active=True).values_list('mentor_id',flat=True))))
    conflict=_offline_version_conflict(request,obj)
    if conflict:return conflict
    if request.method=='DELETE':
        before=_serialize(obj); obj.delete(); append_audit(request=request,action='candidates.deleted',object_type='Candidate',object_id=pk,before=before); return HttpResponse(status=204)
    before=_serialize(obj)
    try:
        data=_body(request); _apply(obj,data,['mentor','target_degree_date','proficiency_date','referred_by','status'],request.tenant); obj.full_clean(); obj.save()
        if 'mentor_ids' in data:
            desired={str(x) for x in (data.get('mentor_ids') or [])}; members={str(x.id):x for x in Member.objects.filter(tenant=request.tenant,id__in=desired)}
            if set(members)!=desired: raise ValueError('invalid_or_cross_tenant_mentor')
            for assignment in obj.mentor_assignments.filter(is_active=True):
                if str(assignment.mentor_id) not in desired: assignment.is_active=False; assignment.ended_at=timezone.now(); assignment.save(update_fields=['is_active','ended_at','updated_at'])
            for mid,member in members.items(): CandidateMentorAssignment.objects.update_or_create(tenant=request.tenant,candidate=obj,mentor=member,role='mentor',defaults={'is_active':True,'ended_at':None})
    except (ValueError,TypeError,ValidationError) as exc:return JsonResponse({'error':'validation_error','detail':str(exc)},status=400)
    append_audit(request=request,action='candidates.updated',object_type='Candidate',object_id=obj.pk,before=before,after=_serialize(obj)); return JsonResponse(_serialize(obj))

@require_http_methods(['GET','POST'])
def meetings(request):
    if request.method=='GET':
        return _crud_list_create(request,Meeting,MEETING_FIELDS,'meetings',['meeting_type','status'])
    if not has_permission(request.tenant_membership.role,'meetings.create',effective_custom_permissions(request.tenant_membership)):
        return JsonResponse({'error':'forbidden'},status=403)
    try: data=_body(request)
    except ValueError as exc:return JsonResponse({'error':str(exc)},status=400)
    profile=LodgeProfile.objects.filter(tenant=request.tenant).first(); profile_defaults=(profile.meeting_schedule if profile else {}) or {}; tenant_defaults=((request.tenant.settings or {}).get('meeting_defaults') or {})
    merged={**tenant_defaults,**profile_defaults,**data}
    if not str(merged.get('title') or '').strip() or not merged.get('starts_at'):
        return JsonResponse({'error':'title_and_starts_at_required'},status=400)
    if merged.get('recurrence_rule'):
        try: parse_rrule(str(merged['recurrence_rule']))
        except ValueError as exc:return JsonResponse({'error':str(exc)},status=400)
    obj=Meeting(tenant=request.tenant)
    try:
        _apply(obj,merged,MEETING_FIELDS,request.tenant)
        if not obj.ends_at and merged.get('default_duration_minutes'):
            obj.ends_at=obj.starts_at+timedelta(minutes=max(1,min(int(merged['default_duration_minutes']),1440)))
        obj.full_clean();obj.save()
        agenda=data.get('agenda_items')
        if agenda is not None:
            if not isinstance(agenda,list) or len(agenda)>200:raise ValueError('invalid_agenda_items')
            for i,item in enumerate(agenda):
                if not isinstance(item,dict) or not str(item.get('title') or '').strip():raise ValueError('invalid_agenda_item')
                AgendaItem.objects.create(tenant=request.tenant,meeting=obj,kind=item.get('kind','custom'),standard_key=str(item.get('standard_key') or '')[:80],title=str(item['title'])[:240],notes=str(item.get('notes') or ''),sort_order=i,is_hidden=bool(item.get('is_hidden',False)))
            sync_meeting_agenda_json(obj)
    except (ValueError,TypeError,ValidationError) as exc:
        if obj.pk:obj.delete()
        return JsonResponse({'error':'validation_error','detail':str(exc)},status=400)
    append_audit(request=request,action='meetings.created',object_type='Meeting',object_id=obj.id,after=_serialize(obj),metadata={'defaults_applied':bool(profile_defaults or tenant_defaults)})
    return JsonResponse(_serialize(obj),status=201)

def meeting_detail(request,pk): return _crud_detail(request,Meeting,pk,MEETING_FIELDS,'meetings')

@require_http_methods(['GET','POST'])
@permission_required('meetings.view')
def meeting_agenda(request,meeting_id):
    meeting=Meeting.objects.filter(pk=meeting_id,tenant=request.tenant).first()
    if not meeting:return JsonResponse({'error':'not_found'},status=404)
    if request.method=='GET':return JsonResponse({'results':[_serialize(x) for x in meeting.agenda_items.order_by('sort_order','created_at')]})
    if not has_permission(request.tenant_membership.role,'meetings.edit',effective_custom_permissions(request.tenant_membership)):return JsonResponse({'error':'forbidden'},status=403)
    data=_body(request);title=str(data.get('title') or '').strip()
    if not title:return JsonResponse({'error':'title_required'},status=400)
    kind=str(data.get('kind') or 'custom');
    if kind not in dict(AgendaItem.KIND):return JsonResponse({'error':'invalid_agenda_kind'},status=400)
    order=data.get('sort_order',meeting.agenda_items.count())
    try:order=max(0,min(int(order),10000))
    except Exception:return JsonResponse({'error':'invalid_sort_order'},status=400)
    obj=AgendaItem.objects.create(tenant=request.tenant,meeting=meeting,kind=kind,standard_key=str(data.get('standard_key') or '')[:80],title=title[:240],notes=str(data.get('notes') or ''),sort_order=order,is_hidden=bool(data.get('is_hidden',False)))
    sync_meeting_agenda_json(meeting);append_audit(request=request,action='meetings.agenda.created',object_type='AgendaItem',object_id=obj.id,after=_serialize(obj));return JsonResponse(_serialize(obj),status=201)

@require_http_methods(['PATCH','DELETE'])
@permission_required('meetings.edit')
def agenda_item_detail(request,meeting_id,pk):
    obj=AgendaItem.objects.filter(pk=pk,meeting_id=meeting_id,tenant=request.tenant).select_related('meeting').first()
    if not obj:return JsonResponse({'error':'not_found'},status=404)
    before=_serialize(obj)
    if request.method=='DELETE':
        meeting=obj.meeting;obj.delete();sync_meeting_agenda_json(meeting);append_audit(request=request,action='meetings.agenda.deleted',object_type='AgendaItem',object_id=pk,before=before);return HttpResponse(status=204)
    data=_body(request)
    for key in ['title','notes','standard_key','is_hidden']:
        if key in data:setattr(obj,key,data[key])
    if 'kind' in data:
        if data['kind'] not in dict(AgendaItem.KIND):return JsonResponse({'error':'invalid_agenda_kind'},status=400)
        obj.kind=data['kind']
    if 'sort_order' in data:
        try:obj.sort_order=max(0,min(int(data['sort_order']),10000))
        except Exception:return JsonResponse({'error':'invalid_sort_order'},status=400)
    if not str(obj.title or '').strip():return JsonResponse({'error':'title_required'},status=400)
    obj.full_clean();obj.save();sync_meeting_agenda_json(obj.meeting);append_audit(request=request,action='meetings.agenda.updated',object_type='AgendaItem',object_id=obj.id,before=before,after=_serialize(obj));return JsonResponse(_serialize(obj))

@require_http_methods(['POST'])
@permission_required('meetings.edit')
def agenda_reorder(request,meeting_id):
    meeting=Meeting.objects.filter(pk=meeting_id,tenant=request.tenant).first()
    if not meeting:return JsonResponse({'error':'not_found'},status=404)
    ids=_body(request).get('ids')
    if not isinstance(ids,list) or len(ids)>200 or len(ids)!=len(set(map(str,ids))):return JsonResponse({'error':'invalid_agenda_order'},status=400)
    items=list(meeting.agenda_items.filter(id__in=ids))
    if len(items)!=meeting.agenda_items.count() or len(items)!=len(ids):return JsonResponse({'error':'agenda_order_must_include_all_items'},status=400)
    mapping={str(x.id):x for x in items}
    with transaction.atomic():
        for order,item_id in enumerate(ids):mapping[str(item_id)].sort_order=order;mapping[str(item_id)].save(update_fields=['sort_order','updated_at'])
        sync_meeting_agenda_json(meeting)
    append_audit(request=request,action='meetings.agenda.reordered',object_type='Meeting',object_id=meeting.id,metadata={'count':len(ids)});return JsonResponse({'results':[_serialize(mapping[str(x)]) for x in ids]})

@require_http_methods(['POST'])
@permission_required('meetings.edit')
def meeting_generate_occurrences(request,meeting_id):
    meeting=Meeting.objects.filter(pk=meeting_id,tenant=request.tenant).first()
    if not meeting:return JsonResponse({'error':'not_found'},status=404)
    if not meeting.recurrence_rule:return JsonResponse({'error':'recurrence_rule_required'},status=409)
    try: result=generate_recurring_occurrences(meeting,horizon_days=int(_body(request).get('horizon_days',180)))
    except (ValueError,TypeError) as exc:return JsonResponse({'error':str(exc)},status=400)
    append_audit(request=request,action='meetings.recurrence.generated',object_type='Meeting',object_id=meeting.id,metadata=result);return JsonResponse(result)
@require_http_methods(['GET','POST'])
def dues_cycles(request):
    if request.method=='GET':return _crud_list_create(request,DuesCycle,DUESCYCLE_FIELDS,'dues',['status','is_open'])
    if not has_permission(request.tenant_membership.role,'dues.admin',effective_custom_permissions(request.tenant_membership)):return JsonResponse({'error':'forbidden'},status=403)
    data=_body(request);obj=DuesCycle(tenant=request.tenant,status='draft',is_open=False)
    if not data.get('currency'):data['currency']=str((request.tenant.settings or {}).get('default_currency') or 'PHP').upper()
    try:
        _apply(obj,data,['name','period_start','period_end','due_on','standard_amount','late_fee','currency','proration_enabled'],request.tenant)
        obj.currency=str(obj.currency).upper();
        if len(obj.currency)!=3:raise ValueError('invalid_currency')
        if obj.period_start>obj.period_end or obj.due_on<obj.period_start:raise ValueError('invalid_cycle_dates')
        obj.full_clean();obj.save()
    except (ValueError,TypeError,ValidationError) as exc:return JsonResponse({'error':'validation_error','detail':str(exc)},status=400)
    append_audit(request=request,action='dues.cycle.created',object_type='DuesCycle',object_id=obj.id,after=_serialize(obj));return JsonResponse(_serialize(obj),status=201)

@require_http_methods(['GET','PATCH','DELETE'])
def dues_cycle_detail(request,pk):
    obj=DuesCycle.objects.filter(pk=pk,tenant=request.tenant).first()
    if not obj:return JsonResponse({'error':'not_found'},status=404)
    action={'GET':'view','PATCH':'admin','DELETE':'admin'}[request.method]
    if not has_permission(request.tenant_membership.role,f'dues.{action}',effective_custom_permissions(request.tenant_membership)):return JsonResponse({'error':'forbidden'},status=403)
    if request.method=='GET':return JsonResponse(_serialize(obj))
    if obj.assessments.exists():return JsonResponse({'error':'cycle_terms_locked_after_assessment_generation'},status=409)
    if request.method=='DELETE':
        if obj.status!='draft':return JsonResponse({'error':'only_draft_cycles_can_be_deleted'},status=409)
        before=_serialize(obj);obj.delete();append_audit(request=request,action='dues.cycle.deleted',object_type='DuesCycle',object_id=pk,before=before);return HttpResponse(status=204)
    if obj.status!='draft':return JsonResponse({'error':'only_draft_cycle_terms_can_be_edited'},status=409)
    before=_serialize(obj);data=_body(request)
    try:
        _apply(obj,data,['name','period_start','period_end','due_on','standard_amount','late_fee','currency','proration_enabled'],request.tenant);obj.currency=str(obj.currency).upper()
        if len(obj.currency)!=3 or obj.period_start>obj.period_end or obj.due_on<obj.period_start:raise ValueError('invalid_cycle_configuration')
        obj.full_clean();obj.save()
    except (ValueError,TypeError,ValidationError) as exc:return JsonResponse({'error':'validation_error','detail':str(exc)},status=400)
    append_audit(request=request,action='dues.cycle.updated',object_type='DuesCycle',object_id=obj.id,before=before,after=_serialize(obj));return JsonResponse(_serialize(obj))

@require_http_methods(['GET'])
@permission_required('dues.view')
def dues_cycle_preview(request,cycle_id):
    cycle=DuesCycle.objects.filter(pk=cycle_id,tenant=request.tenant).first()
    if not cycle:return JsonResponse({'error':'not_found'},status=404)
    from .domain.finance import prorated_dues
    rows=[];total=Decimal('0')
    for member in Member.objects.filter(tenant=request.tenant,status='active').order_by('last_name','first_name')[:50000]:
        base=Decimal('0') if member.is_dues_exempt else (member.custom_dues_amount if member.custom_dues_amount is not None else cycle.standard_amount);amount=prorated_dues(base,cycle.period_start,cycle.period_end,member.joined_on,cycle.proration_enabled);total+=amount;rows.append({'member_id':str(member.id),'name':f'{member.first_name} {member.last_name}','amount':str(amount),'exempt':member.is_dues_exempt})
    return JsonResponse({'cycle':_serialize(cycle),'population':len(rows),'total':str(total),'currency':cycle.currency,'preview':rows[:100]})

@require_http_methods(['POST'])
@permission_required('dues.admin')
def dues_cycle_transition(request,cycle_id):
    cycle=DuesCycle.objects.filter(pk=cycle_id,tenant=request.tenant).first()
    if not cycle:return JsonResponse({'error':'not_found'},status=404)
    target=str(_body(request).get('target') or '').lower();allowed={('draft','active'),('active','closed')}
    if (cycle.status,target) not in allowed:return JsonResponse({'error':'invalid_cycle_transition','from':cycle.status,'to':target},status=409)
    before=cycle.status;cycle.status=target;cycle.is_open=target=='active';cycle.save(update_fields=['status','is_open','updated_at']);append_audit(request=request,action='dues.cycle.transitioned',object_type='DuesCycle',object_id=cycle.id,before={'status':before},after={'status':target});return JsonResponse(_serialize(cycle))

@require_http_methods(['GET','POST'])
def dues_assessments(request):
    if request.method=='GET': return _crud_list_create(request,DuesAssessment,ASSESSMENT_FIELDS,'dues',['status','member_id','cycle_id'])
    if not has_permission(request.tenant_membership.role,'dues.create',effective_custom_permissions(request.tenant_membership)): return JsonResponse({'error':'forbidden'},status=403)
    try: data=_body(request)
    except ValueError as exc: return JsonResponse({'error':str(exc)},status=400)
    obj=DuesAssessment(tenant=request.tenant)
    try: _apply(obj,data,['member','cycle','amount'],request.tenant); obj.currency=obj.cycle.currency; obj.full_clean(); obj.save()
    except (ValueError,TypeError,ValidationError) as exc: return JsonResponse({'error':'validation_error','detail':str(exc)},status=400)
    append_audit(request=request,action='dues.created',object_type='DuesAssessment',object_id=obj.pk,after=_serialize(obj)); return JsonResponse(_serialize(obj),status=201)

@require_http_methods(['GET','PATCH','PUT','DELETE'])
def dues_assessment_detail(request,pk):
    try: obj=DuesAssessment.objects.select_related('cycle','member').get(pk=pk,tenant=request.tenant)
    except DuesAssessment.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
    action={'GET':'view','PATCH':'edit','PUT':'edit','DELETE':'delete'}[request.method]
    if not has_permission(request.tenant_membership.role,f'dues.{action}',effective_custom_permissions(request.tenant_membership)): return JsonResponse({'error':'forbidden'},status=403)
    if request.method=='GET': return JsonResponse(_serialize(obj))
    if obj.payments.exists() and request.method=='DELETE': return JsonResponse({'error':'assessment_with_payments_is_immutable'},status=409)
    if request.method=='DELETE':
        before=_serialize(obj); obj.delete(); append_audit(request=request,action='dues.deleted',object_type='DuesAssessment',object_id=pk,before=before); return HttpResponse(status=204)
    if obj.payments.exists(): return JsonResponse({'error':'assessment_financial_terms_locked_after_payment'},status=409)
    before=_serialize(obj)
    try: data=_body(request); _apply(obj,data,['member','cycle','amount'],request.tenant); obj.status=assessment_status(obj.amount,obj.paid_amount,obj.waived_amount,obj.late_fee,timezone.now().date()>obj.cycle.due_on,obj.adjustment_amount); obj.full_clean(); obj.save()
    except (ValueError,TypeError,ValidationError) as exc: return JsonResponse({'error':'validation_error','detail':str(exc)},status=400)
    append_audit(request=request,action='dues.updated',object_type='DuesAssessment',object_id=obj.pk,before=before,after=_serialize(obj)); return JsonResponse(_serialize(obj))
def payments(request):
    if request.method!='GET': return JsonResponse({'error':'use_record_payment_workflow'},status=405)
    return _crud_list_create(request,Payment,PAYMENT_FIELDS,'dues',['member_id','provider'])
def payment_detail(request,pk):
    if request.method!='GET': return JsonResponse({'error':'financial_records_are_immutable'},status=405)
    return _crud_detail(request,Payment,pk,PAYMENT_FIELDS,'dues')
@require_http_methods(['GET','POST'])
def expense_categories(request):
    if request.method=='GET':
        if not has_permission(request.tenant_membership.role,'finance.view',effective_custom_permissions(request.tenant_membership)):return JsonResponse({'error':'forbidden'},status=403)
        return JsonResponse({'results':[_serialize(x) for x in ExpenseCategory.objects.filter(tenant=request.tenant).order_by('name')]})
    if not has_permission(request.tenant_membership.role,'finance.edit',effective_custom_permissions(request.tenant_membership)):return JsonResponse({'error':'forbidden'},status=403)
    data=_body(request);name=str(data.get('name') or '').strip()
    if not name:return JsonResponse({'error':'category_name_required'},status=400)
    obj,created=ExpenseCategory.objects.update_or_create(tenant=request.tenant,name=name[:100],defaults={'is_active':bool(data.get('is_active',True)),'requires_approval':bool(data.get('requires_approval',False))});append_audit(request=request,action='finance.expense_category.saved',object_type='ExpenseCategory',object_id=obj.id);return JsonResponse(_serialize(obj),status=201 if created else 200)

@require_http_methods(['GET','POST'])
def expenses(request):
    if request.method=='GET':return _crud_list_create(request,Expense,EXPENSE_FIELDS,'finance',['category','is_charity','approval_status'])
    if not has_permission(request.tenant_membership.role,'finance.edit',effective_custom_permissions(request.tenant_membership)):return JsonResponse({'error':'forbidden'},status=403)
    data=_body(request);obj=Expense(tenant=request.tenant)
    try:
        _apply(obj,data,EXPENSE_FIELDS,request.tenant)
        if obj.category_definition_id:
            category=ExpenseCategory.objects.filter(pk=obj.category_definition_id,tenant=request.tenant,is_active=True).first()
            if not category:raise ValueError('invalid_or_inactive_expense_category')
            obj.category=category.name;obj.approval_status='pending' if category.requires_approval else 'draft'
        obj.amount=Decimal(str(obj.amount))
        if obj.amount<=0:raise ValueError('expense_amount_must_be_positive')
        obj.currency=str(obj.currency or (request.tenant.settings or {}).get('default_currency') or 'PHP').upper()
        if len(obj.currency)!=3:raise ValueError('invalid_currency')
        obj.full_clean();obj.save()
    except (ValueError,TypeError,ValidationError) as exc:return JsonResponse({'error':'validation_error','detail':str(exc)},status=400)
    append_audit(request=request,action='finance.expense.created',object_type='Expense',object_id=obj.id,after=_serialize(obj));return JsonResponse(_serialize(obj),status=201)

@require_http_methods(['GET','PATCH','PUT','DELETE'])
def expense_detail(request,pk):
    obj=Expense.objects.filter(pk=pk,tenant=request.tenant).first()
    if not obj:return JsonResponse({'error':'not_found'},status=404)
    action={'GET':'view','PATCH':'edit','PUT':'edit','DELETE':'delete'}[request.method]
    if not has_permission(request.tenant_membership.role,f'finance.{action}',effective_custom_permissions(request.tenant_membership)):return JsonResponse({'error':'forbidden'},status=403)
    if request.method=='GET':return JsonResponse(_serialize(obj))
    if obj.approval_status=='approved':return JsonResponse({'error':'approved_expense_is_immutable'},status=409)
    conflict=_offline_version_conflict(request,obj)
    if conflict:return conflict
    if request.method=='DELETE':
        before=_serialize(obj);obj.delete();append_audit(request=request,action='finance.expense.deleted',object_type='Expense',object_id=pk,before=before);return HttpResponse(status=204)
    before=_serialize(obj);data=_body(request)
    try:
        _apply(obj,data,EXPENSE_FIELDS,request.tenant)
        if obj.category_definition_id:
            category=ExpenseCategory.objects.filter(pk=obj.category_definition_id,tenant=request.tenant,is_active=True).first()
            if not category:raise ValueError('invalid_or_inactive_expense_category')
            obj.category=category.name;obj.approval_status='pending' if category.requires_approval else 'draft'
        obj.amount=Decimal(str(obj.amount))
        if obj.amount<=0:raise ValueError('expense_amount_must_be_positive')
        obj.currency=str(obj.currency).upper()
        if len(obj.currency)!=3:raise ValueError('invalid_currency')
        obj.approval_reason='';obj.approved_by=None;obj.approved_at=None;obj.full_clean();obj.save()
    except (ValueError,TypeError,ValidationError) as exc:return JsonResponse({'error':'validation_error','detail':str(exc)},status=400)
    append_audit(request=request,action='finance.expense.updated',object_type='Expense',object_id=obj.id,before=before,after=_serialize(obj));return JsonResponse(_serialize(obj))

@require_http_methods(['POST'])
@permission_required('finance.approve')
def expense_approval(request,pk):
    obj=Expense.objects.filter(pk=pk,tenant=request.tenant).first()
    if not obj:return JsonResponse({'error':'not_found'},status=404)
    if obj.approval_status not in {'draft','pending'}:return JsonResponse({'error':'expense_approval_already_decided','status':obj.approval_status},status=409)
    data=_body(request);decision=str(data.get('decision') or '').lower();reason=str(data.get('reason') or '').strip()
    if decision not in {'approved','rejected'}:return JsonResponse({'error':'invalid_decision'},status=400)
    if decision=='rejected' and len(reason)<5:return JsonResponse({'error':'rejection_reason_required'},status=400)
    before={'approval_status':obj.approval_status};obj.approval_status=decision;obj.approval_reason=reason[:2000];obj.approved_by=request.user;obj.approved_at=timezone.now();obj.save(update_fields=['approval_status','approval_reason','approved_by','approved_at','updated_at'])
    append_audit(request=request,action=f'finance.expense.{decision}',object_type='Expense',object_id=obj.id,before=before,after={'approval_status':decision},metadata={'reason':reason[:500]});return JsonResponse(_serialize(obj))

@require_http_methods(['GET','POST'])
def campaigns(request):
    if request.method=='GET': return _crud_list_create(request,CommunicationCampaign,CAMPAIGN_FIELDS,'communications',['status','channel'])
    if not has_permission(request.tenant_membership.role,'communications.create',effective_custom_permissions(request.tenant_membership)): return JsonResponse({'error':'forbidden'},status=403)
    try: data=_body(request)
    except ValueError as exc: return JsonResponse({'error':str(exc)},status=400)
    if str(data.get('channel','email')).lower()!='email': return JsonResponse({'error':'unsupported_channel'},status=400)
    obj=CommunicationCampaign(tenant=request.tenant,created_by=request.user,channel='email',status='draft')
    template_key=str(data.get('template_key') or '').strip()
    if template_key:
        template=EmailTemplate.objects.filter(tenant=request.tenant,key=template_key,is_active=True).first() or EmailTemplate.objects.filter(tenant=None,key=template_key,is_active=True).first()
        if not template: return JsonResponse({'error':'email_template_not_found'},status=404)
        data={**data,'subject':data.get('subject') or template.subject,'body':data.get('body') or template.body,'template_key':template_key}
    try:
        _apply(obj,data,['name','subject','body','segment','template_key','administrative_notice','scheduled_for'],request.tenant)
        from .services.template_rendering import validate_merge_fields,TemplateRenderError
        validate_merge_fields([],obj.subject,obj.body); obj.full_clean(); obj.save()
    except TemplateRenderError as exc: return JsonResponse({'error':exc.code,'fields':exc.fields},status=400)
    except (ValueError,TypeError,ValidationError) as exc: return JsonResponse({'error':'validation_error','detail':str(exc)},status=400)
    append_audit(request=request,action='communications.created',object_type='CommunicationCampaign',object_id=obj.pk,after=_serialize(obj)); return JsonResponse(_serialize(obj),status=201)

@require_http_methods(['GET','PATCH','PUT','DELETE'])
def campaign_detail(request,pk):
    try: obj=CommunicationCampaign.objects.get(pk=pk,tenant=request.tenant)
    except CommunicationCampaign.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
    action={'GET':'view','PATCH':'edit','PUT':'edit','DELETE':'delete'}[request.method]
    if not has_permission(request.tenant_membership.role,f'communications.{action}',effective_custom_permissions(request.tenant_membership)): return JsonResponse({'error':'forbidden'},status=403)
    if request.method=='GET': return JsonResponse(_serialize(obj))
    conflict=_offline_version_conflict(request,obj)
    if conflict: return conflict
    if obj.status not in ('draft','failed'): return JsonResponse({'error':'campaign_locked_after_queue','status':obj.status},status=409)
    if request.method=='DELETE':
        before=_serialize(obj); obj.delete(); append_audit(request=request,action='communications.deleted',object_type='CommunicationCampaign',object_id=pk,before=before); return HttpResponse(status=204)
    before=_serialize(obj)
    try:
        data=_body(request); _apply(obj,data,['name','subject','body','segment','template_key','administrative_notice','scheduled_for'],request.tenant); obj.channel='email'
        from .services.template_rendering import validate_merge_fields,TemplateRenderError
        validate_merge_fields([],obj.subject,obj.body); obj.full_clean(); obj.save()
    except TemplateRenderError as exc: return JsonResponse({'error':exc.code,'fields':exc.fields},status=400)
    except (ValueError,TypeError,ValidationError) as exc: return JsonResponse({'error':'validation_error','detail':str(exc)},status=400)
    append_audit(request=request,action='communications.updated',object_type='CommunicationCampaign',object_id=obj.pk,before=before,after=_serialize(obj)); return JsonResponse(_serialize(obj))

@require_http_methods(['POST'])
@permission_required('members.create')
def member_import(request):
    upload=request.FILES.get('file')
    if not upload: return JsonResponse({'error':'csv_or_xlsx_file_required'},status=400)
    if upload.size>10*1024*1024: return JsonResponse({'error':'file_too_large','max_mb':10},status=413)
    mapping={}
    try: mapping=json.loads(request.POST.get('mapping','{}') or '{}')
    except json.JSONDecodeError: return JsonResponse({'error':'invalid_mapping_json'},status=400)
    name=(upload.name or '').lower(); rows=[]
    if name.endswith('.xlsx'):
        try:
            from openpyxl import load_workbook
            wb=load_workbook(upload.file,read_only=True,data_only=True); ws=wb.active; iterator=ws.iter_rows(values_only=True); headers=[str(x or '').strip() for x in next(iterator)]
            for vals in iterator: rows.append({headers[i]:vals[i] for i in range(min(len(headers),len(vals)))})
        except Exception as exc: return JsonResponse({'error':'invalid_xlsx','detail':type(exc).__name__},status=400)
    elif name.endswith('.csv') or upload.content_type in ('text/csv','application/csv','text/plain'):
        try:
            text=io.TextIOWrapper(upload.file,encoding='utf-8-sig'); reader=csv.DictReader(text); headers=list(reader.fieldnames or []); rows=list(reader)
        except Exception as exc: return JsonResponse({'error':'invalid_csv','detail':type(exc).__name__},status=400)
    else: return JsonResponse({'error':'unsupported_import_type','supported':['csv','xlsx']},status=415)
    if len(rows)>10000:return JsonResponse({'error':'import_row_limit_exceeded','max_rows':10000},status=413)
    source_fields=set(headers); target_for_source={src:(mapping.get(src) or src) for src in source_fields}; mapped_targets=set(target_for_source.values()); required={'first_name','last_name'}
    if not required.issubset(mapped_targets): return JsonResponse({'error':'required_columns','columns':sorted(required),'available':sorted(source_fields)},status=400)
    validated_rows=[]; errors=[];duplicates=[];existing_rows=list(Member.objects.filter(tenant=request.tenant).values('id','first_name','last_name','email','phone'))
    for i,row in enumerate(rows,2):
        mapped={target_for_source[src]:value for src,value in row.items() if target_for_source.get(src) in MEMBER_FIELDS and value not in (None,'')}
        if not mapped.get('first_name') or not mapped.get('last_name'): errors.append({'row':i,'error':'missing_name'}); continue
        try:
            candidate=Member(tenant=request.tenant);_apply(candidate,mapped,MEMBER_FIELDS,request.tenant);candidate.full_clean()
        except Exception as exc:
            errors.append({'row':i,'error':str(exc)[:300]});continue
        row_matches=[]
        for existing in existing_rows:
            reasons=duplicate_reasons(mapped,existing)
            if reasons:row_matches.append({'member_id':str(existing['id']),'reasons':reasons})
        for prior_index,prior in validated_rows:
            reasons=duplicate_reasons(mapped,prior)
            if reasons:row_matches.append({'import_row':prior_index,'reasons':reasons})
        if row_matches:duplicates.append({'row':i,'matches':row_matches})
        validated_rows.append((i,mapped))
    preview=[row for _,row in validated_rows[:100]]
    if request.GET.get('commit')!='true': return JsonResponse({'mapping':target_for_source,'preview':preview,'valid_rows':len(validated_rows),'errors':errors,'duplicates':duplicates,'dry_run':True})
    duplicate_policy=str(request.POST.get('duplicate_policy') or '').strip().lower();override_reason=str(request.POST.get('duplicate_override_reason') or '').strip()
    if duplicates and duplicate_policy not in {'skip','create'}:return JsonResponse({'error':'duplicate_decision_required','duplicates':duplicates,'allowed_policies':['skip','create']},status=409)
    if duplicates and duplicate_policy=='create' and len(override_reason)<5:return JsonResponse({'error':'duplicate_override_reason_required'},status=400)
    duplicate_rows={item['row'] for item in duplicates}
    if len(validated_rows)>500:
        queued_rows=[{'row':index,'data':{key:(value.isoformat() if hasattr(value,'isoformat') else str(value) if isinstance(value,Decimal) else value) for key,value in row.items()}} for index,row in validated_rows if not (index in duplicate_rows and duplicate_policy=='skip')]
        job=SystemJob.objects.create(tenant=request.tenant,kind='member_import',status='queued',payload={'actor_id':request.user.id,'rows':queued_rows,'mapping':target_for_source,'duplicate_policy':duplicate_policy or 'none','duplicate_override_reason':override_reason},attempts=0)
        try:
            from .tasks import import_members_job
            task=import_members_job.delay(str(job.id));job.result={'task_id':str(task.id or '')};job.save(update_fields=['result','updated_at'])
        except Exception as exc:
            job.status='failed';job.finished_at=timezone.now();job.last_error=f'queue_unavailable:{type(exc).__name__}';job.save(update_fields=['status','finished_at','last_error','updated_at']);return JsonResponse({'error':'background_queue_unavailable','job_id':str(job.id)},status=503)
        append_audit(request=request,action='members.import_queued',object_type='SystemJob',object_id=job.id,metadata={'valid_rows':len(queued_rows),'duplicates':len(duplicates),'duplicate_policy':duplicate_policy or 'none','mapping':target_for_source});return JsonResponse({'queued':True,'job_id':str(job.id),'task_id':job.result['task_id'],'valid_rows':len(queued_rows),'errors':errors,'duplicates':duplicates},status=202)
    created=[]; validation_errors=[]
    with transaction.atomic():
        for index,row in validated_rows:
            if index in duplicate_rows and duplicate_policy=='skip':continue
            try:
                obj=Member(tenant=request.tenant); _apply(obj,row,MEMBER_FIELDS,request.tenant); obj.full_clean(); obj.save(); created.append(str(obj.id))
            except Exception as exc: validation_errors.append({'row':index,'error':str(exc)[:300]})
        if validation_errors: transaction.set_rollback(True)
    if validation_errors: return JsonResponse({'error':'import_validation_failed','created':0,'errors':errors+validation_errors},status=400)
    append_audit(request=request,action='members.imported',metadata={'count':len(created),'errors':len(errors),'duplicates':len(duplicates),'duplicate_policy':duplicate_policy or 'none','duplicate_override_reason':override_reason,'mapping':target_for_source,'format':'xlsx' if name.endswith('.xlsx') else 'csv'})
    return JsonResponse({'created':len(created),'errors':errors,'duplicates':duplicates,'skipped_duplicates':len(duplicate_rows) if duplicate_policy=='skip' else 0},status=201)

@require_http_methods(['GET'])
@permission_required('members.view')
def member_duplicates(request):
    rows=list(Member.objects.filter(tenant=request.tenant).values('id','first_name','last_name','email','phone'))
    pairs=[]
    for i,a in enumerate(rows):
        for b in rows[i+1:]:
            if likely_duplicate(a,b): pairs.append({'a':str(a['id']),'b':str(b['id'])})
    return JsonResponse({'pairs':pairs})

@require_http_methods(['POST'])
@permission_required('candidates.edit')
def candidate_transition(request,pk):
    c=_candidate_for_request(request,pk)
    if not c: return JsonResponse({'error':'not_found'},status=404)
    data=_body(request); old=c.stage; target=str(data.get('target',''))
    stages=(request.tenant.settings or {}).get('candidate_stages') or ['enquiry','application','investigation','ballot','accepted','EA','FC','MM']
    needs_override=bool(data.get('allow_regression',False) or data.get('allow_skip',False))
    # Skips/regressions require explicit approval capability, not merely a client flag.
    if needs_override and not has_permission(request.tenant_membership.role,'candidates.approve',effective_custom_permissions(request.tenant_membership)):
        return JsonResponse({'error':'approval_required'},status=403)
    outstanding_required=c.tasks.filter(required_for_stage=True,completed_at__isnull=True).exclude(status__in=['completed','cancelled'])
    if outstanding_required.exists() and not needs_override:
        return JsonResponse({'error':'required_candidate_tasks_incomplete','task_ids':[str(x.id) for x in outstanding_required[:25]]},status=409)
    tr=transition_candidate_pipeline(old,target,stages,allow_regression=bool(data.get('allow_regression',False)),allow_skip=bool(data.get('allow_skip',False)))
    if not tr.allowed: return JsonResponse({'error':'invalid_transition','reason':tr.reason,'from':old,'to':target},status=409)
    history=list(c.stage_history or []); history.append({'from':old,'to':target,'at':timezone.now().isoformat(),'by':request.user.id})
    c.stage=target; c.stage_order=stages.index(target); c.stage_history=history; c.save(update_fields=['stage','stage_order','stage_history','updated_at'])
    append_audit(request=request,action='candidates.transitioned',object_type='Candidate',object_id=c.pk,before={'stage':old},after={'stage':c.stage},metadata={'override':needs_override})
    return JsonResponse(_serialize(c))

@require_http_methods(['GET','POST'])
def attendance(request,meeting_id):
    try: meeting=Meeting.objects.get(pk=meeting_id,tenant=request.tenant)
    except Meeting.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
    if request.method=='GET':
        if not has_permission(request.tenant_membership.role,'meetings.view',effective_custom_permissions(request.tenant_membership)): return JsonResponse({'error':'forbidden'},status=403)
        return JsonResponse({'results':[_serialize(x) for x in Attendance.objects.filter(meeting=meeting,tenant=request.tenant).select_related('member')]})
    if not has_permission(request.tenant_membership.role,'meetings.edit',effective_custom_permissions(request.tenant_membership)): return JsonResponse({'error':'forbidden'},status=403)
    data=_body(request); member_id=data.get('member_id'); status=data.get('status','present')
    try: member=Member.objects.get(pk=member_id,tenant=request.tenant)
    except Member.DoesNotExist: return JsonResponse({'error':'member_not_found'},status=404)
    obj,_=Attendance.objects.update_or_create(tenant=request.tenant,meeting=meeting,member=member,defaults={'status':status,'checked_in_at':timezone.now() if status=='present' else None})
    append_audit(request=request,action='meetings.attendance.updated',object_type='Attendance',object_id=obj.pk,after={'member_id':str(member.id),'status':status})
    return JsonResponse(_serialize(obj))

@require_http_methods(['GET','POST'])
def minutes_view(request,meeting_id):
    try: meeting=Meeting.objects.get(pk=meeting_id,tenant=request.tenant)
    except Meeting.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
    if request.method=='GET':
        if not has_permission(request.tenant_membership.role,'meetings.view',effective_custom_permissions(request.tenant_membership)): return JsonResponse({'error':'forbidden'},status=403)
        qs=MinuteVersion.objects.filter(meeting=meeting,tenant=request.tenant).order_by('-version'); return JsonResponse({'results':[_serialize(x) for x in qs]})
    if not has_permission(request.tenant_membership.role,'meetings.edit',effective_custom_permissions(request.tenant_membership)): return JsonResponse({'error':'forbidden'},status=403)
    data=_body(request); latest=MinuteVersion.objects.filter(meeting=meeting,tenant=request.tenant).order_by('-version').first()
    if latest and latest.state in {'approved','locked'}: return JsonResponse({'error':'controlled_reopen_required','state':latest.state,'reopen_endpoint':f'/api/meetings/{meeting.id}/minutes/{latest.version}/reopen'},status=409)
    body=str(data.get('body') or '').strip()
    if not body:return JsonResponse({'error':'minutes_body_required'},status=400)
    version=(latest.version+1 if latest else 1)
    obj=MinuteVersion.objects.create(tenant=request.tenant,meeting=meeting,version=version,state='draft',body=body,changed_by=request.user)
    append_audit(request=request,action='meetings.minutes.version_created',object_type='MinuteVersion',object_id=obj.pk,after={'version':version,'source_state':latest.state if latest else None})
    return JsonResponse(_serialize(obj),status=201)

@require_http_methods(['POST'])
@permission_required('meetings.approve')
def minutes_transition(request,meeting_id,version):
    try: obj=MinuteVersion.objects.select_related('meeting').get(meeting_id=meeting_id,version=version,tenant=request.tenant)
    except MinuteVersion.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
    data=_body(request);target=str(data.get('target','')).lower();comment=str(data.get('comment') or '').strip()
    tr=transition_minutes(obj.state,target)
    if not tr.allowed: return JsonResponse({'error':'invalid_transition','reason':tr.reason,'from':obj.state,'to':target},status=409)
    old=obj.state; obj.state=target
    if target=='review': obj.reviewer_comments=comment[:5000]
    if target=='approved': obj.approved_by=request.user;obj.approved_at=timezone.now()
    if target=='locked': obj.locked_by=request.user;obj.locked_at=timezone.now()
    obj.save(update_fields=['state','reviewer_comments','approved_by','approved_at','locked_by','locked_at','updated_at'])
    generated=None
    if target=='approved':
        from .services.pdfs import minutes_pdf
        pdf=minutes_pdf(obj.meeting,obj,obj.meeting.attendance.select_related('member').all(),obj.meeting.visitors.all(),request.tenant.name);generated=_persist_generated_document(request,'meeting_minutes','MinuteVersion',obj.id,pdf,f'meeting-minutes-v{obj.version}')
    append_audit(request=request,action='meetings.minutes.transitioned',object_type='MinuteVersion',object_id=obj.pk,before={'state':old},after={'state':target},metadata={'comment':comment[:500],'generated_document_id':str(generated.id) if generated else ''})
    payload=_serialize(obj);payload['generated_document_id']=str(generated.id) if generated else None;return JsonResponse(payload)

@require_http_methods(['POST'])
@permission_required('meetings.approve')
def minutes_reopen(request,meeting_id,version):
    try: source=MinuteVersion.objects.select_for_update().get(meeting_id=meeting_id,version=version,tenant=request.tenant)
    except MinuteVersion.DoesNotExist:return JsonResponse({'error':'not_found'},status=404)
    if source.state not in {'approved','locked'}:return JsonResponse({'error':'only_approved_or_locked_minutes_can_be_reopened'},status=409)
    reason=str(_body(request).get('reason') or '').strip()
    if len(reason)<5:return JsonResponse({'error':'reopen_reason_required'},status=400)
    with transaction.atomic():
        latest=MinuteVersion.objects.select_for_update().filter(meeting_id=meeting_id,tenant=request.tenant).order_by('-version').first()
        if latest and latest.version!=source.version:return JsonResponse({'error':'only_latest_version_can_be_reopened','latest_version':latest.version},status=409)
        obj=MinuteVersion.objects.create(tenant=request.tenant,meeting=source.meeting,version=source.version+1,state='draft',body=source.body,changed_by=request.user,reopened_from=source,reopen_reason=reason)
    append_audit(request=request,action='meetings.minutes.reopened',object_type='MinuteVersion',object_id=obj.id,before={'source_version':source.version,'source_state':source.state},after={'version':obj.version,'state':'draft'},metadata={'reason':reason[:500]});return JsonResponse(_serialize(obj),status=201)

@require_http_methods(['POST'])
@permission_required('dues.edit')
def record_payment(request):
    data=_body(request)
    try:amount=Decimal(str(data.get('amount','0')))
    except Exception:return JsonResponse({'error':'invalid_amount'},status=400)
    if amount<=0:return JsonResponse({'error':'amount_must_be_positive'},status=400)
    with transaction.atomic():
        try:assessment=DuesAssessment.objects.select_for_update().select_related('cycle','member').get(pk=data.get('assessment_id'),tenant=request.tenant)
        except DuesAssessment.DoesNotExist:return JsonResponse({'error':'assessment_not_found'},status=404)
        due=amount_due(assessment.amount,assessment.paid_amount,assessment.waived_amount,assessment.late_fee,assessment.adjustment_amount)
        if amount>due:return JsonResponse({'error':'overpayment_not_allowed','balance':str(due)},status=409)
        locked_tenant=Tenant.objects.select_for_update().get(pk=request.tenant.pk);settings=dict(locked_tenant.settings or {})
        finance_cfg=settings.get('finance') or {};configured_methods=finance_cfg.get('payment_methods') or ['cash','bank_transfer','cheque','manual'];configured_methods=configured_methods if isinstance(configured_methods,list) else [];allowed_methods={str(x).lower() for x in configured_methods if str(x).strip()};method=str(data.get('method') or 'manual').lower();method={'bank':'bank_transfer'}.get(method,method)
        if method not in allowed_methods:return JsonResponse({'error':'payment_method_not_enabled','allowed':sorted(allowed_methods)},status=400)
        currency=str(data.get('currency') or assessment.currency).upper()
        if currency!=assessment.currency:return JsonResponse({'error':'payment_currency_must_match_assessment','expected':assessment.currency},status=409)
        seqs=dict(settings.get('receipt_sequences') or {});year=str(timezone.now().year);seq=int(seqs.get(year,0))+1;seqs[year]=seq;settings['receipt_sequences']=seqs;locked_tenant.settings=settings;locked_tenant.save(update_fields=['settings','updated_at'])
        receipt_prefix=str(settings.get('receipt_prefix') or locked_tenant.lodge_number or locked_tenant.slug);rec=receipt_number(receipt_prefix,int(year),seq)
        payment=Payment.objects.create(tenant=locked_tenant,member=assessment.member,assessment=assessment,amount=amount,currency=currency,method=method,provider=data.get('provider',''),provider_reference=str(data.get('provider_reference') or '')[:180],receipt_number=rec,notes=str(data.get('notes') or ''))
        allocation=PaymentAllocation.objects.create(tenant=locked_tenant,payment=payment,assessment=assessment,amount=amount)
        assessment.paid_amount+=amount;assessment.status=assessment_status(assessment.amount,assessment.paid_amount,assessment.waived_amount,assessment.late_fee,timezone.now().date()>assessment.cycle.due_on,assessment.adjustment_amount);assessment.save(update_fields=['paid_amount','status','updated_at'])
    append_audit(request=request,action='dues.payment.recorded',object_type='Payment',object_id=payment.pk,after={'amount':str(amount),'currency':currency,'method':method,'receipt_number':rec},metadata={'allocation_id':str(allocation.id)})
    return JsonResponse({'payment':_serialize(payment),'allocation':_serialize(allocation),'assessment':_serialize(assessment)},status=201)

@require_http_methods(['GET'])
@permission_required('reports.view')
def reports_summary(request):
    members=Member.objects.filter(tenant=request.tenant); meetings=Meeting.objects.filter(tenant=request.tenant)
    assessments=DuesAssessment.objects.filter(tenant=request.tenant); expenses_q=Expense.objects.filter(tenant=request.tenant)
    by_status=list(members.values('status').annotate(count=Count('id')).order_by('status'))
    by_degree=list(members.values('degree').annotate(count=Count('id')).order_by('degree'))
    attendance_rows=[]
    for m in meetings.order_by('-starts_at')[:24]: attendance_rows.append({'meeting':m.title,'date':m.starts_at.date().isoformat(),'present':m.attendance.filter(status='present').count()})
    billed=assessments.aggregate(v=Sum('amount'))['v'] or Decimal('0'); paid=assessments.aggregate(v=Sum('paid_amount'))['v'] or Decimal('0'); expenses=expenses_q.aggregate(v=Sum('amount'))['v'] or Decimal('0')
    candidates_q=Candidate.objects.filter(tenant=request.tenant); candidate_stage=list(candidates_q.values('stage','stage_order').annotate(count=Count('id')).order_by('stage_order'))
    candidate_stages=((request.tenant.settings or {}).get('candidate_stages') or ['enquiry','application','investigation','ballot','accepted','EA','FC','MM'])
    candidate_conversion=candidate_conversion_funnel(candidates_q.only('stage','stage_history'),candidate_stages)
    mentor_workload=list(CandidateMentorAssignment.objects.filter(tenant=request.tenant,is_active=True).values('mentor_id','mentor__first_name','mentor__last_name').annotate(count=Count('candidate')).order_by('-count')[:50])
    durations=[]
    for c in candidates_q[:2000]:
        history=c.stage_history or []
        for i,item in enumerate(history[:-1]):
            try:
                a=timezone.datetime.fromisoformat(item.get('at')); b=timezone.datetime.fromisoformat(history[i+1].get('at')); durations.append({'stage':item.get('to'),'days':max(0,(b-a).total_seconds()/86400)})
            except Exception: pass
    stage_duration={}
    for row in durations: stage_duration.setdefault(row['stage'],[]).append(row['days'])
    stage_duration={k:round(sum(v)/len(v),1) for k,v in stage_duration.items() if v}
    return JsonResponse({'membership':{'status':by_status,'degree':by_degree,'data_completeness_avg':round(sum(completeness(x) for x in members.values('first_name','last_name','email','phone','status'))/max(members.count(),1),1)},'attendance':attendance_rows,'candidates':{'by_stage':candidate_stage,'conversion':candidate_conversion,'average_days_in_stage':stage_duration,'mentor_workload':mentor_workload,'total':candidates_q.count()},'finance':{'billed':str(billed),'paid':str(paid),'expenses':str(expenses),'net':str(paid-expenses)}})


@require_http_methods(['GET'])
@permission_required('reports.view')
def report_catalog_view(request):
    return JsonResponse({'results':report_catalog(include_platform=False),'formats':sorted(FORMATS)})


def _platform_report_access(request, report_id):
    if not getattr(request,'user',None) or not request.user.is_authenticated:
        return JsonResponse({'error':'authentication_required'},status=401)
    try:permission=platform_permission_for_report(report_id)
    except ValueError as exc:return JsonResponse({'error':str(exc)},status=404)
    if not user_has_platform_permission(request.user,permission):
        return JsonResponse({'error':'platform_forbidden','permission':permission},status=403)
    return None


@require_http_methods(['GET'])
def platform_report_catalog_view(request):
    if not getattr(request,'user',None) or not request.user.is_authenticated:
        return JsonResponse({'error':'authentication_required'},status=401)
    results=[]
    for definition in report_catalog(include_platform=True):
        if not definition.get('platform'):continue
        if user_has_platform_permission(request.user,platform_permission_for_report(definition['id'])):
            results.append(definition)
    return JsonResponse({'results':results,'formats':sorted(FORMATS)})


@require_http_methods(['GET'])
def platform_report_preview_view(request,report_type):
    try:report_id=canonical_report_id(report_type)
    except ValueError as exc:return JsonResponse({'error':str(exc)},status=404)
    denied=_platform_report_access(request,report_id)
    if denied:return denied
    filters=_report_filters_from_request(request)
    try:payload=platform_report_preview(report_id,filters,limit=request.GET.get('limit',100))
    except (ValueError,TypeError) as exc:return JsonResponse({'error':str(exc)},status=400)
    payload['definition']={'id':report_id,**payload['definition']}
    payload['metadata']={'context':'platform','filters':filters,'generated_at':timezone.now().isoformat(),'total_rows':payload['total_rows']}
    return JsonResponse(payload)


@require_http_methods(['GET'])
def platform_report_generate(request,report_type):
    try:report_id=canonical_report_id(report_type)
    except ValueError as exc:return JsonResponse({'error':str(exc)},status=404)
    denied=_platform_report_access(request,report_id)
    if denied:return denied
    fmt=str(request.GET.get('format','pdf')).lower();filters=_report_filters_from_request(request)
    try:data,mime,ext,row_count=render_platform_report_bytes(report_id,fmt,filters)
    except (ValueError,TypeError) as exc:return JsonResponse({'error':str(exc)},status=400)
    digest=hashlib.sha256(data).hexdigest();filename=safe_report_filename(PLATFORM_REPORT_CONTEXT,report_id,ext)
    response=HttpResponse(data,content_type=mime);response['Content-Disposition']=f'attachment; filename="{filename}"';response['X-Report-SHA256']=digest;response['Cache-Control']='private, no-store';response['X-Content-Type-Options']='nosniff'
    append_audit(request=request,actor=request.user,tenant=None,action='platform.reports.generated',object_type='PlatformReport',object_id=report_id,metadata={'format':fmt,'row_count':row_count,'sha256':digest,'filters':filters})
    return response


def _report_filters_from_request(request):
    excluded={'format','async','approval_id','limit'}
    return {k:v for k,v in request.GET.items() if k not in excluded and str(v).strip()!=''}


def _report_requires_approval(tenant,report_id,row_count):
    cfg=((tenant.settings or {}).get('reports') or {})
    definition=REPORT_CATALOG[report_id]
    if definition.get('sensitive') and bool(cfg.get('sensitive_approval_enabled',True)):
        return True
    try: threshold=int(cfg.get('approval_record_threshold') or 0)
    except (TypeError,ValueError): threshold=0
    return bool(threshold and row_count>=threshold)


def _report_approval(request,report_id,fmt,filters,row_count):
    if not _report_requires_approval(request.tenant,report_id,row_count): return None
    approval_id=request.GET.get('approval_id')
    approval=ExportApprovalRequest.objects.filter(pk=approval_id,tenant=request.tenant,requester=request.user,export_type=f'report:{report_id}',format=fmt,filters=filters,status='approved').filter(Q(expires_at__isnull=True)|Q(expires_at__gt=timezone.now())).first() if approval_id else None
    return approval or False


@require_http_methods(['GET'])
@permission_required('reports.view')
def report_preview_view(request,report_type):
    filters=_report_filters_from_request(request)
    try:
        report_id=canonical_report_id(report_type)
        if REPORT_CATALOG[report_id].get('platform'): return JsonResponse({'error':'platform_report_requires_operator_context'},status=403)
        payload=report_preview(request.tenant,report_id,filters,limit=request.GET.get('limit',100))
    except (ValueError,TypeError) as exc: return JsonResponse({'error':str(exc)},status=400)
    payload['definition']={'id':report_id,**payload['definition']}
    payload['metadata']={'tenant':request.tenant.name,'tenant_id':str(request.tenant.id),'filters':filters,'generated_at':timezone.now().isoformat(),'total_rows':payload['total_rows']}
    return JsonResponse(payload)


@require_http_methods(['GET'])
@permission_required('reports.export')
def report_generate(request,report_type):
    fmt=str(request.GET.get('format','pdf')).lower(); filters=_report_filters_from_request(request)
    try:
        report_id=canonical_report_id(report_type); definition=REPORT_CATALOG[report_id]
    except ValueError as exc:return JsonResponse({'error':str(exc)},status=404)
    if definition.get('platform'): return JsonResponse({'error':'platform_report_requires_operator_context'},status=403)
    if fmt not in FORMATS or fmt not in definition.get('formats',[]):return JsonResponse({'error':'unsupported_report_format','allowed':definition.get('formats',[])},status=400)
    try: row_count=estimate_report_rows(request.tenant,report_id,filters)
    except (ValueError,TypeError) as exc:return JsonResponse({'error':str(exc)},status=400)
    approval=_report_approval(request,report_id,fmt,filters,row_count)
    if approval is False:
        return JsonResponse({'error':'report_approval_required','request_endpoint':'/api/exports/approvals','export_type':f'report:{report_id}','format':fmt,'filters':filters},status=428)
    cfg=((request.tenant.settings or {}).get('reports') or {})
    try: async_threshold=max(100,int(cfg.get('async_threshold_rows',5000)))
    except (TypeError,ValueError): async_threshold=5000
    async_requested=str(request.GET.get('async','')).lower() in {'1','true','yes'} or row_count>=async_threshold
    if async_requested:
        obj=GeneratedReport.objects.create(tenant=request.tenant,report_type=report_id,format=fmt,filters=filters,reporting_period='',generated_by=request.user,status='queued',progress=0,row_count=row_count,metadata={'report_id':report_id,'report_name':definition['name'],'approval_id':str(approval.id) if approval else None})
        try:
            from .tasks import generate_report_async
            task=generate_report_async.delay(str(obj.id)); obj.job_id=str(task.id or '')[:80];obj.progress=5;obj.save(update_fields=['job_id','progress','updated_at'])
        except Exception as exc:
            obj.status='failed';obj.progress=100;obj.error=f'queue_unavailable:{type(exc).__name__}';obj.save(update_fields=['status','progress','error','updated_at'])
            append_audit(request=request,action='reports.queue_failed',object_type='GeneratedReport',object_id=obj.id,metadata={'report_id':report_id,'format':fmt,'error':type(exc).__name__})
            return JsonResponse({'error':'background_queue_unavailable','generated_report_id':str(obj.id)},status=503)
        append_audit(request=request,action='reports.queued',object_type='GeneratedReport',object_id=obj.id,metadata={'report_id':report_id,'format':fmt,'row_count':row_count,'job_id':obj.job_id,'approval_id':str(approval.id) if approval else None})
        return JsonResponse({'id':str(obj.id),'status':obj.status,'progress':obj.progress,'job_id':obj.job_id,'estimated_rows':row_count},status=202)
    try: obj,data,mime=generate_report(request.tenant,report_id,fmt,filters,request.user)
    except ValueError as exc:return JsonResponse({'error':str(exc)},status=400)
    filename=obj.metadata.get('filename') or safe_report_filename(request.tenant,report_id,fmt)
    response=HttpResponse(data,content_type=mime); response['Content-Disposition']=f'attachment; filename="{filename}"'; response['X-Report-SHA256']=obj.sha256; response['Cache-Control']='private, no-store';response['X-Content-Type-Options']='nosniff'
    append_audit(request=request,action='reports.generated',object_type='GeneratedReport',object_id=obj.id,metadata={'report_id':report_id,'format':fmt,'row_count':obj.row_count,'sha256':obj.sha256,'approval_id':str(approval.id) if approval else None}); return response


@require_http_methods(['GET'])
@permission_required('reports.view')
def generated_reports(request):
    try: limit=max(1,min(int(request.GET.get('limit',100)),500))
    except (TypeError,ValueError): return JsonResponse({'error':'invalid_limit'},status=400)
    qs=GeneratedReport.objects.filter(tenant=request.tenant).order_by('-created_at')[:limit]
    return JsonResponse({'results':[_serialize(x) for x in qs]})


@require_http_methods(['GET'])
@permission_required('reports.export')
def generated_report_download(request,pk):
    obj=GeneratedReport.objects.filter(pk=pk,tenant=request.tenant,status='completed').first()
    if not obj or not obj.file_path:return JsonResponse({'error':'not_found_or_not_ready'},status=404)
    try: stream=default_storage.open(obj.file_path,'rb')
    except Exception:return JsonResponse({'error':'stored_report_unavailable'},status=410)
    mime=(obj.metadata or {}).get('mime_type') or {'pdf':'application/pdf','csv':'text/csv','xlsx':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'}.get(obj.format,'application/octet-stream')
    filename=(obj.metadata or {}).get('filename') or safe_report_filename(request.tenant,obj.report_type,obj.format,obj.created_at)
    response=FileResponse(stream,content_type=mime);response['Content-Disposition']=f'attachment; filename="{filename}"';response['Cache-Control']='private, no-store';response['X-Content-Type-Options']='nosniff';response['X-Report-SHA256']=obj.sha256
    append_audit(request=request,action='reports.downloaded',object_type='GeneratedReport',object_id=obj.id,metadata={'report_id':obj.report_type,'format':obj.format,'sha256':obj.sha256})
    return response


@require_http_methods(['GET','POST'])
@permission_required('reports.view')
def saved_reports(request):
    if request.method=='GET':
        qs=SavedReport.objects.filter(tenant=request.tenant).filter(Q(owner=request.user)|Q(is_shared=True)).order_by('name')[:500]
        return JsonResponse({'results':[_serialize(x) for x in qs]})
    data=_body(request); report_type=str(data.get('report_type',''))
    try: report_type=canonical_report_id(report_type)
    except ValueError as exc:return JsonResponse({'error':str(exc)},status=400)
    if REPORT_CATALOG[report_type].get('platform'):return JsonResponse({'error':'platform_report_requires_operator_context'},status=403)
    obj=SavedReport.objects.create(tenant=request.tenant,owner=request.user,name=str(data.get('name') or REPORT_CATALOG[report_type]['name'])[:160],report_type=report_type,filters=data.get('filters') or {},columns=data.get('columns') or [],grouping=data.get('grouping') or [],sorting=data.get('sorting') or [],is_shared=bool(data.get('is_shared',False)))
    append_audit(request=request,action='reports.saved_view.created',object_type='SavedReport',object_id=obj.id); return JsonResponse(_serialize(obj),status=201)

@require_http_methods(['PATCH','DELETE'])
@permission_required('reports.view')
def saved_report_detail(request,pk):
    obj=SavedReport.objects.filter(pk=pk,tenant=request.tenant,owner=request.user).first()
    if not obj:return JsonResponse({'error':'not_found'},status=404)
    if request.method=='DELETE':obj.delete();append_audit(request=request,action='reports.saved_view.deleted',object_type='SavedReport',object_id=pk);return HttpResponse(status=204)
    data=_body(request)
    for key in ['name','report_type','filters','columns','grouping','sorting','is_shared']:
        if key in data:setattr(obj,key,data[key])
    try: obj.report_type=canonical_report_id(obj.report_type)
    except ValueError as exc:return JsonResponse({'error':str(exc)},status=400)
    if REPORT_CATALOG[obj.report_type].get('platform'):return JsonResponse({'error':'platform_report_requires_operator_context'},status=403)
    obj.full_clean();obj.save();append_audit(request=request,action='reports.saved_view.updated',object_type='SavedReport',object_id=obj.id);return JsonResponse(_serialize(obj))

def _export_filter_params(request):
    return {k:v for k,v in request.GET.items() if k not in {'format','approval_id'} and v not in (None,'')}


def _apply_export_filters(qs, export_type, filters):
    filters=dict(filters or {})
    if export_type=='members':
        if filters.get('status'):qs=qs.filter(status=filters['status'])
        if filters.get('degree'):qs=qs.filter(degree=filters['degree'])
    elif export_type=='meetings':
        if filters.get('status'):qs=qs.filter(status=filters['status'])
        if filters.get('date_from'):qs=qs.filter(starts_at__date__gte=filters['date_from'])
        if filters.get('date_to'):qs=qs.filter(starts_at__date__lte=filters['date_to'])
    elif export_type=='finance':
        if filters.get('member_id'):qs=qs.filter(member_id=filters['member_id'])
        if filters.get('method'):qs=qs.filter(method=filters['method'])
        if filters.get('date_from'):qs=qs.filter(paid_at__date__gte=filters['date_from'])
        if filters.get('date_to'):qs=qs.filter(paid_at__date__lte=filters['date_to'])
        qs=qs.filter(is_reversed=False)
    elif export_type=='audit':
        if filters.get('action'):qs=qs.filter(action__startswith=filters['action'])
        if filters.get('actor_id'):qs=qs.filter(actor_id=filters['actor_id'])
        if filters.get('object_type'):qs=qs.filter(object_type=filters['object_type'])
        if filters.get('object_id'):qs=qs.filter(object_id=filters['object_id'])
        if filters.get('date_from'):qs=qs.filter(occurred_at__date__gte=filters['date_from'])
        if filters.get('date_to'):qs=qs.filter(occurred_at__date__lte=filters['date_to'])
    return qs


@require_http_methods(['GET'])
def export_view(request,export_type):
    mapping={'members':(Member,['first_name','last_name','email','phone','status','degree'],'members'),'meetings':(Meeting,['title','meeting_type','starts_at','location','status'],'meetings'),'finance':(Payment,['receipt_number','amount','currency','method','provider','provider_reference','paid_at','member_id','is_reversed','reversal_reason'],'finance'),'audit':(AuditEvent,['occurred_at','action','object_type','object_id','actor_id'],'audit')}
    if export_type not in mapping:return JsonResponse({'error':'unknown_export'},status=404)
    model,fields,module=mapping[export_type]
    if not has_permission(request.tenant_membership.role,f'{module}.export' if module!='audit' else 'reports.export',effective_custom_permissions(request.tenant_membership)):return JsonResponse({'error':'forbidden'},status=403)
    fmt=request.GET.get('format','csv').lower();filters=_export_filter_params(request);security=((request.tenant.settings or {}).get('security') or {});approval=None
    if security.get('export_requires_approval') and export_type in {'members','finance','audit'}:
        approval_id=request.GET.get('approval_id')
        approval=ExportApprovalRequest.objects.filter(pk=approval_id,tenant=request.tenant,requester=request.user,export_type=export_type,format=fmt,status='approved').first() if approval_id else None
        if not approval or (approval.expires_at and approval.expires_at<=timezone.now()):return JsonResponse({'error':'export_approval_required','request_endpoint':'/api/exports/approvals'},status=428)
        if dict(approval.filters or {})!=filters:return JsonResponse({'error':'approved_filters_do_not_match_request'},status=409)
    qs=_apply_export_filters(model.objects.filter(tenant=request.tenant),export_type,filters)
    return export_queryset(request,qs,fields,export_type,fmt,approval=approval,filters=filters)


@require_http_methods(['GET','POST'])
def export_approvals(request):
    if request.method=='GET':
        can_approve=has_permission(request.tenant_membership.role,'reports.approve',effective_custom_permissions(request.tenant_membership))
        qs=ExportApprovalRequest.objects.filter(tenant=request.tenant).select_related('requester','approved_by')
        if not can_approve:qs=qs.filter(requester=request.user)
        rows=[]
        for x in qs.order_by('-created_at')[:200]:
            row=_serialize(x);row['requester_name']=x.requester.get_username();row['approver_name']=x.approved_by.get_username() if x.approved_by else '';rows.append(row)
        return JsonResponse({'results':rows})
    if not has_permission(request.tenant_membership.role,'reports.export',effective_custom_permissions(request.tenant_membership)):return JsonResponse({'error':'forbidden'},status=403)
    data=_body(request);export_type=str(data.get('export_type') or '');fmt=str(data.get('format','csv')).lower();reason=str(data.get('reason') or '').strip();filters=dict(data.get('filters') or {})
    valid=export_type in {'members','finance','audit','meetings'}
    if export_type.startswith('report:'):
        try:rid=canonical_report_id(export_type.split(':',1)[1]);valid=not REPORT_CATALOG[rid].get('platform');export_type=f'report:{rid}'
        except (ValueError,KeyError):valid=False
    if not valid or fmt not in {'csv','xlsx','pdf'}:return JsonResponse({'error':'invalid_export_request'},status=400)
    if len(reason)<8:return JsonResponse({'error':'reason_required_min_8_chars'},status=400)
    try:hours=max(1,min(int((((request.tenant.settings or {}).get('security') or {}).get('export_approval_hours',24))),168))
    except (TypeError,ValueError):hours=24
    obj=ExportApprovalRequest.objects.create(tenant=request.tenant,requester=request.user,export_type=export_type,format=fmt,filters=filters,reason=reason,expires_at=timezone.now()+timezone.timedelta(hours=hours))
    append_audit(request=request,action='export.approval.requested',object_type='ExportApprovalRequest',object_id=obj.id,metadata={'export_type':export_type,'format':fmt,'filters':filters,'reason':reason,'expires_at':obj.expires_at.isoformat()});return JsonResponse(_serialize(obj),status=201)


@require_http_methods(['POST'])
@permission_required('reports.approve')
def decide_export_approval(request,pk):
    try:obj=ExportApprovalRequest.objects.get(pk=pk,tenant=request.tenant,status='pending')
    except ExportApprovalRequest.DoesNotExist:return JsonResponse({'error':'not_found'},status=404)
    if obj.expires_at and obj.expires_at<=timezone.now():obj.status='expired';obj.decided_at=timezone.now();obj.save(update_fields=['status','decided_at','updated_at']);return JsonResponse({'error':'approval_request_expired'},status=409)
    data=_body(request);decision=data.get('decision');decision_reason=str(data.get('reason') or '').strip()
    if decision not in ('approved','rejected'):return JsonResponse({'error':'invalid_decision'},status=400)
    if len(decision_reason)<5:return JsonResponse({'error':'decision_reason_required_min_5_chars'},status=400)
    dual_control=bool((((request.tenant.settings or {}).get('security') or {}).get('export_dual_control',True)))
    if obj.requester_id==request.user.id and dual_control:return JsonResponse({'error':'self_approval_not_allowed'},status=409)
    before={'status':obj.status,'requester_id':obj.requester_id,'reason':obj.reason,'expires_at':obj.expires_at.isoformat() if obj.expires_at else None}
    obj.status=decision;obj.approved_by=request.user;obj.decided_at=timezone.now();obj.decision_reason=decision_reason;obj.save(update_fields=['status','approved_by','decided_at','decision_reason','updated_at'])
    append_audit(request=request,action=f'export.approval.{decision}',object_type='ExportApprovalRequest',object_id=obj.id,before=before,after={'status':decision,'approver_id':request.user.id,'decision_reason':decision_reason});return JsonResponse(_serialize(obj))


@require_http_methods(['GET'])
@permission_required('audit.view')
def audit_log(request):
    qs=AuditEvent.objects.filter(tenant=request.tenant).select_related('actor')
    if request.GET.get('action'):qs=qs.filter(action__startswith=request.GET['action'])
    if request.GET.get('actor_id'):qs=qs.filter(actor_id=request.GET['actor_id'])
    if request.GET.get('object_type'):qs=qs.filter(object_type=request.GET['object_type'])
    if request.GET.get('object_id'):qs=qs.filter(object_id=request.GET['object_id'])
    if request.GET.get('date_from'):qs=qs.filter(occurred_at__date__gte=request.GET['date_from'])
    if request.GET.get('date_to'):qs=qs.filter(occurred_at__date__lte=request.GET['date_to'])
    try:limit=max(1,min(int(request.GET.get('limit',200)),500))
    except (TypeError,ValueError):limit=200
    total=qs.count();rows=[]
    for x in qs.order_by('-occurred_at')[:limit]:
        row=_serialize(x);row['actor_name']=x.actor.get_username() if x.actor else 'System';rows.append(row)
    return JsonResponse({'results':rows,'count':total,'filters':{k:v for k,v in request.GET.items() if v}})


@require_http_methods(['GET'])
@permission_required('audit.view')
def audit_verify(request):
    return JsonResponse(verify_audit_chain(request.tenant))


@require_http_methods(['GET'])
@permission_required('audit.view')
def compliance_exports(request):
    qs=ExportEvent.objects.filter(tenant=request.tenant).select_related('actor','approved_by','approval_request').order_by('-created_at')
    if request.GET.get('export_type'):qs=qs.filter(export_type=request.GET['export_type'])
    if request.GET.get('format'):qs=qs.filter(format=request.GET['format'])
    if request.GET.get('requester_id'):qs=qs.filter(actor_id=request.GET['requester_id'])
    try:limit=max(1,min(int(request.GET.get('limit',200)),500))
    except (TypeError,ValueError):limit=200
    rows=[]
    for x in qs[:limit]:
        row=_serialize(x);row['requester']=x.actor.get_username() if x.actor else 'System';row['approver']=x.approved_by.get_username() if x.approved_by else '';rows.append(row)
    return JsonResponse({'results':rows,'count':qs.count()})


@require_http_methods(['GET','PATCH'])
@permission_required('settings.view')
def tenant_settings(request):
    profile,_=LodgeProfile.objects.get_or_create(tenant=request.tenant)
    if request.method=='GET': return JsonResponse({'tenant':_serialize(request.tenant),'profile':_serialize(profile)})
    if not has_permission(request.tenant_membership.role,'settings.edit',effective_custom_permissions(request.tenant_membership)): return JsonResponse({'error':'forbidden'},status=403)
    data=_body(request); before={'tenant':_serialize(request.tenant),'profile':_serialize(profile)}
    try:
        profile_data=dict(data.get('profile') or {}) if isinstance(data.get('profile',{}),dict) else {}
        if 'custom_fields' in profile_data:profile_data['custom_fields']=normalize_member_custom_fields(profile_data.get('custom_fields'))
        if 'retention_policy' in profile_data:profile_data['retention_policy']=normalize_retention_policy(profile_data.get('retention_policy'))
        _apply(profile,profile_data,['address','meeting_schedule','letterhead','signature_blocks','custom_fields','retention_policy'],request.tenant)
        profile.full_clean(); profile.save()
        tenant_data=data.get('tenant',{}) if isinstance(data.get('tenant',{}),dict) else {}
        for key in ['name','lodge_number','jurisdiction','locale','timezone']:
            if key in tenant_data: setattr(request.tenant,key,tenant_data[key])
        if 'settings' in tenant_data:
            incoming=dict(tenant_data.get('settings') or {}); current=dict(request.tenant.settings or {})
            for reserved in ('receipt_sequences',): incoming.pop(reserved,None)
            if 'security' in incoming:
                incoming['security']=validate_security_settings(incoming.get('security'))
            if 'candidate_stages' in incoming:
                used_stages=set()
                for row in Candidate.objects.filter(tenant=request.tenant).values('stage','stage_history'):
                    if row.get('stage'):used_stages.add(str(row['stage']))
                    for item in row.get('stage_history') or []:
                        if isinstance(item,dict):
                            for key in ('from','to'):
                                if item.get(key):used_stages.add(str(item[key]))
                incoming['candidate_stages']=normalize_candidate_stages(incoming.get('candidate_stages'),used_stages)
            request.tenant.settings={**current,**incoming}
        request.tenant.full_clean(); request.tenant.save()
    except (ValueError,TypeError,ValidationError) as exc:
        return JsonResponse({'error':'validation_error','detail':str(exc)},status=400)
    after={'tenant':_serialize(request.tenant),'profile':_serialize(profile)}
    append_audit(request=request,action='settings.updated',object_type='Tenant',object_id=request.tenant.id,before=before,after=after)
    return JsonResponse(after)

@require_http_methods(['GET','POST'])
@permission_required('settings.admin')
def tenant_users(request):
    if request.method=='GET':
        qs=TenantMembership.objects.filter(tenant=request.tenant).select_related('user','custom_role')
        return JsonResponse({'results':[{'id':str(m.id),'user_id':m.user_id,'username':m.user.get_username(),'email':m.user.email,'role':m.role,'custom_role_id':str(m.custom_role_id) if m.custom_role_id else None,'custom_role_name':m.custom_role.name if m.custom_role_id else None,'custom_permissions':effective_custom_permissions(m),'mfa_required':m.mfa_required,'is_active':m.is_active,'last_login':m.user.last_login.isoformat() if m.user.last_login else None} for m in qs]})
    data=_body(request); requested_role=data.get('role','viewer'); user_id=data.get('user_id'); email=str(data.get('email') or '').strip().lower()
    if requested_role not in dict(TenantMembership.ROLE_CHOICES): return JsonResponse({'error':'invalid_role'},status=400)
    if requested_role in ('owner','admin') and request.tenant_membership.role!='owner': return JsonResponse({'error':'owner_required_for_privileged_role'},status=403)
    User=get_user_model(); membership_created=False; account_created=False
    if user_id:
        try: user=User.objects.get(pk=user_id)
        except User.DoesNotExist: return JsonResponse({'error':'user_not_found'},status=404)
    else:
        if not email or '@' not in email: return JsonResponse({'error':'valid_email_required'},status=400)
        user=User.objects.filter(email__iexact=email).first()
        if not user:
            base=''.join(ch for ch in email.split('@',1)[0] if ch.isalnum() or ch in '._-')[:120] or 'officer'; username=base; n=1
            while User.objects.filter(username=username).exists(): n+=1; username=f'{base[:110]}-{n}'
            user=User(username=username,email=email,is_active=False); user.set_unusable_password(); user.save(); account_created=True

    custom_role=None
    if requested_role=='custom':
        custom_role_id=data.get('custom_role_id')
        if custom_role_id:
            custom_role=CustomRoleDefinition.objects.filter(pk=custom_role_id,tenant=request.tenant).first()
            if not custom_role:return JsonResponse({'error':'invalid_or_cross_tenant_custom_role'},status=400)
            requested_permissions=list(custom_role.permissions or [])
        else:
            # Backwards-compatible explicit custom bundle, still privilege-escalation checked.
            requested_permissions=data.get('custom_permissions',[])
    else:
        requested_permissions=data.get('custom_permissions',[])
    try:
        custom_permissions=ensure_grantable_permissions(request.tenant_membership.role,effective_custom_permissions(request.tenant_membership),requested_permissions)
    except ValueError as exc:
        return JsonResponse({'error':'invalid_permissions','detail':str(exc)},status=400)
    except PermissionError as exc:
        return JsonResponse({'error':'privilege_escalation_blocked','detail':str(exc)},status=403)

    existing=TenantMembership.objects.filter(tenant=request.tenant,user=user).first()
    if existing and existing.role=='owner' and requested_role!='owner' and TenantMembership.objects.filter(tenant=request.tenant,role='owner',is_active=True).count()<=1: return JsonResponse({'error':'cannot_demote_last_owner'},status=409)
    m,membership_created=TenantMembership.objects.update_or_create(tenant=request.tenant,user=user,defaults={'role':requested_role,'custom_role':custom_role,'custom_permissions':([] if custom_role else custom_permissions),'mfa_required':bool(data.get('mfa_required',False)),'is_active':bool(data.get('is_active',True))})
    invite_sent=False
    if bool(data.get('send_invite',False)) or account_created:
        _,token=issue_account_token(user=user,purpose='invitation',tenant=request.tenant,issued_by=request.user,ttl_seconds=django_settings.INVITATION_TOKEN_TTL_SECONDS,metadata={'membership_id':str(m.id)})
        base=django_settings.FRONTEND_BASE_URL.rstrip('/'); link=f'{base}/invite?token={token}'
        try:
            message=f'You have been invited to {request.tenant.name}. Set your password securely using this one-time link:'+'\n\n'+link
            send_mail(f'Your {request.tenant.name} LodgeFlow invitation',message,django_settings.DEFAULT_FROM_EMAIL,[user.email],fail_silently=False); invite_sent=True
        except Exception as exc:
            append_audit(request=request,action='settings.user_invitation.email_failed',object_type='TenantMembership',object_id=m.pk,metadata={'error':type(exc).__name__})
    append_audit(request=request,action='settings.user_membership.changed',object_type='TenantMembership',object_id=m.pk,after={'role':m.role,'custom_role_id':str(m.custom_role_id) if m.custom_role_id else None,'permissions':effective_custom_permissions(m),'mfa_required':m.mfa_required,'is_active':m.is_active,'invite_sent':invite_sent})
    return JsonResponse({'id':str(m.id),'user_id':user.id,'username':user.get_username(),'email':user.email,'role':m.role,'custom_role_id':str(m.custom_role_id) if m.custom_role_id else None,'invite_sent':invite_sent},status=201 if membership_created else 200)

@require_http_methods(['POST'])
def accept_invitation(request):
    data=_body(request); token=str(data.get('token') or ''); password=str(data.get('password') or '')
    if not token:return JsonResponse({'error':'invalid_token'},status=400)
    try:
        with transaction.atomic():
            record=AccountActionToken.objects.select_related('user').filter(digest=__import__('hashlib').sha256(token.encode()).hexdigest()).first()
            if not record:return JsonResponse({'error':'invalid_token'},status=400)
            try: validate_password(password,user=record.user)
            except ValidationError as exc:return JsonResponse({'error':'password_policy','details':list(exc.messages)},status=400)
            record=consume_account_token(raw_token=token,purpose='invitation');user=record.user
            if not record.tenant_id or not record.tenant.is_active or not TenantMembership.objects.filter(user=user,tenant=record.tenant,is_active=True).exists():
                raise AccountTokenError('invitation_membership_inactive')
            user.set_password(password);user.is_active=True;user.save(update_fields=['password','is_active']);login(request,user)
            memberships=list(TenantMembership.objects.filter(user=user,is_active=True,tenant__is_active=True).select_related('tenant','custom_role'))
            append_audit(request=request,actor=user,tenant=record.tenant,action='auth.invitation.accepted',object_type='User',object_id=user.pk)
    except AccountTokenError as exc:return JsonResponse({'error':exc.code},status=400)
    return JsonResponse({'ok':True,'user':{'id':user.id,'email':user.email},'tenants':[{'id':str(m.tenant_id),'name':m.tenant.name} for m in memberships]})

@require_http_methods(['POST'])
def password_reset_request(request):
    client_key=getattr(request,'ip_hash','') or request.META.get('REMOTE_ADDR','unknown')
    if not rate_limit_allow(f'password-reset:{client_key}',limit=6,window=900,critical=True):return JsonResponse({'error':'too_many_attempts'},status=429)
    data=_body(request);email=str(data.get('email') or '').strip().lower();response={'accepted':True,'detail':'If the account exists, a password-reset email will be sent.'}
    user=get_user_model().objects.filter(email__iexact=email,is_active=True).first() if email else None
    if user and user.has_usable_password():
        _,token=issue_account_token(user=user,purpose='password_reset',ttl_seconds=django_settings.PASSWORD_RESET_TOKEN_TTL_SECONDS,metadata={'requested_ip_hash':getattr(request,'ip_hash','')})
        link=f"{django_settings.FRONTEND_BASE_URL.rstrip('/')}/reset-password?token={token}"
        try:send_mail('Reset your LodgeFlow password','A password reset was requested for your account. Use this single-use link:\n\n'+link,django_settings.DEFAULT_FROM_EMAIL,[user.email],fail_silently=False)
        except Exception as exc:append_audit(request=request,actor=user,action='auth.password_reset.email_failed',object_type='User',object_id=user.pk,metadata={'error':type(exc).__name__})
        append_audit(request=request,actor=user,action='auth.password_reset.requested',object_type='User',object_id=user.pk)
    return JsonResponse(response,status=202)

@require_http_methods(['POST'])
def password_reset_confirm(request):
    client_key=getattr(request,'ip_hash','') or request.META.get('REMOTE_ADDR','unknown')
    if not rate_limit_allow(f'password-reset-confirm:{client_key}',limit=8,window=900,critical=True):return JsonResponse({'error':'too_many_attempts'},status=429)
    data=_body(request);token=str(data.get('token') or '');password=str(data.get('password') or '')
    if not token:return JsonResponse({'error':'invalid_token'},status=400)
    record=AccountActionToken.objects.select_related('user').filter(digest=__import__('hashlib').sha256(token.encode()).hexdigest()).first()
    if not record:return JsonResponse({'error':'invalid_token'},status=400)
    try:validate_password(password,user=record.user)
    except ValidationError as exc:return JsonResponse({'error':'password_policy','details':list(exc.messages)},status=400)
    try:
        with transaction.atomic():
            record=consume_account_token(raw_token=token,purpose='password_reset');user=record.user;user.set_password(password);user.save(update_fields=['password']);append_audit(request=request,actor=user,action='auth.password_reset.completed',object_type='User',object_id=user.pk)
    except AccountTokenError as exc:return JsonResponse({'error':exc.code},status=400)
    return JsonResponse({'ok':True})

@require_http_methods(['GET'])
def public_page(request,slug):
    try: p=CMSPage.objects.get(slug=slug,locale=request.GET.get('locale','en'),status='published')
    except CMSPage.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
    return JsonResponse(_serialize(p))

@require_http_methods(['GET'])
def active_announcements(request):
    now=timezone.now(); qs=PlatformAnnouncement.objects.filter(is_active=True).filter(Q(starts_at=None)|Q(starts_at__lte=now)).filter(Q(ends_at=None)|Q(ends_at__gte=now))
    return JsonResponse({'results':[_serialize(x) for x in qs]})

@require_http_methods(['GET'])
def public_site_config(request):
    """Expose only explicitly public presentation settings, navigation and active announcements."""
    locale=request.GET.get('locale','en'); settings={x.key:x.value for x in SiteSetting.objects.filter(is_public=True).order_by('key')}; nav=NavigationMenu.objects.filter(key='primary',locale=locale).first()
    now=timezone.now(); announcements=PlatformAnnouncement.objects.filter(is_active=True).filter(Q(starts_at=None)|Q(starts_at__lte=now)).filter(Q(ends_at=None)|Q(ends_at__gte=now)).order_by('-created_at')[:5]
    return JsonResponse({'settings':settings,'navigation':(nav.items if nav else []),'announcements':[_serialize(x) for x in announcements]})

@require_http_methods(['GET','POST'])
@platform_permission_required('tenants.view')
def platform_tenants(request):
    if request.method=='GET': return JsonResponse({'results':[_serialize(x) for x in Tenant.objects.all().order_by('name')[:500]]})
    if not user_has_platform_permission(request.user,'tenants.edit'): return JsonResponse({'error':'platform_forbidden','permission':'tenants.edit'},status=403)
    data=_body(request); t=Tenant.objects.create(name=data['name'],slug=data['slug'],lodge_number=data.get('lodge_number',''),jurisdiction=data.get('jurisdiction',''),plan=data.get('plan','annual')); append_audit(request=request,actor=request.user,tenant=t,action='platform.tenant.created',object_type='Tenant',object_id=t.pk); return JsonResponse(_serialize(t),status=201)

@require_http_methods(['GET','PATCH','DELETE'])
@platform_permission_required('tenants.view')
def platform_tenant_detail(request,pk):
    try: tenant=Tenant.objects.get(pk=pk)
    except Tenant.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
    if request.method=='GET': return JsonResponse(_serialize(tenant))
    if not user_has_platform_permission(request.user,'tenants.edit'): return JsonResponse({'error':'platform_forbidden','permission':'tenants.edit'},status=403)
    before=_serialize(tenant)
    if request.method=='DELETE':
        tenant.is_active=False; tenant.save(update_fields=['is_active','updated_at']); append_audit(request=request,actor=request.user,tenant=tenant,action='platform.tenant.disabled',object_type='Tenant',object_id=tenant.pk,before=before,after={'is_active':False}); return JsonResponse({'id':str(tenant.id),'is_active':False})
    data=_body(request)
    for key in ['name','slug','lodge_number','jurisdiction','plan','locale','timezone','is_active']:
        if key in data: setattr(tenant,key,data[key])
    if 'settings' in data:
        incoming=dict(data.get('settings') or {}); current=dict(tenant.settings or {}); incoming.pop('receipt_sequences',None); tenant.settings={**current,**incoming}
    tenant.full_clean(); tenant.save(); append_audit(request=request,actor=request.user,tenant=tenant,action='platform.tenant.updated',object_type='Tenant',object_id=tenant.pk,before=before,after=_serialize(tenant)); return JsonResponse(_serialize(tenant))

PROVIDER_REQUIRED_FIELDS={
    'stripe':('secret_key','webhook_secret'),
    'paymongo':('secret_key','webhook_secret'),
    'xendit':('secret_key','callback_token'),
}

def _validate_provider_credentials(provider,values):
    if not isinstance(values,dict) or not values:raise ValueError('credentials_required')
    required=PROVIDER_REQUIRED_FIELDS.get(provider,())
    missing=[key for key in required if not str(values.get(key) or '').strip()]
    if missing:raise ValueError('missing_provider_fields:'+','.join(missing))
    return {str(k):str(v) if not isinstance(v,(dict,list)) else v for k,v in values.items()}


def _tenant_or_global_credential(tenant,provider,mode='live'):
    """Prefer a tenant-specific provider secret, falling back to the platform default."""
    specific=EncryptedCredential.objects.filter(tenant=tenant,provider=provider,mode=mode,is_enabled=True).first()
    return specific or EncryptedCredential.objects.filter(tenant__isnull=True,provider=provider,mode=mode,is_enabled=True).first()


@require_http_methods(['GET','POST'])
@login_required
def platform_credentials(request):
    can_infra=user_has_platform_permission(request.user,'credentials.infrastructure');can_billing=user_has_platform_permission(request.user,'credentials.billing')
    if not (can_infra or can_billing):return JsonResponse({'error':'platform_forbidden','permission':'credentials'},status=403)
    if request.method=='GET':
        qs=EncryptedCredential.objects.select_related('tenant').all()
        if can_billing and not can_infra:qs=qs.filter(provider__in=['stripe','paymongo','xendit'])
        return JsonResponse({'results':[{'id':str(x.id),'tenant_id':str(x.tenant_id) if x.tenant_id else None,'provider':x.provider,'mode':x.mode,'key_version':x.key_version,'health_status':x.health_status,'is_enabled':x.is_enabled,'last_checked_at':x.last_checked_at.isoformat() if x.last_checked_at else None} for x in qs[:1000]]})
    if not has_recent_step_up(request):return JsonResponse({'error':'step_up_required','reauthenticate':True},status=403)
    data=_body(request);tenant_id=data.get('tenant_id');provider=str(data.get('provider') or '').lower().strip();mode=str(data.get('mode','test')).lower();values=data.get('credentials',{})
    if mode not in {'test','live'}:return JsonResponse({'error':'invalid_provider_mode'},status=400)
    if provider in PROVIDER_REQUIRED_FIELDS:
        if not (can_billing or can_infra):return JsonResponse({'error':'platform_forbidden'},status=403)
    elif not can_infra:return JsonResponse({'error':'platform_forbidden','permission':'credentials.infrastructure'},status=403)
    if tenant_id and not Tenant.objects.filter(pk=tenant_id).exists():return JsonResponse({'error':'tenant_not_found'},status=404)
    try:values=_validate_provider_credentials(provider,values)
    except ValueError as exc:return JsonResponse({'error':str(exc)},status=400)
    obj,_=EncryptedCredential.objects.update_or_create(tenant_id=tenant_id or None,provider=provider,mode=mode,defaults={'ciphertext':encrypt_json(values),'key_version':int(data.get('key_version',1)),'health_status':'unchecked','last_checked_at':None})
    security_alert(kind='credential_updated',message=f'{provider} {mode} credential updated',tenant=obj.tenant,actor=request.user,severity='high',metadata={'provider':provider,'mode':mode,'key_version':obj.key_version})
    append_audit(request=request,actor=request.user,tenant=obj.tenant,action='platform.credential.updated',object_type='EncryptedCredential',object_id=obj.pk,metadata={'provider':provider,'mode':mode,'key_version':obj.key_version});return JsonResponse({'id':str(obj.id),'provider':provider,'mode':mode,'key_version':obj.key_version,'health_status':obj.health_status,'is_enabled':obj.is_enabled},status=201)


@require_http_methods(['GET','POST'])
def experiment_variant(request,key):
    subject=(str(request.user.id) if request.user.is_authenticated else request.COOKIES.get('lf_subject') or request.headers.get('X-Anonymous-ID') or getattr(request,'ip_hash','anonymous'))
    try: exp=ABExperiment.objects.get(key=key,is_active=True); variants=tuple(exp.variants); weights=tuple(exp.weights)
    except ABExperiment.DoesNotExist: variants=('A','B'); weights=(0.5,0.5); exp=None
    variant=assign_variant(key,subject,variants,weights)
    if exp:
        assignment,_=ABAssignment.objects.get_or_create(experiment=exp,subject_key=subject,defaults={'variant':variant,'tenant':getattr(request,'tenant',None)})
        if request.method=='POST': ABExposure.objects.create(assignment=assignment,event=_body(request).get('event','view'),metadata=_body(request).get('metadata',{}))
    response=JsonResponse({'experiment':key,'variant':variant});
    if not request.COOKIES.get('lf_subject'): response.set_cookie('lf_subject',subject,max_age=31536000,httponly=True,samesite='Lax',secure=not __import__('django.conf').conf.settings.DEBUG)
    return response

@csrf_exempt
@require_http_methods(['POST'])
def webhook(request,provider,tenant_id=None,mode='live'):
    """Signature-verified, tenant-scoped provider ingress.

    Tenant and provider mode are resolved from the registered callback URL, never
    from custom headers that Stripe/PayMongo/Xendit cannot be expected to send.
    """
    provider=provider.lower().strip(); mode=str(mode or 'live').lower().strip()
    if tenant_id is None:
        return JsonResponse({'error':'tenant_scoped_webhook_url_required'},status=410)
    if mode not in {'test','live'}: return JsonResponse({'error':'invalid_provider_mode'},status=404)
    client_key=getattr(request,'ip_hash','') or request.META.get('REMOTE_ADDR','unknown')
    if not rate_limit_allow(f'webhook:{provider}:{tenant_id}:{client_key}',limit=240,window=60,critical=True): return JsonResponse({'error':'too_many_webhooks'},status=429)
    payload=request.body
    if len(payload)>2*1024*1024: return JsonResponse({'error':'payload_too_large'},status=413)
    try: data=json.loads(payload or b'{}')
    except json.JSONDecodeError: return JsonResponse({'error':'invalid_json'},status=400)
    try: tenant=Tenant.objects.get(pk=tenant_id,is_active=True)
    except (Tenant.DoesNotExist,ValueError): return JsonResponse({'error':'tenant_not_found'},status=404)
    cred=_tenant_or_global_credential(tenant,provider,mode)
    if not cred: return JsonResponse({'error':'provider_not_configured'},status=503)
    try: secrets=decrypt_json(cred.ciphertext)
    except Exception: return JsonResponse({'error':'provider_configuration_unavailable'},status=503)
    if provider=='stripe':
        try:data=StripeAdapter.parse_webhook(payload,request.headers.get('Stripe-Signature',''),secrets.get('webhook_secret',''))
        except Exception:return JsonResponse({'error':'invalid_signature'},status=400)
    elif provider=='xendit':
        if not XenditAdapter.verify_callback_token(secrets.get('callback_token',''),request.headers.get('x-callback-token','')): return JsonResponse({'error':'invalid_signature'},status=400)
    elif provider=='paymongo':
        if not PayMongoAdapter.verify_signature(payload,request.headers.get('Paymongo-Signature',''),secrets.get('webhook_secret',''),mode=mode): return JsonResponse({'error':'invalid_signature'},status=400)
    else: return JsonResponse({'error':'unsupported_provider'},status=404)
    event_id=data.get('id') or data.get('event_id') or data.get('data',{}).get('id') or request.headers.get('X-Event-ID')
    if not event_id: return JsonResponse({'error':'event_id_required'},status=400)
    obj,created=WebhookEvent.objects.get_or_create(provider=provider,event_id=str(event_id),mode=mode,defaults={'tenant':tenant,'event_type':data.get('type') or data.get('event',''),'payload':data,'status':'received','attempts':0})
    if not created:
        if obj.tenant_id!=tenant.id or obj.mode!=mode:return JsonResponse({'error':'event_tenant_or_mode_conflict'},status=409)
        return JsonResponse({'status':'duplicate_ignored','id':str(obj.id),'mode':obj.mode})
    from .tasks import process_webhook_event
    try:
        process_webhook_event.delay(str(obj.id))
        queued=True
    except Exception as exc:
        queued=False; obj.status='received'; obj.last_error=f'queue_unavailable:{type(exc).__name__}'; obj.save(update_fields=['status','last_error','updated_at'])
    append_audit(tenant=tenant,action='billing.webhook.verified',object_type='WebhookEvent',object_id=obj.id,metadata={'provider':provider,'event_type':obj.event_type,'mode':mode,'queued':queued})
    # A valid provider event is acknowledged after durable persistence; reconciliation
    # continues in Celery and is idempotent/replayable.
    return JsonResponse({'status':'accepted' if queued else 'accepted_for_retry','id':str(obj.id)},status=202)


from django.contrib.auth import get_user_model
from .services.platform_permissions import platform_permission_required, user_has_platform_permission

@require_http_methods(['GET','POST'])
@platform_permission_required('features.view')
def platform_features(request):
    if request.method=='GET':
        qs=FeatureFlag.objects.select_related('tenant').all().order_by('key')[:1000]; return JsonResponse({'results':[_serialize(x) for x in qs]})
    if not user_has_platform_permission(request.user,'features.edit'): return JsonResponse({'error':'platform_forbidden','permission':'features.edit'},status=403)
    data=_body(request); obj,_=FeatureFlag.objects.update_or_create(tenant_id=data.get('tenant_id') or None,key=data['key'],defaults={'enabled':bool(data.get('enabled',False)),'config':data.get('config',{})})
    append_audit(request=request,actor=request.user,tenant=obj.tenant,action='platform.feature_flag.updated',object_type='FeatureFlag',object_id=obj.pk,after={'key':obj.key,'enabled':obj.enabled})
    return JsonResponse(_serialize(obj),status=201)

@require_http_methods(['GET','POST'])
@platform_permission_required('tenants.view')
def platform_cms_pages(request):
    if not request.user.is_superuser: return JsonResponse({'error':'super_admin_required'},status=403)
    if request.method=='GET': return JsonResponse({'results':[_serialize(x) for x in CMSPage.objects.all().order_by('slug')[:1000]]})
    if not request.user.is_superuser: return JsonResponse({'error':'platform_write_requires_super_admin'},status=403)
    data=_body(request); locale=str(data.get('locale','en')).strip() or 'en'
    try: blocks=validate_cms_blocks(data.get('blocks',[]))
    except ValidationError as exc: return JsonResponse({'error':'cms_validation','detail':exc.message_dict},status=400)
    defaults={k:data[k] for k in ['title','status','meta_title','meta_description','canonical_url','og','schema','robots_index'] if k in data}; defaults['blocks']=blocks
    try:
        if defaults.get('canonical_url'):defaults['canonical_url']=validate_public_url(defaults['canonical_url'],allow_relative=False,https_required=not django_settings.DEBUG)
        if isinstance(defaults.get('og'),dict) and defaults['og'].get('image'):
            defaults['og']={**defaults['og'],'image':validate_public_url(defaults['og']['image'],https_required=not django_settings.DEBUG)}
    except UnsafeURLError as exc:return JsonResponse({'error':'cms_validation','detail':{'url':str(exc)}},status=400)
    existing=CMSPage.objects.filter(slug=data['slug'],locale=locale).first(); before=_serialize(existing) if existing else {}
    obj,created=CMSPage.objects.update_or_create(slug=data['slug'],locale=locale,defaults=defaults)
    append_audit(request=request,actor=request.user,action='platform.cms_page.saved',object_type='CMSPage',object_id=obj.pk,before=before,after=_serialize(obj),metadata={'block_count':len(blocks),'created':created}); return JsonResponse(_serialize(obj),status=201 if created else 200)

@require_http_methods(['GET'])
@platform_permission_required('tenants.view')
def platform_cms_page_revisions(request,pk):
    if not request.user.is_superuser: return JsonResponse({'error':'super_admin_required'},status=403)
    try: page=CMSPage.objects.get(pk=pk)
    except CMSPage.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
    rows=AuditEvent.objects.filter(object_type='CMSPage',object_id=str(page.id),action='platform.cms_page.saved').order_by('-occurred_at')[:50]
    return JsonResponse({'page':_serialize(page),'results':[{'id':str(x.id),'occurred_at':x.occurred_at.isoformat(),'actor_id':x.actor_id,'before':x.before,'after':x.after} for x in rows]})

@require_http_methods(['POST'])
@platform_permission_required('tenants.view')
def platform_cms_page_rollback(request,pk):
    if not request.user.is_superuser: return JsonResponse({'error':'super_admin_required'},status=403)
    try: page=CMSPage.objects.get(pk=pk)
    except CMSPage.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
    data=_body(request); revision_id=data.get('revision_id')
    try: event=AuditEvent.objects.get(pk=revision_id,object_type='CMSPage',object_id=str(page.id),action='platform.cms_page.saved')
    except (AuditEvent.DoesNotExist,ValueError): return JsonResponse({'error':'revision_not_found'},status=404)
    snapshot=event.before or event.after
    if not snapshot or not isinstance(snapshot.get('blocks'),list): return JsonResponse({'error':'revision_not_recoverable'},status=409)
    before=_serialize(page)
    for key in ['slug','title','status','blocks','meta_title','meta_description','canonical_url','og','schema','robots_index','locale']:
        if key in snapshot: setattr(page,key,snapshot[key])
    try: page.blocks=validate_cms_blocks(page.blocks); page.full_clean(); page.save()
    except ValidationError as exc: return JsonResponse({'error':'revision_validation','detail':getattr(exc,'message_dict',str(exc))},status=400)
    append_audit(request=request,actor=request.user,action='platform.cms_page.rolled_back',object_type='CMSPage',object_id=page.id,before=before,after=_serialize(page),metadata={'revision_id':str(event.id)})
    return JsonResponse(_serialize(page))

@require_http_methods(['GET','POST'])
@platform_permission_required('tenants.view')
def platform_media(request):
    if not request.user.is_superuser: return JsonResponse({'error':'super_admin_required'},status=403)
    if request.method=='GET':
        qs=MediaAsset.objects.all().order_by('-created_at'); q=str(request.GET.get('q') or '').strip(); visibility=str(request.GET.get('visibility') or '').strip()
        if q: qs=qs.filter(Q(title__icontains=q)|Q(alt_text__icontains=q)|Q(file_path__icontains=q))
        if visibility in {'public','private'}: qs=qs.filter(visibility=visibility)
        return JsonResponse({'results':[_serialize(x) for x in qs[:300]]})
    upload=request.FILES.get('file')
    if not upload: return JsonResponse({'error':'file_required'},status=400)
    from .services.file_security import validate_upload,FileValidationError
    try: data_bytes,scan_status=validate_upload(upload)
    except FileValidationError as exc: return JsonResponse({'error':exc.code},status=exc.status)
    decorative=str(request.POST.get('decorative','false')).lower() in {'1','true','yes','on'}; visibility=str(request.POST.get('visibility','private')).strip().lower(); public=visibility=='public'; alt=str(request.POST.get('alt_text') or '').strip()
    if visibility not in {'public','private'}: return JsonResponse({'error':'invalid_visibility'},status=400)
    if public and not decorative and not alt: return JsonResponse({'error':'alt_text_required_for_meaningful_public_image'},status=400)
    if public and scan_status!='clean': return JsonResponse({'error':'public_media_requires_clean_malware_scan'},status=409)
    metadata=build_media_metadata(original_name=upload.name,content_type=upload.content_type,payload=data_bytes,visibility=visibility,decorative=decorative)
    path=default_storage.save(f'platform/media/{uuid.uuid4()}-{metadata["storage_name"]}',ContentFile(data_bytes))
    obj=MediaAsset.objects.create(title=str(request.POST.get('title') or upload.name)[:180],file_path=path,mime_type=metadata['mime_type'],original_name=metadata['original_name'],file_size=metadata['file_size'],sha256=metadata['sha256'],scan_status=scan_status,visibility=visibility,decorative=decorative,alt_text='' if decorative else alt,accessibility_tags=['decorative'] if decorative else [])
    append_audit(request=request,actor=request.user,action='platform.media.created',object_type='MediaAsset',object_id=obj.id,after=_serialize(obj),metadata={'size':obj.file_size,'sha256':obj.sha256,'scan_status':obj.scan_status})
    return JsonResponse(_serialize(obj),status=201)

@require_http_methods(['PATCH','DELETE'])
@platform_permission_required('tenants.view')
def platform_media_detail(request,pk):
    if not request.user.is_superuser: return JsonResponse({'error':'super_admin_required'},status=403)
    try: obj=MediaAsset.objects.get(pk=pk)
    except MediaAsset.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
    before=_serialize(obj)
    if request.method=='DELETE':
        needle=str(obj.id); path=obj.file_path
        referenced=any(needle in json.dumps(p.blocks) or path in json.dumps(p.blocks) for p in CMSPage.objects.only('blocks').all())
        if referenced: return JsonResponse({'error':'media_asset_is_referenced'},status=409)
        if path and default_storage.exists(path): default_storage.delete(path)
        obj.delete(); append_audit(request=request,actor=request.user,action='platform.media.deleted',object_type='MediaAsset',object_id=pk,before=before); return HttpResponse(status=204)
    data=_body(request); decorative=bool(data.get('decorative',obj.decorative)); visibility=str(data.get('visibility') or obj.visibility)
    if visibility not in {'public','private'}: return JsonResponse({'error':'invalid_visibility'},status=400)
    if visibility=='public' and obj.scan_status!='clean': return JsonResponse({'error':'public_media_requires_clean_malware_scan'},status=409)
    alt=str(data.get('alt_text',obj.alt_text) or '').strip()
    if visibility=='public' and not decorative and not alt: return JsonResponse({'error':'alt_text_required_for_meaningful_public_image'},status=400)
    if 'title' in data: obj.title=str(data['title'])[:180]
    obj.alt_text='' if decorative else alt; obj.visibility=visibility; obj.decorative=decorative; obj.accessibility_tags=['decorative'] if decorative else []
    obj.full_clean(); obj.save(); append_audit(request=request,actor=request.user,action='platform.media.updated',object_type='MediaAsset',object_id=obj.id,before=before,after=_serialize(obj)); return JsonResponse(_serialize(obj))

@require_http_methods(['GET'])
@platform_permission_required('tenants.view')
def platform_media_download(request,pk):
    if not request.user.is_superuser: return JsonResponse({'error':'super_admin_required'},status=403)
    try: obj=MediaAsset.objects.get(pk=pk)
    except MediaAsset.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
    if obj.scan_status not in {'clean','not_configured','unavailable_optional'}: return JsonResponse({'error':'media_scan_not_approved'},status=409)
    if not obj.file_path or not default_storage.exists(obj.file_path): return JsonResponse({'error':'media_file_missing'},status=404)
    response=FileResponse(default_storage.open(obj.file_path,'rb'),as_attachment=False,filename=obj.original_name or 'media.bin',content_type=obj.mime_type or 'application/octet-stream')
    response['Cache-Control']='private, no-store'; response['X-Content-Type-Options']='nosniff'; response['X-Media-SHA256']=obj.sha256
    append_audit(request=request,actor=request.user,action='platform.media.downloaded',object_type='MediaAsset',object_id=obj.id,metadata={'sha256':obj.sha256})
    return response

@require_http_methods(['GET'])
def public_media_download(request,pk):
    try: obj=MediaAsset.objects.get(pk=pk)
    except MediaAsset.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
    if not public_download_allowed(obj.visibility,obj.scan_status): return JsonResponse({'error':'not_found'},status=404)
    if not obj.file_path or not default_storage.exists(obj.file_path): return JsonResponse({'error':'not_found'},status=404)
    response=FileResponse(default_storage.open(obj.file_path,'rb'),as_attachment=False,filename=obj.original_name or 'media.bin',content_type=obj.mime_type or 'application/octet-stream')
    response['Cache-Control']='public, max-age=86400, immutable'; response['X-Content-Type-Options']='nosniff'; response['Content-Security-Policy']="default-src 'none'; sandbox"
    return response

@require_http_methods(['GET','POST'])
@platform_permission_required('tenants.view')
def platform_navigation(request):
    if not request.user.is_superuser: return JsonResponse({'error':'super_admin_required'},status=403)
    if request.method=='GET': return JsonResponse({'results':[_serialize(x) for x in NavigationMenu.objects.all()[:500]]})
    if not request.user.is_superuser: return JsonResponse({'error':'platform_write_requires_super_admin'},status=403)
    data=_body(request); locale=str(data.get('locale','en')).strip() or 'en'
    try: items=validate_navigation_items(data.get('items',[]))
    except ValidationError as exc: return JsonResponse({'error':'navigation_validation','detail':exc.message_dict},status=400)
    obj,created=NavigationMenu.objects.update_or_create(key=data['key'],locale=locale,defaults={'items':items}); append_audit(request=request,actor=request.user,action='platform.navigation.saved',object_type='NavigationMenu',object_id=obj.pk,after={'key':obj.key,'item_count':len(items)}); return JsonResponse(_serialize(obj),status=201 if created else 200)

@require_http_methods(['GET','POST'])
@platform_permission_required('tenants.view')
def platform_site_settings(request):
    if not request.user.is_superuser: return JsonResponse({'error':'super_admin_required'},status=403)
    if request.method=='GET': return JsonResponse({'results':[_serialize(x) for x in SiteSetting.objects.all().order_by('key')[:500]]})
    if not request.user.is_superuser: return JsonResponse({'error':'platform_write_requires_super_admin'},status=403)
    data=_body(request); key=str(data.get('key') or '').strip(); value=data.get('value',{}); is_public=bool(data.get('is_public',False))
    if key=='analytics':
        try:value=normalize_analytics_config(value)
        except ValueError as exc:return JsonResponse({'error':str(exc)},status=400)
        is_public=True
    try:
        if key=='public_brand' and isinstance(value,dict) and value.get('logo_url'):
            value={**value,'logo_url':validate_public_url(value['logo_url'],https_required=not django_settings.DEBUG)}
        if key=='site' and isinstance(value,dict) and value.get('base_url'):
            value={**value,'base_url':validate_public_url(value['base_url'],allow_relative=False,https_required=not django_settings.DEBUG)}
        if key=='billing_plans' and isinstance(value,dict):
            value={**value}
            for field in ('success_url','cancel_url','portal_return_url'):
                if value.get(field):value[field]=validate_public_url(value[field],https_required=not django_settings.DEBUG)
    except UnsafeURLError as exc:return JsonResponse({'error':'unsafe_url','detail':str(exc)},status=400)
    obj,_=SiteSetting.objects.update_or_create(key=key,defaults={'value':value,'is_public':is_public}); append_audit(request=request,actor=request.user,action='platform.site_setting.saved',object_type='SiteSetting',object_id=obj.pk,metadata={'key':key,'is_public':is_public}); return JsonResponse(_serialize(obj),status=201)

@require_http_methods(['GET'])
@platform_permission_required('logs.view')
def platform_logs(request):
    kind=request.GET.get('kind','audit')
    if kind=='webhooks': qs=WebhookEvent.objects.all().order_by('-created_at')[:250]
    elif kind=='jobs': qs=SystemJob.objects.all().order_by('-created_at')[:250]
    elif kind=='exports': qs=ExportEvent.objects.all().order_by('-created_at')[:250]
    elif kind=='billing': qs=BillingEvent.objects.all().order_by('-created_at')[:250]
    elif kind=='security': qs=SecurityAlert.objects.select_related('tenant','actor','acknowledged_by').all().order_by('-created_at')[:250]
    else: qs=AuditEvent.objects.all().order_by('-occurred_at')[:250]
    return JsonResponse({'kind':kind,'results':[_serialize(x) for x in qs]})

@require_http_methods(['GET'])
@permission_required('dashboard.view')
def tenant_job_status(request,pk):
    job=SystemJob.objects.filter(pk=pk,tenant=request.tenant).first()
    if not job:return JsonResponse({'error':'not_found'},status=404)
    return JsonResponse({'id':str(job.id),'kind':job.kind,'status':job.status,'attempts':job.attempts,'result':job.result,'last_error':job.last_error,'created_at':job.created_at.isoformat(),'started_at':job.started_at.isoformat() if job.started_at else None,'finished_at':job.finished_at.isoformat() if job.finished_at else None})

@require_http_methods(['POST'])
@platform_permission_required('logs.view')
def retry_system_job(request,pk):
    if not has_recent_step_up(request):return JsonResponse({'error':'step_up_required','reauthenticate':True},status=403)
    job=SystemJob.objects.filter(pk=pk,status='failed').first()
    if not job:return JsonResponse({'error':'failed_job_not_found'},status=404)
    payload=job.payload or {};task=None
    if not retryable_job(job.kind,payload):return JsonResponse({'error':'job_not_safely_retryable','kind':job.kind},status=409)
    try:
        if job.kind=='member_import':
            from .tasks import import_members_job
            task=import_members_job.delay(str(job.id));job.status='queued';job.last_error='';job.finished_at=None;job.save(update_fields=['status','last_error','finished_at','updated_at'])
        elif job.kind=='scheduled_report' and payload.get('schedule_id'):
            from .tasks import run_report_schedule
            task=run_report_schedule.delay(str(payload['schedule_id']))
        elif job.kind=='retention_enforcement':
            from .tasks import enforce_retention_policies
            task=enforce_retention_policies.delay(bool(payload.get('dry_run',False)))
        else:return JsonResponse({'error':'job_not_safely_retryable','kind':job.kind},status=409)
    except Exception as exc:return JsonResponse({'error':'background_queue_unavailable','detail':type(exc).__name__},status=503)
    append_audit(request=request,actor=request.user,tenant=job.tenant,action='platform.job.retry_queued',object_type='SystemJob',object_id=job.id,metadata={'kind':job.kind,'task_id':str(task.id or '')});return JsonResponse({'queued':True,'job_id':str(job.id),'task_id':str(task.id or ''),'kind':job.kind},status=202)

@require_http_methods(['POST'])
@platform_permission_required('logs.view')
def acknowledge_security_alert(request,pk):
    try: alert=SecurityAlert.objects.get(pk=pk)
    except SecurityAlert.DoesNotExist:return JsonResponse({'error':'not_found'},status=404)
    if alert.acknowledged_at:return JsonResponse(_serialize(alert))
    alert.acknowledged_at=timezone.now();alert.acknowledged_by=request.user;alert.save(update_fields=['acknowledged_at','acknowledged_by','updated_at'])
    append_audit(request=request,actor=request.user,tenant=alert.tenant,action='platform.security_alert.acknowledged',object_type='SecurityAlert',object_id=alert.id,metadata={'kind':alert.kind,'severity':alert.severity})
    return JsonResponse(_serialize(alert))

@require_http_methods(['GET','POST'])
@platform_permission_required('billing.view')
def platform_billing(request):
    if request.method=='GET':
        subscriptions=Subscription.objects.select_related('tenant').all().order_by('tenant__name')
        events=BillingEvent.objects.select_related('tenant').order_by('-created_at')[:150]
        webhooks=WebhookEvent.objects.select_related('tenant').order_by('-created_at')[:100]
        requests=PaymentRequest.objects.select_related('tenant').order_by('-created_at')[:100]
        credentials=EncryptedCredential.objects.filter(provider__in=PROVIDER_REQUIRED_FIELDS.keys(),is_enabled=True)
        provider_health=[]
        for provider in PROVIDER_REQUIRED_FIELDS:
            for mode in ('test','live'):
                q=credentials.filter(provider=provider,mode=mode);provider_health.append({'provider':provider,'mode':mode,'configured':q.exists(),'healthy':q.filter(health_status='healthy').exists(),'credential_count':q.count()})
        checkout_policy=(SiteSetting.objects.filter(key='billing_plans').first().value if SiteSetting.objects.filter(key='billing_plans').exists() else {}) or {}
        return JsonResponse({'subscriptions':[_serialize(x) for x in subscriptions],'events':[_serialize(x) for x in events],'webhooks':[_serialize(x) for x in webhooks],'payment_requests':[_serialize(x) for x in requests],'provider_health':provider_health,'checkout_policy':checkout_policy,'summary':{'active_subscriptions':subscriptions.filter(status__in=['active','trialing']).count(),'past_due':subscriptions.filter(status='past_due').count(),'webhook_quarantined':WebhookEvent.objects.filter(status='quarantined').count(),'webhook_failed':WebhookEvent.objects.filter(status='failed').count(),'payment_requests_pending':PaymentRequest.objects.filter(status__in=['created','pending']).count()}})
    if not user_has_platform_permission(request.user,'billing.edit'):return JsonResponse({'error':'platform_forbidden','permission':'billing.edit'},status=403)
    if not has_recent_step_up(request):return JsonResponse({'error':'step_up_required','reauthenticate':True},status=403)
    data=_body(request)
    if data.get('action')=='configure_checkout':
        from urllib.parse import urlsplit
        policy=dict(data.get('policy') or {});clean={}
        for plan in ('annual','monthly','enterprise','trial'):
            item=dict(policy.get(plan) or {});price=str(item.get('stripe_price_id') or '').strip()
            if price:clean[plan]={'stripe_price_id':price[:180]}
        for key in ('success_url','cancel_url','portal_return_url'):
            raw=str(policy.get(key) or '').strip()
            if raw:clean[key]=raw[:500]
        origins=[]
        for raw in policy.get('approved_return_origins',[]) if isinstance(policy.get('approved_return_origins',[]),list) else []:
            parsed=urlsplit(str(raw).strip());origin=f'{parsed.scheme}://{parsed.netloc}'.rstrip('/') if parsed.scheme and parsed.netloc else ''
            if not origin or parsed.path not in ('','/') or parsed.query or parsed.fragment:return JsonResponse({'error':'approved_origin_must_be_scheme_and_host_only'},status=400)
            if parsed.scheme!='https' and not django_settings.DEBUG:return JsonResponse({'error':'approved_origin_requires_https'},status=400)
            if parsed.scheme not in {'http','https'}:return JsonResponse({'error':'invalid_approved_origin_scheme'},status=400)
            origins.append(origin)
        clean['approved_return_origins']=sorted(set(origins));obj,_=SiteSetting.objects.update_or_create(key='billing_plans',defaults={'value':clean,'is_public':False})
        append_audit(request=request,actor=request.user,action='platform.billing.checkout_policy_updated',object_type='SiteSetting',object_id=obj.id,metadata={'plans':sorted(k for k in clean if k in {'annual','monthly','enterprise','trial'}),'approved_return_origins':clean['approved_return_origins']});return JsonResponse({'checkout_policy':clean},status=201)
    tenant=Tenant.objects.filter(pk=data.get('tenant_id')).first()
    if not tenant:return JsonResponse({'error':'tenant_not_found'},status=404)
    provider=str(data.get('provider','stripe')).lower();status=str(data.get('status','active'));plan=str(data.get('plan','annual'))
    if provider!='stripe':return JsonResponse({'error':'subscription_provider_must_be_stripe'},status=400)
    if status not in {'active','trialing','past_due','cancelled','paused','incomplete','pending'}:return JsonResponse({'error':'invalid_subscription_status'},status=400)
    before=_serialize(Subscription.objects.filter(tenant=tenant).first()) if Subscription.objects.filter(tenant=tenant).exists() else {}
    defaults={k:data[k] for k in ['provider_customer_id','provider_subscription_id','current_period_end','trial_ends_at'] if k in data};defaults.update({'provider':provider,'plan':plan,'status':status,'last_synced_at':timezone.now()})
    obj,_=Subscription.objects.update_or_create(tenant=tenant,defaults=defaults)
    append_audit(request=request,actor=request.user,tenant=tenant,action='platform.subscription.updated',object_type='Subscription',object_id=obj.pk,before=before,after={'provider':provider,'plan':obj.plan,'status':obj.status,'provider_customer_id':obj.provider_customer_id,'provider_subscription_id':obj.provider_subscription_id});return JsonResponse(_serialize(obj),status=201)


@require_http_methods(['POST'])
@platform_permission_required('impersonation.create')
def start_impersonation(request):
    if not has_recent_step_up(request):return JsonResponse({'error':'step_up_required','reauthenticate':True},status=403)
    data=_body(request); reason=(data.get('reason') or '').strip()
    if len(reason)<10: return JsonResponse({'error':'reason_required_min_10_chars'},status=400)
    tenant=Tenant.objects.filter(pk=data.get('tenant_id'),is_active=True).first(); target=get_user_model().objects.filter(pk=data.get('target_user_id'),is_active=True).first()
    if not tenant or not target:return JsonResponse({'error':'tenant_or_target_not_found'},status=404)
    if target.id==request.user.id:return JsonResponse({'error':'cannot_impersonate_self'},status=400)
    if not TenantMembership.objects.filter(tenant=tenant,user=target,is_active=True).exists(): return JsonResponse({'error':'target_not_in_tenant'},status=400)
    try: minutes=max(5,min(int(data.get('duration_minutes',30)),60))
    except (TypeError,ValueError):return JsonResponse({'error':'invalid_duration'},status=400)
    # End any stale/parallel session by this operator before opening a new boundary.
    ImpersonationSession.objects.filter(actor=request.user,is_active=True,ended_at__isnull=True).update(is_active=False,ended_at=timezone.now())
    imp=ImpersonationSession.objects.create(actor=request.user,target_user=target,tenant=tenant,reason=reason,expires_at=timezone.now()+timezone.timedelta(minutes=minutes))
    request.session['impersonation_id']=str(imp.id); request.session['impersonated_user_id']=target.id
    security_alert(kind='impersonation_started',message=f'Platform impersonation started for tenant {tenant.name}',tenant=tenant,actor=request.user,severity='high',metadata={'target_user_id':target.id,'session_id':str(imp.id),'expires_at':imp.expires_at.isoformat()})
    append_audit(request=request,actor=request.user,tenant=tenant,action='platform.impersonation.started',object_type='ImpersonationSession',object_id=imp.pk,metadata={'target_user_id':target.id,'reason':reason,'expires_at':imp.expires_at.isoformat()}); return JsonResponse({'id':str(imp.id),'active':True,'expires_at':imp.expires_at.isoformat()})

@require_http_methods(['POST'])
@login_required
def stop_impersonation(request):
    imp_id=request.session.get('impersonation_id')
    if not imp_id: return JsonResponse({'error':'no_active_impersonation'},status=400)
    actor=getattr(request,'real_user',request.user)
    try: imp=ImpersonationSession.objects.get(pk=imp_id,actor=actor,is_active=True)
    except ImpersonationSession.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
    imp.is_active=False; imp.ended_at=timezone.now(); imp.save(update_fields=['is_active','ended_at','updated_at']); request.session.pop('impersonation_id',None); request.session.pop('impersonated_user_id',None)
    append_audit(request=request,actor=actor,tenant=imp.tenant,action='platform.impersonation.ended',object_type='ImpersonationSession',object_id=imp.pk); return JsonResponse({'active':False})


@require_http_methods(['GET','POST'])
def public_lead_form(request,key):
    try: form=LeadForm.objects.get(key=key,is_active=True)
    except LeadForm.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
    if request.method=='GET': return JsonResponse({'key':form.key,'name':form.name,'fields':form.fields})
    client_key=getattr(request,'ip_hash','') or request.META.get('REMOTE_ADDR','unknown')
    if not rate_limit_allow(f'lead:{form.key}:{client_key}',limit=10,window=3600,critical=True): return JsonResponse({'error':'too_many_submissions'},status=429)
    data=_body(request)
    try: clean=validate_lead_submission(form.fields,data)
    except ValidationError as exc: return JsonResponse({'error':'validation_error','fields':exc.message_dict},status=400)
    sub=LeadSubmission.objects.create(form=form,data=clean,source=request.headers.get('Referer','')[:240])
    routing=form.routing or {}; emails=[str(x).strip() for x in (routing.get('emails') or []) if '@' in str(x)]
    if emails:
        try: send_mail(f'New {form.name} submission',json.dumps(clean,indent=2,ensure_ascii=False),django_settings.DEFAULT_FROM_EMAIL,emails,fail_silently=True)
        except Exception: pass
    return JsonResponse({'ok':True,'submission_id':str(sub.id)},status=201)

from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from .services.pdfs import minutes_pdf, receipt_pdf


def _persist_generated_document(request,kind,source_type,source_id,data,template_key='',template_version='1'):
    import hashlib
    digest=hashlib.sha256(data).hexdigest();name=f'tenants/{request.tenant.id}/generated/documents/{kind}-{source_id}-{digest[:12]}.pdf';path=default_storage.save(name,ContentFile(data))
    return GeneratedDocument.objects.create(tenant=request.tenant,kind=kind,source_type=source_type,source_id=str(source_id),template_key=template_key,template_version=str(template_version)[:40] or '1',file_path=path,sha256=digest,generated_by=request.user,metadata={'size_bytes':len(data),'mime_type':'application/pdf','storage':'private'})

@require_http_methods(['POST'])
@permission_required('documents.create')
def document_upload(request):
    upload=request.FILES.get('file')
    if not upload: return JsonResponse({'error':'file_required'},status=400)
    from .services.file_security import validate_upload,FileValidationError,safe_storage_name
    import hashlib
    try: data_bytes,scan_status=validate_upload(upload)
    except FileValidationError as exc: return JsonResponse({'error':exc.code},status=exc.status)
    safe_name=safe_storage_name(upload.name)
    try:
        tags=json.loads(request.POST.get('tags','[]') or '[]'); permissions=json.loads(request.POST.get('permissions','{}') or '{}')
    except json.JSONDecodeError: return JsonResponse({'error':'invalid_document_metadata_json'},status=400)
    if not isinstance(tags,list) or not isinstance(permissions,dict): return JsonResponse({'error':'invalid_document_metadata'},status=400)
    digest=hashlib.sha256(data_bytes).hexdigest();path=default_storage.save(f'tenants/{request.tenant.id}/documents/{digest[:16]}-{safe_name}',ContentFile(data_bytes))
    try:
        obj=Document.objects.create(tenant=request.tenant,title=(request.POST.get('title') or upload.name)[:240],folder=request.POST.get('folder','')[:240],tags=tags,file_path=path,mime_type=upload.content_type,original_name=upload.name[:255],file_size=len(data_bytes),sha256=digest,scan_status=scan_status,permissions=permissions,created_by=request.user)
    except Exception:
        default_storage.delete(path); raise
    append_audit(request=request,action='documents.uploaded',object_type='Document',object_id=obj.pk,after={'mime_type':upload.content_type,'size':len(data_bytes),'sha256':digest,'scan_status':scan_status})
    return JsonResponse(_serialize(obj),status=201)

@require_http_methods(['GET'])
@permission_required('meetings.view')
def meeting_pdf(request,meeting_id):
    try: meeting=Meeting.objects.get(pk=meeting_id,tenant=request.tenant)
    except Meeting.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
    minute=MinuteVersion.objects.filter(meeting=meeting,tenant=request.tenant).order_by('-version').first()
    if not minute: return JsonResponse({'error':'minutes_not_found'},status=404)
    data=minutes_pdf(meeting,minute,meeting.attendance.select_related('member').all(),meeting.visitors.all(),request.tenant.name); doc=_persist_generated_document(request,'meeting_minutes','Meeting',meeting.id,data,'meeting-minutes-v1')
    response=HttpResponse(data,content_type='application/pdf'); response['Content-Disposition']=f'attachment; filename="meeting-{meeting.id}-minutes.pdf"'; response['X-Document-SHA256']=doc.sha256
    append_audit(request=request,action='meetings.minutes.pdf_generated',object_type='Meeting',object_id=meeting.id); return response

@require_http_methods(['GET'])
@permission_required('dues.view')
def payment_receipt(request,payment_id):
    try: payment=Payment.objects.select_related('member').get(pk=payment_id,tenant=request.tenant)
    except Payment.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
    data=receipt_pdf(payment,request.tenant.name); doc=_persist_generated_document(request,'receipt','Payment',payment.id,data,'receipt-v1'); response=HttpResponse(data,content_type='application/pdf'); response['Content-Disposition']=f'attachment; filename="receipt-{payment.receipt_number}.pdf"'; response['X-Document-SHA256']=doc.sha256
    append_audit(request=request,action='dues.receipt.generated',object_type='Payment',object_id=payment.id); return response

@require_http_methods(['GET'])
@permission_required('communications.view')
def campaign_recipients(request,pk):
    try: c=CommunicationCampaign.objects.get(pk=pk,tenant=request.tenant)
    except CommunicationCampaign.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
    qs=campaign_members(c); limit=50
    preview=[{'member_id':str(m.id),'name':f'{m.first_name} {m.last_name}'.strip(),'email':m.email} for m in qs[:limit]]
    return JsonResponse({'count':qs.count(),'preview':preview,'truncated':qs.count()>limit,'segment':c.segment or {}})

@require_http_methods(['GET'])
@permission_required('communications.view')
def campaign_delivery_logs(request,pk):
    try: c=CommunicationCampaign.objects.get(pk=pk,tenant=request.tenant)
    except CommunicationCampaign.DoesNotExist:return JsonResponse({'error':'not_found'},status=404)
    qs=c.deliveries.order_by('-updated_at')[:500]
    return JsonResponse({'campaign_id':str(c.id),'status':c.status,'results':[{'id':str(x.id),'member_id':str(x.member_id) if x.member_id else None,'recipient':x.recipient,'status':x.status,'attempts':x.attempts,'provider_id':x.provider_id,'error':x.error,'last_attempt_at':x.last_attempt_at.isoformat() if x.last_attempt_at else None,'delivered_at':x.delivered_at.isoformat() if x.delivered_at else None} for x in qs]})

@require_http_methods(['POST'])
@permission_required('communications.edit')
def campaign_send(request,pk):
    from .tasks import send_campaign
    with transaction.atomic():
        try: c=CommunicationCampaign.objects.select_for_update().get(pk=pk,tenant=request.tenant)
        except CommunicationCampaign.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
        if c.status not in ('draft','failed'): return JsonResponse({'error':'campaign_not_sendable','status':c.status},status=409)
        recipient_count=campaign_members(c).count()
        if recipient_count<=0: return JsonResponse({'error':'campaign_has_no_recipients'},status=409)
        future=bool(c.scheduled_for and c.scheduled_for>timezone.now()); c.status='scheduled' if future else 'sending'; c.save(update_fields=['status','updated_at'])
    if future:
        append_audit(request=request,action='communications.scheduled',object_type='CommunicationCampaign',object_id=c.id,metadata={'recipient_count':recipient_count,'scheduled_for':c.scheduled_for.isoformat()});return JsonResponse({'queued':False,'scheduled':True,'task_id':None,'recipient_count':recipient_count,'status':c.status,'scheduled_for':c.scheduled_for.isoformat()})
    try:result=send_campaign.delay(str(c.id))
    except Exception as exc:
        CommunicationCampaign.objects.filter(pk=c.id,status='sending').update(status='failed',updated_at=timezone.now());append_audit(request=request,action='communications.queue_failed',object_type='CommunicationCampaign',object_id=c.id,metadata={'error':type(exc).__name__,'recipient_count':recipient_count});return JsonResponse({'error':'background_queue_unavailable'},status=503)
    append_audit(request=request,action='communications.queued',object_type='CommunicationCampaign',object_id=c.id,metadata={'task_id':result.id,'recipient_count':recipient_count,'scheduled':False}); return JsonResponse({'queued':True,'task_id':result.id,'recipient_count':recipient_count,'status':c.status})

@require_http_methods(['POST'])
@permission_required('communications.edit')
def campaign_cancel(request,pk):
    with transaction.atomic():
        try: c=CommunicationCampaign.objects.select_for_update().get(pk=pk,tenant=request.tenant)
        except CommunicationCampaign.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
        if c.status not in {'draft','scheduled'}: return JsonResponse({'error':'campaign_not_cancellable','status':c.status},status=409)
        before={'status':c.status}; c.status='cancelled'; c.save(update_fields=['status','updated_at'])
    append_audit(request=request,action='communications.cancelled',object_type='CommunicationCampaign',object_id=c.id,before=before,after={'status':'cancelled'})
    return JsonResponse(_serialize(c))

@require_http_methods(['GET','POST'])
def report_schedules(request):
    if request.method=='GET':
        if not has_permission(request.tenant_membership.role,'reports.view',effective_custom_permissions(request.tenant_membership)): return JsonResponse({'error':'forbidden'},status=403)
        return JsonResponse({'results':[_serialize(x) for x in ReportSchedule.objects.filter(tenant=request.tenant).order_by('-created_at')]})
    if not has_permission(request.tenant_membership.role,'reports.admin',effective_custom_permissions(request.tenant_membership)): return JsonResponse({'error':'forbidden'},status=403)
    data=_body(request); fmt=str(data.get('format','pdf')).lower()
    try:
        report_type=canonical_report_id(str(data.get('report_type',''))); definition=REPORT_CATALOG[report_type]
    except ValueError as exc:return JsonResponse({'error':str(exc)},status=400)
    if definition.get('platform') or not definition.get('schedulable'): return JsonResponse({'error':'report_not_schedulable'},status=400)
    if fmt not in FORMATS or fmt not in definition.get('formats',[]): return JsonResponse({'error':'unsupported_format','allowed':definition.get('formats',[])},status=400)
    try: recipients=validate_report_recipients(request.tenant,data.get('recipients',[]))
    except ValidationError as exc: return JsonResponse({'error':'recipient_policy','detail':exc.message_dict},status=400)
    schedule=str(data.get('schedule') or '').strip()
    try:
        from .services.scheduling import validate_cron,next_cron_run
        if not validate_cron(schedule):raise ValueError('Cron expression must contain five valid fields.')
        next_run=next_cron_run(schedule,request.tenant.timezone or 'UTC')
        if next_run is None:raise ValueError('Cron expression does not produce a run time within one year.')
    except ValueError as exc:return JsonResponse({'error':'invalid_schedule','detail':str(exc)},status=400)
    attachments=[]
    for item in data.get('attachments') or []:
        try: attachment_id=canonical_report_id(str(item))
        except ValueError as exc:return JsonResponse({'error':'invalid_attachment_report','detail':str(exc)},status=400)
        if REPORT_CATALOG[attachment_id].get('platform'):return JsonResponse({'error':'platform_attachment_not_allowed'},status=400)
        if attachment_id not in attachments:attachments.append(attachment_id)
    obj=ReportSchedule.objects.create(tenant=request.tenant,report_type=report_type,format=fmt,schedule=schedule,recipients=recipients,filters=data.get('filters',{}),attachments=attachments,is_active=bool(data.get('is_active',True)),created_by=request.user,next_run_at=next_run)
    append_audit(request=request,action='reports.schedule.created',object_type='ReportSchedule',object_id=obj.id,after=_serialize(obj)); return JsonResponse(_serialize(obj),status=201)


@require_http_methods(['PATCH','DELETE'])
def report_schedule_detail(request,pk):
    if not has_permission(request.tenant_membership.role,'reports.admin',effective_custom_permissions(request.tenant_membership)): return JsonResponse({'error':'forbidden'},status=403)
    try: obj=ReportSchedule.objects.get(pk=pk,tenant=request.tenant)
    except ReportSchedule.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
    conflict=_offline_version_conflict(request,obj)
    if conflict: return conflict
    before=_serialize(obj)
    if request.method=='DELETE': obj.delete(); append_audit(request=request,action='reports.schedule.deleted',object_type='ReportSchedule',object_id=pk,before=before); return HttpResponse(status=204)
    data=_body(request)
    for key in ['format','filters','is_active']:
        if key in data: setattr(obj,key,data[key])
    if 'report_type' in data:
        try: obj.report_type=canonical_report_id(str(data['report_type']))
        except ValueError as exc:return JsonResponse({'error':str(exc)},status=400)
    if 'schedule' in data:obj.schedule=str(data.get('schedule') or '').strip()
    if 'recipients' in data:
        try: obj.recipients=validate_report_recipients(request.tenant,data.get('recipients'))
        except ValidationError as exc: return JsonResponse({'error':'recipient_policy','detail':exc.message_dict},status=400)
    if 'attachments' in data:
        attachments=[]
        for item in data.get('attachments') or []:
            try: rid=canonical_report_id(str(item))
            except ValueError as exc:return JsonResponse({'error':'invalid_attachment_report','detail':str(exc)},status=400)
            if REPORT_CATALOG[rid].get('platform'):return JsonResponse({'error':'platform_attachment_not_allowed'},status=400)
            if rid not in attachments:attachments.append(rid)
        obj.attachments=attachments
    definition=REPORT_CATALOG.get(obj.report_type,{})
    if definition.get('platform') or not definition.get('schedulable'):return JsonResponse({'error':'report_not_schedulable'},status=400)
    if obj.format not in FORMATS or obj.format not in definition.get('formats',[]):return JsonResponse({'error':'unsupported_format','allowed':definition.get('formats',[])},status=400)
    try:
        from .services.scheduling import validate_cron,next_cron_run
        validate_cron(obj.schedule);obj.next_run_at=next_cron_run(obj.schedule,request.tenant.timezone or 'UTC') if obj.is_active else None
    except ValueError as exc:return JsonResponse({'error':'invalid_schedule','detail':str(exc)},status=400)
    obj.full_clean(); obj.save(); append_audit(request=request,action='reports.schedule.updated',object_type='ReportSchedule',object_id=obj.pk,before=before,after=_serialize(obj)); return JsonResponse(_serialize(obj))


@require_http_methods(['POST'])
def report_schedule_run(request,pk):
    if not has_permission(request.tenant_membership.role,'reports.admin',effective_custom_permissions(request.tenant_membership)):return JsonResponse({'error':'forbidden'},status=403)
    obj=ReportSchedule.objects.filter(pk=pk,tenant=request.tenant,is_active=True).first()
    if not obj:return JsonResponse({'error':'not_found_or_disabled'},status=404)
    try:
        from .tasks import run_report_schedule
        task=run_report_schedule.delay(str(obj.id))
    except Exception as exc:return JsonResponse({'error':'background_queue_unavailable','detail':type(exc).__name__},status=503)
    append_audit(request=request,action='reports.schedule.manual_run',object_type='ReportSchedule',object_id=obj.id,metadata={'task_id':task.id})
    return JsonResponse({'queued':True,'task_id':task.id},status=202)


@require_http_methods(['GET','POST'])
def email_templates(request):
    if request.method=='GET':
        if not has_permission(request.tenant_membership.role,'communications.view',effective_custom_permissions(request.tenant_membership)): return JsonResponse({'error':'forbidden'},status=403)
        qs=EmailTemplate.objects.filter(Q(tenant=request.tenant)|Q(tenant=None),is_active=True).order_by('key','-tenant_id'); return JsonResponse({'results':[_serialize(x) for x in qs]})
    if not has_permission(request.tenant_membership.role,'communications.admin',effective_custom_permissions(request.tenant_membership)): return JsonResponse({'error':'forbidden'},status=403)
    data=_body(request);key=str(data.get('key') or '').strip();name=str(data.get('name') or '').strip();subject=str(data.get('subject') or '');body=str(data.get('body') or '')
    if not key or not name or not subject or not body:return JsonResponse({'error':'template_key_name_subject_body_required'},status=400)
    from .services.template_rendering import validate_merge_fields,TemplateRenderError
    try:
        merge_fields=validate_merge_fields(data.get('merge_fields') or [],subject,body)
        obj=EmailTemplate.objects.filter(tenant=request.tenant,key=key).first() or EmailTemplate(tenant=request.tenant,key=key)
        obj.name=name;obj.subject=subject;obj.body=body;obj.merge_fields=merge_fields;obj.is_active=bool(data.get('is_active',True));obj.full_clean();obj.save()
    except TemplateRenderError as exc: return JsonResponse({'error':exc.code,'fields':exc.fields},status=400)
    except ValidationError as exc:return JsonResponse({'error':'template_validation','detail':getattr(exc,'message_dict',str(exc))},status=400)
    append_audit(request=request,action='communications.template.saved',object_type='EmailTemplate',object_id=obj.id,metadata={'merge_fields':merge_fields}); return JsonResponse(_serialize(obj),status=201)


def _email_template_for_request(request,pk):
    return EmailTemplate.objects.filter(pk=pk,is_active=True).filter(Q(tenant=request.tenant)|Q(tenant=None)).first()

def _template_preview_context(request):
    now=timezone.now();return {'member.first_name':'Alex','member.last_name':'Member','member.email':request.user.email or 'alex@example.invalid','member.phone':'','member.address':'','member.status':'active','member.degree':'MM','member.joined_on':'2026-01-01','lodge.name':request.tenant.name,'lodge.number':request.tenant.lodge_number,'lodge.jurisdiction':request.tenant.jurisdiction,'lodge.locale':request.tenant.locale,'document.generated_date':now.date().isoformat(),'document.generated_at':now.isoformat()}

@require_http_methods(['POST'])
@permission_required('communications.view')
def email_template_preview(request,pk):
    template=_email_template_for_request(request,pk)
    if not template:return JsonResponse({'error':'not_found'},status=404)
    from .services.template_rendering import render_template,TemplateRenderError
    context={**_template_preview_context(request),**(_body(request).get('context') or {})}
    try: subject=render_template(template.subject,context,template.merge_fields);body=render_template(template.body,context,template.merge_fields)
    except TemplateRenderError as exc:return JsonResponse({'error':exc.code,'fields':exc.fields},status=400)
    return JsonResponse({'id':str(template.id),'subject':subject,'body':body,'merge_fields':template.merge_fields})

@require_http_methods(['POST'])
@permission_required('communications.edit')
def email_template_test_send(request,pk):
    template=_email_template_for_request(request,pk)
    if not template:return JsonResponse({'error':'not_found'},status=404)
    data=_body(request);recipient=str(data.get('recipient') or request.user.email or '').strip().lower()
    allowed={str(request.user.email or '').strip().lower()}
    User=get_user_model();allowed.update(x.lower() for x in User.objects.filter(tenant_memberships__tenant=request.tenant,tenant_memberships__is_active=True,is_active=True).exclude(email='').values_list('email',flat=True))
    allowed.discard('')
    if not recipient or recipient not in allowed:return JsonResponse({'error':'test_recipient_must_be_authorized_tenant_user'},status=403)
    from .services.template_rendering import render_template,TemplateRenderError
    context={**_template_preview_context(request),**(data.get('context') or {})}
    try: subject=render_template(template.subject,context,template.merge_fields);body=render_template(template.body,context,template.merge_fields)
    except TemplateRenderError as exc:return JsonResponse({'error':exc.code,'fields':exc.fields},status=400)
    from django.core.mail import EmailMessage
    message_id=f'<lodgeflow-template-{template.id}-{uuid.uuid4().hex}@{request.tenant.slug}.invalid>'
    try: EmailMessage(subject=f'[TEST] {subject}',body=body,from_email=django_settings.DEFAULT_FROM_EMAIL,to=[recipient],headers={'Message-ID':message_id,'X-LodgeFlow-Template-Test':str(template.id)}).send(fail_silently=False)
    except Exception as exc:return JsonResponse({'error':'test_send_failed','detail':type(exc).__name__},status=502)
    append_audit(request=request,action='communications.template.test_sent',object_type='EmailTemplate',object_id=template.id,metadata={'recipient_user_scoped':True,'message_id':message_id})
    return JsonResponse({'ok':True,'message_id':message_id})

MILESTONE_FIELDS=['member','kind','title','milestone_date','notes','recurring_yearly']
PARTICIPATION_FIELDS=['member','meeting','kind','points','notes']
VISITOR_FIELDS=['meeting','name','home_lodge','lodge_number','jurisdiction','email','notes']
def milestones(request): return _crud_list_create(request,Milestone,MILESTONE_FIELDS,'members',['kind','member_id'],'milestone_date')
def milestone_detail(request,pk): return _crud_detail(request,Milestone,pk,MILESTONE_FIELDS,'members')
def participation_records(request): return _crud_list_create(request,ParticipationRecord,PARTICIPATION_FIELDS,'meetings',['kind','member_id','meeting_id'])
def participation_detail(request,pk): return _crud_detail(request,ParticipationRecord,pk,PARTICIPATION_FIELDS,'meetings')
def visitors(request): return _crud_list_create(request,Visitor,VISITOR_FIELDS,'meetings',['home_lodge','meeting_id'])
def visitor_detail(request,pk): return _crud_detail(request,Visitor,pk,VISITOR_FIELDS,'meetings')

@require_http_methods(['GET'])
@permission_required('reports.view')
def insights(request):
    from .services.insights import all_insights
    return JsonResponse(all_insights(request.tenant))

@require_http_methods(['GET'])
@permission_required('dues.view')
def dues_arrears(request):
    today=timezone.now().date();rows=[];totals={}
    qs=DuesAssessment.objects.filter(tenant=request.tenant).select_related('member','cycle').exclude(status__in=['paid','waived']).order_by('cycle__due_on','member__last_name')[:50000]
    for a in qs:
        due=amount_due(a.amount,a.paid_amount,a.waived_amount,a.late_fee,a.adjustment_amount)
        if due<=0:continue
        raw_days=(today-a.cycle.due_on).days;days=max(0,raw_days);bucket='current' if raw_days<=0 else ('1-30' if days<=30 else '31-60' if days<=60 else '61-90' if days<=90 else '90+')
        totals[a.currency]=totals.get(a.currency,Decimal('0'))+due
        rows.append({'assessment_id':str(a.id),'member_id':str(a.member_id),'member':f'{a.member.first_name} {a.member.last_name}','cycle':a.cycle.name,'due_on':a.cycle.due_on.isoformat(),'days_overdue':days,'aging_bucket':bucket,'currency':a.currency,'outstanding':str(due),'status':a.status})
    return JsonResponse({'count':len(rows),'total':str(next(iter(totals.values()))) if len(totals)==1 else None,'totals_by_currency':{k:str(v) for k,v in sorted(totals.items())},'results':rows[:5000]})

@require_http_methods(['POST'])
@permission_required('dues.admin')
def generate_dues_assessments(request,cycle_id):
    from .domain.finance import prorated_dues
    try: cycle=DuesCycle.objects.get(pk=cycle_id,tenant=request.tenant,status='active')
    except DuesCycle.DoesNotExist: return JsonResponse({'error':'cycle_not_active'},status=409)
    created=existing=0
    with transaction.atomic():
        for member in Member.objects.filter(tenant=request.tenant,status='active'):
            base=Decimal('0') if member.is_dues_exempt else (member.custom_dues_amount if member.custom_dues_amount is not None else cycle.standard_amount)
            amount=prorated_dues(base,cycle.period_start,cycle.period_end,member.joined_on,cycle.proration_enabled)
            obj,is_created=DuesAssessment.objects.get_or_create(tenant=request.tenant,member=member,cycle=cycle,defaults={'amount':amount,'late_fee':Decimal('0'),'adjustment_amount':Decimal('0'),'currency':cycle.currency,'status':'waived' if member.is_dues_exempt else 'open','waived_amount':amount if member.is_dues_exempt else Decimal('0')})
            created+=int(is_created);existing+=int(not is_created)
    append_audit(request=request,action='dues.assessments.generated',object_type='DuesCycle',object_id=cycle.id,metadata={'created':created,'existing':existing,'idempotent':True})
    return JsonResponse({'created':created,'existing':existing})

@require_http_methods(['POST'])
@permission_required('dues.admin')
def apply_late_fees(request,cycle_id):
    try: cycle=DuesCycle.objects.get(pk=cycle_id,tenant=request.tenant,status='active')
    except DuesCycle.DoesNotExist: return JsonResponse({'error':'cycle_not_active'},status=409)
    if timezone.now().date()<=cycle.due_on: return JsonResponse({'updated':0,'reason':'not_past_due'})
    updated=0
    for a in cycle.assessments.filter(tenant=request.tenant).exclude(status__in=['paid','waived']):
        a.late_fee=cycle.late_fee; a.status=assessment_status(a.amount,a.paid_amount,a.waived_amount,a.late_fee,True,a.adjustment_amount); a.save(update_fields=['late_fee','status','updated_at']); updated+=1
    append_audit(request=request,action='dues.late_fees.applied',object_type='DuesCycle',object_id=cycle.id,metadata={'updated':updated}); return JsonResponse({'updated':updated})

@require_http_methods(['POST'])
@permission_required('dues.edit')
def waive_dues_assessment(request,pk):
    data=_body(request);reason=str(data.get('reason') or '').strip()
    if len(reason)<5:return JsonResponse({'error':'waiver_reason_required'},status=400)
    with transaction.atomic():
        try: obj=DuesAssessment.objects.select_for_update().select_related('cycle').get(pk=pk,tenant=request.tenant)
        except DuesAssessment.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
        try: amount=Decimal(str(data.get('amount',amount_due(obj.amount,obj.paid_amount,obj.waived_amount,obj.late_fee,obj.adjustment_amount))))
        except Exception: return JsonResponse({'error':'invalid_waiver_amount'},status=400)
        if amount<=0:return JsonResponse({'error':'waiver_must_be_positive'},status=400)
        maximum=amount_due(obj.amount,obj.paid_amount,obj.waived_amount,obj.late_fee,obj.adjustment_amount)
        if amount>maximum:return JsonResponse({'error':'waiver_exceeds_balance','balance':str(maximum)},status=409)
        before=_serialize(obj);obj.waived_amount+=amount;obj.status=assessment_status(obj.amount,obj.paid_amount,obj.waived_amount,obj.late_fee,timezone.now().date()>obj.cycle.due_on,obj.adjustment_amount);obj.save(update_fields=['waived_amount','status','updated_at'])
        adj=FinancialAdjustment.objects.create(tenant=request.tenant,assessment=obj,kind='waiver',amount=amount,reason=reason,created_by=request.user)
    append_audit(request=request,action='dues.assessment.waived',object_type='FinancialAdjustment',object_id=adj.id,before=before,after=_serialize(obj),metadata={'assessment_id':str(obj.id),'waiver_amount':str(amount),'reason':reason[:500]});payload=_serialize(obj);payload['adjustment_record_id']=str(adj.id);return JsonResponse(payload)

@require_http_methods(['POST'])
@permission_required('dues.edit')
def adjust_dues_assessment(request,pk):
    data=_body(request);kind=str(data.get('kind') or '').lower();reason=str(data.get('reason') or '').strip()
    if kind not in {'charge','credit'}:return JsonResponse({'error':'adjustment_kind_must_be_charge_or_credit'},status=400)
    if len(reason)<5:return JsonResponse({'error':'adjustment_reason_required'},status=400)
    try:amount=Decimal(str(data.get('amount','0')))
    except Exception:return JsonResponse({'error':'invalid_adjustment_amount'},status=400)
    if amount<=0:return JsonResponse({'error':'adjustment_must_be_positive'},status=400)
    with transaction.atomic():
        obj=DuesAssessment.objects.select_for_update().select_related('cycle').filter(pk=pk,tenant=request.tenant).first()
        if not obj:return JsonResponse({'error':'not_found'},status=404)
        delta=amount if kind=='charge' else -amount;new_adjustment=obj.adjustment_amount+delta
        gross=Decimal(obj.amount)+Decimal(obj.late_fee)+new_adjustment
        if gross<Decimal(obj.paid_amount)+Decimal(obj.waived_amount):return JsonResponse({'error':'credit_would_exceed_settled_amount'},status=409)
        before=_serialize(obj);obj.adjustment_amount=new_adjustment;obj.status=assessment_status(obj.amount,obj.paid_amount,obj.waived_amount,obj.late_fee,timezone.now().date()>obj.cycle.due_on,obj.adjustment_amount);obj.save(update_fields=['adjustment_amount','status','updated_at'])
        adj=FinancialAdjustment.objects.create(tenant=request.tenant,assessment=obj,kind=kind,amount=amount,reason=reason,created_by=request.user)
    append_audit(request=request,action=f'dues.assessment.{kind}_adjustment',object_type='FinancialAdjustment',object_id=adj.id,before=before,after=_serialize(obj),metadata={'assessment_id':str(obj.id),'reason':reason[:500]});return JsonResponse({'assessment':_serialize(obj),'adjustment':_serialize(adj)},status=201)

@require_http_methods(['POST'])
@permission_required('dues.admin')
def reverse_payment(request,payment_id):
    reason=str(_body(request).get('reason') or '').strip()
    if len(reason)<5:return JsonResponse({'error':'reversal_reason_required'},status=400)
    with transaction.atomic():
        payment=Payment.objects.select_for_update().select_related('assessment__cycle').filter(pk=payment_id,tenant=request.tenant).first()
        if not payment:return JsonResponse({'error':'not_found'},status=404)
        if payment.is_reversed:return JsonResponse({'error':'payment_already_reversed'},status=409)
        if not payment.assessment_id:return JsonResponse({'error':'unallocated_payment_requires_manual_reconciliation'},status=409)
        assessment=DuesAssessment.objects.select_for_update().get(pk=payment.assessment_id,tenant=request.tenant)
        if assessment.paid_amount<payment.amount:return JsonResponse({'error':'assessment_payment_total_inconsistent'},status=409)
        assessment.paid_amount-=payment.amount;assessment.status=assessment_status(assessment.amount,assessment.paid_amount,assessment.waived_amount,assessment.late_fee,timezone.now().date()>assessment.cycle.due_on,assessment.adjustment_amount);assessment.save(update_fields=['paid_amount','status','updated_at'])
        payment.is_reversed=True;payment.reversed_at=timezone.now();payment.reversed_by=request.user;payment.reversal_reason=reason;payment.save(update_fields=['is_reversed','reversed_at','reversed_by','reversal_reason','updated_at'])
        adj=FinancialAdjustment.objects.create(tenant=request.tenant,assessment=assessment,kind='payment_reversal',amount=payment.amount,reason=reason,created_by=request.user,related_payment=payment)
    append_audit(request=request,action='dues.payment.reversed',object_type='Payment',object_id=payment.id,after={'is_reversed':True,'assessment_status':assessment.status},metadata={'reason':reason[:500],'adjustment_id':str(adj.id)});return JsonResponse({'payment':_serialize(payment),'assessment':_serialize(assessment)})

# ---- Production hardening endpoints -------------------------------------------------
def _document_allowed(request,obj):
    rules=obj.permissions or {}
    if not rules: return True
    if not isinstance(rules,dict): return False
    if request.tenant_membership.role in ('owner','admin'): return True
    allowed_roles=set(rules.get('roles') or [])
    allowed_users={str(x) for x in (rules.get('users') or [])}
    return request.tenant_membership.role in allowed_roles or str(request.user.id) in allowed_users

@require_http_methods(['GET','POST'])
def documents(request):
    permission='documents.view' if request.method=='GET' else 'documents.create'
    if not has_permission(request.tenant_membership.role,permission,effective_custom_permissions(request.tenant_membership)): return JsonResponse({'error':'forbidden'},status=403)
    if request.method=='GET':
        qs=Document.objects.filter(tenant=request.tenant).order_by('-created_at')
        folder=request.GET.get('folder'); q=request.GET.get('q')
        if folder: qs=qs.filter(folder=folder)
        if q: qs=qs.filter(Q(title__icontains=q)|Q(tags__icontains=q))
        visible=[x for x in qs[:500] if _document_allowed(request,x)]
        return JsonResponse({'count':len(visible),'results':[_serialize(x) for x in visible]})
    data=_body(request); obj=Document(tenant=request.tenant,created_by=request.user); _apply(obj,data,DOCUMENT_FIELDS,request.tenant); obj.tenant=request.tenant; obj.created_by=request.user
    if obj.template_key or '{{' in (obj.content or ''):
        from .services.template_rendering import validate_merge_fields,TemplateRenderError
        try: validate_merge_fields([],obj.content)
        except TemplateRenderError as exc: return JsonResponse({'error':exc.code,'fields':exc.fields},status=400)
    obj.full_clean(); obj.save()
    append_audit(request=request,action='documents.created',object_type='Document',object_id=obj.pk,after=_serialize(obj)); return JsonResponse(_serialize(obj),status=201)

@require_http_methods(['GET'])
@permission_required('documents.view')
def document_download(request,pk):
    try: obj=Document.objects.get(pk=pk,tenant=request.tenant)
    except Document.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
    if not _document_allowed(request,obj): return JsonResponse({'error':'not_found'},status=404)
    if not obj.file_path: return JsonResponse({'error':'document_has_no_file'},status=409)
    try:
        handle=default_storage.open(obj.file_path,'rb')
    except Exception as exc:
        append_audit(request=request,action='documents.download_failed',object_type='Document',object_id=obj.id,metadata={'error':type(exc).__name__})
        return JsonResponse({'error':'document_storage_unavailable'},status=503)
    from .services.file_security import safe_storage_name
    filename=safe_storage_name(obj.original_name or obj.title,'document.bin')
    response=FileResponse(handle,as_attachment=True,filename=filename,content_type=obj.mime_type or 'application/octet-stream')
    response['Cache-Control']='private, no-store';response['X-Content-Type-Options']='nosniff'
    if obj.sha256: response['X-Document-SHA256']=obj.sha256
    append_audit(request=request,action='documents.downloaded',object_type='Document',object_id=obj.id,metadata={'mime_type':obj.mime_type,'size':obj.file_size})
    return response

@require_http_methods(['GET','PATCH','PUT','DELETE'])
def document_detail(request,pk):
    try: obj=Document.objects.get(pk=pk,tenant=request.tenant)
    except Document.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
    if not _document_allowed(request,obj): return JsonResponse({'error':'not_found'},status=404)
    action={'GET':'view','PATCH':'edit','PUT':'edit','DELETE':'delete'}[request.method]
    if not has_permission(request.tenant_membership.role,f'documents.{action}',effective_custom_permissions(request.tenant_membership)): return JsonResponse({'error':'forbidden'},status=403)
    if request.method=='GET': return JsonResponse(_serialize(obj))
    conflict=_offline_version_conflict(request,obj)
    if conflict: return conflict
    if request.method=='DELETE':
        before=_serialize(obj); stored_path=obj.file_path
        if stored_path:
            try: default_storage.delete(stored_path)
            except Exception as exc: return JsonResponse({'error':'document_storage_delete_failed','detail':type(exc).__name__},status=502)
        obj.delete(); append_audit(request=request,action='documents.deleted',object_type='Document',object_id=pk,before=before); return HttpResponse(status=204)
    before=_serialize(obj); _apply(obj,_body(request),DOCUMENT_FIELDS,request.tenant); obj.tenant=request.tenant
    if obj.template_key or '{{' in (obj.content or ''):
        from .services.template_rendering import validate_merge_fields,TemplateRenderError
        try: validate_merge_fields([],obj.content)
        except TemplateRenderError as exc: return JsonResponse({'error':exc.code,'fields':exc.fields},status=400)
    obj.full_clean(); obj.save(); append_audit(request=request,action='documents.updated',object_type='Document',object_id=obj.pk,before=before,after=_serialize(obj)); return JsonResponse(_serialize(obj))

@require_http_methods(['POST'])
@permission_required('members.edit')
def member_bulk_preflight(request):
    data=_body(request);ids=[str(item) for item in (data.get('ids') or [])];action=str(data.get('action') or '')
    if not ids or len(ids)>500 or len(ids)!=len(set(ids)):return JsonResponse({'error':'unique_ids_required','max_ids':500},status=400)
    parameters={key:data.get(key) for key in ('status','tag') if key in data}
    if action=='set_status' and parameters.get('status') not in dict(Member.STATUS):return JsonResponse({'error':'invalid_status'},status=400)
    if action=='add_tag' and not str(parameters.get('tag') or '').strip():return JsonResponse({'error':'tag_required'},status=400)
    if action not in {'set_status','add_tag'}:return JsonResponse({'error':'unsupported_bulk_action'},status=400)
    impacted=Member.objects.filter(tenant=request.tenant,id__in=ids).count()
    if impacted!=len(ids):return JsonResponse({'error':'one_or_more_members_not_found'},status=404)
    raw=secrets.token_urlsafe(32);fingerprint=bulk_fingerprint(ids,action,parameters)
    BulkActionConfirmation.objects.create(tenant=request.tenant,user=request.user,token_digest=hashlib.sha256(raw.encode()).hexdigest(),fingerprint=fingerprint,action=action,impacted_count=impacted,expires_at=timezone.now()+timezone.timedelta(minutes=10))
    return JsonResponse({'confirmation_token':raw,'action':action,'impacted_count':impacted,'expires_in_seconds':600})

@require_http_methods(['POST'])
@permission_required('members.edit')
def member_bulk(request):
    data=_body(request); ids=[str(item) for item in (data.get('ids') or [])]
    if not ids or len(ids)>500 or len(ids)!=len(set(ids)): return JsonResponse({'error':'unique_ids_required','max_ids':500},status=400)
    qs=Member.objects.filter(tenant=request.tenant,id__in=ids); action=str(data.get('action') or ''); changed=0;parameters={key:data.get(key) for key in ('status','tag') if key in data}
    with transaction.atomic():
        raw=str(data.get('confirmation_token') or '');fingerprint=bulk_fingerprint(ids,action,parameters)
        confirmation=BulkActionConfirmation.objects.select_for_update().filter(tenant=request.tenant,user=request.user,token_digest=hashlib.sha256(raw.encode()).hexdigest()).first() if raw else None
        if not confirmation:return JsonResponse({'error':'bulk_confirmation_required'},status=428)
        if confirmation.consumed_at:return JsonResponse({'error':'bulk_confirmation_already_used'},status=409)
        if confirmation.expires_at<=timezone.now():return JsonResponse({'error':'bulk_confirmation_expired'},status=409)
        if confirmation.fingerprint!=fingerprint or confirmation.action!=action or confirmation.impacted_count!=len(ids):return JsonResponse({'error':'bulk_confirmation_mismatch'},status=409)
        if qs.count()!=len(ids):return JsonResponse({'error':'one_or_more_members_not_found'},status=404)
        confirmation.consumed_at=timezone.now();confirmation.save(update_fields=['consumed_at','updated_at'])
        if action=='set_status':
            status=data.get('status')
            if status not in dict(Member.STATUS): return JsonResponse({'error':'invalid_status'},status=400)
            changed=qs.update(status=status,updated_at=timezone.now())
        elif action=='add_tag':
            tag=str(data.get('tag','')).strip()
            if not tag: return JsonResponse({'error':'tag_required'},status=400)
            for m in qs:
                tags=list(m.tags or [])
                if tag not in tags: tags.append(tag); m.tags=tags; m.save(update_fields=['tags','updated_at']); changed+=1
        else: return JsonResponse({'error':'unsupported_bulk_action'},status=400)
    append_audit(request=request,action='members.bulk_updated',metadata={'action':action,'requested':len(ids),'changed':changed,'confirmation_id':str(confirmation.id)}); return JsonResponse({'changed':changed})

@require_http_methods(['POST'])
@permission_required('members.edit')
def member_merge(request):
    data=_body(request); target_id=data.get('target_id'); source_ids=[str(x) for x in (data.get('source_ids') or []) if str(x)!=str(target_id)]
    if not target_id or not source_ids: return JsonResponse({'error':'target_and_source_ids_required'},status=400)
    try: target=Member.objects.get(pk=target_id,tenant=request.tenant)
    except Member.DoesNotExist: return JsonResponse({'error':'target_not_found'},status=404)
    sources=list(Member.objects.filter(tenant=request.tenant,id__in=source_ids))
    if len(sources)!=len(set(source_ids)): return JsonResponse({'error':'one_or_more_sources_not_found'},status=404)
    with transaction.atomic():
        for src in sources:
            # Preserve financial/audit history; source identity remains as an inactive merged record.
            for field in ('email','phone','address','notes'):
                if not getattr(target,field) and getattr(src,field): setattr(target,field,getattr(src,field))
            target.tags=sorted(set((target.tags or [])+(src.tags or [])))
            target.custom_data={**(src.custom_data or {}),**(target.custom_data or {})}
            Attendance.objects.filter(tenant=request.tenant,member=src).exclude(meeting_id__in=Attendance.objects.filter(tenant=request.tenant,member=target).values('meeting_id')).update(member=target)
            OfficerTerm.objects.filter(tenant=request.tenant,member=src).update(member=target)
            Candidate.objects.filter(tenant=request.tenant,mentor=src).update(mentor=target)
            Milestone.objects.filter(tenant=request.tenant,member=src).update(member=target)
            ParticipationRecord.objects.filter(tenant=request.tenant,member=src).update(member=target)
            src.status='inactive'; src.custom_data={**(src.custom_data or {}),'merged_into':str(target.id),'merged_at':timezone.now().isoformat()}; src.save(update_fields=['status','custom_data','updated_at'])
        target.save()
    append_audit(request=request,action='members.merged',object_type='Member',object_id=target.id,metadata={'sources':source_ids}); return JsonResponse({'target':_serialize(target),'merged_sources':source_ids})

@require_http_methods(['POST'])
def member_anonymize(request,pk):
    if request.tenant_membership.role not in ('owner','admin'): return JsonResponse({'error':'owner_or_admin_required'},status=403)
    try: m=Member.objects.get(pk=pk,tenant=request.tenant)
    except Member.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
    before={'id':str(m.id),'email':m.email,'phone':m.phone,'status':m.status}
    token=str(m.id).replace('-','')[:10]
    m.first_name='Anonymized'; m.last_name=f'Member {token}'; m.email=''; m.phone=''; m.address=''; m.date_of_birth=None; m.notes=''; m.tags=[]; m.custom_data={'anonymized_at':timezone.now().isoformat()}; m.status='inactive'; m.save()
    append_audit(request=request,action='members.anonymized',object_type='Member',object_id=m.id,before=before,after={'status':'inactive'}); return JsonResponse({'ok':True,'member':_serialize(m)})

@require_http_methods(['GET'])
@permission_required('dues.view')
def member_statement(request,pk):
    try: m=Member.objects.get(pk=pk,tenant=request.tenant)
    except Member.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
    currency_filter=str(request.GET.get('currency') or '').upper().strip()
    assessments=DuesAssessment.objects.filter(tenant=request.tenant,member=m).select_related('cycle').order_by('-cycle__period_start')
    payments=Payment.objects.filter(tenant=request.tenant,member=m).order_by('-paid_at')
    if currency_filter:
        assessments=assessments.filter(currency=currency_filter);payments=payments.filter(currency=currency_filter)
    balances={}
    for a in assessments: balances[a.currency]=balances.get(a.currency,Decimal('0'))+amount_due(a.amount,a.paid_amount,a.waived_amount,a.late_fee,a.adjustment_amount)
    balance=sum(balances.values(),Decimal('0')) if len(balances)<=1 else Decimal('0')
    adjustments=FinancialAdjustment.objects.filter(tenant=request.tenant,assessment__member=m).select_related('assessment').order_by('-created_at')
    if currency_filter:adjustments=adjustments.filter(assessment__currency=currency_filter)
    append_audit(request=request,action='dues.statement.viewed',object_type='Member',object_id=m.id)
    if str(request.GET.get('format','json')).lower()=='pdf':
        if len(balances)>1:return JsonResponse({'error':'statement_currency_required','currencies':sorted(balances)},status=409)
        from .services.pdfs import member_statement_pdf
        data=member_statement_pdf(m,assessments,payments,balance,request.tenant.name,request.GET.get('period','All activity'),adjustments=list(adjustments)); doc=_persist_generated_document(request,'member_statement','Member',m.id,data,'member-statement-v1')
        response=HttpResponse(data,content_type='application/pdf');response['Content-Disposition']=f'attachment; filename="member-statement-{m.id}.pdf"';response['X-Document-SHA256']=doc.sha256;return response
    return JsonResponse({'member':{'id':str(m.id),'name':f'{m.first_name} {m.last_name}'},'balance':str(balance) if len(balances)<=1 else None,'balances_by_currency':{k:str(v) for k,v in balances.items()},'assessments':[_serialize(a) for a in assessments],'payments':[_serialize(p) for p in payments],'adjustments':[_serialize(x) for x in adjustments]})


@require_http_methods(['GET','POST'])
@permission_required('candidates.view')
def candidate_tasks(request,pk):
    c=_candidate_for_request(request,pk)
    if not c: return JsonResponse({'error':'not_found'},status=404)
    if request.method=='GET': return JsonResponse({'results':[_serialize(x) for x in c.tasks.filter(tenant=request.tenant).order_by('completed_at','due_on')]})
    if not has_permission(request.tenant_membership.role,'candidates.edit',effective_custom_permissions(request.tenant_membership)): return JsonResponse({'error':'forbidden'},status=403)
    data=_body(request); assignee=None
    if data.get('assignee_id'):
        assignee=Member.objects.filter(pk=data['assignee_id'],tenant=request.tenant).first()
        if not assignee:return JsonResponse({'error':'invalid_or_cross_tenant_assignee'},status=400)
    completed=bool(data.get('completed')) or data.get('status')=='completed'
    obj=CandidateTask.objects.create(tenant=request.tenant,candidate=c,title=data['title'],due_on=data.get('due_on') or None,notes=data.get('notes',''),assignee=assignee,status='completed' if completed else data.get('status','todo'),required_for_stage=bool(data.get('required_for_stage',False)),completed_at=timezone.now() if completed else None)
    append_audit(request=request,action='candidates.task.created',object_type='CandidateTask',object_id=obj.id); return JsonResponse(_serialize(obj),status=201)

@require_http_methods(['PATCH','DELETE'])
@permission_required('candidates.edit')
def candidate_task_detail(request,pk,task_id):
    candidate=_candidate_for_request(request,pk)
    if not candidate: return JsonResponse({'error':'not_found'},status=404)
    try: obj=CandidateTask.objects.get(pk=task_id,candidate=candidate,tenant=request.tenant)
    except CandidateTask.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
    if request.method=='DELETE': obj.delete(); append_audit(request=request,action='candidates.task.deleted',object_type='CandidateTask',object_id=task_id); return HttpResponse(status=204)
    data=_body(request)
    if 'title' in data: obj.title=data['title']
    if 'due_on' in data: obj.due_on=data['due_on'] or None
    if 'notes' in data: obj.notes=data['notes']
    if 'required_for_stage' in data: obj.required_for_stage=bool(data['required_for_stage'])
    if 'assignee_id' in data:
        obj.assignee=Member.objects.filter(pk=data['assignee_id'],tenant=request.tenant).first() if data['assignee_id'] else None
        if data['assignee_id'] and not obj.assignee:return JsonResponse({'error':'invalid_or_cross_tenant_assignee'},status=400)
    if 'status' in data: obj.status=data['status']
    if 'completed' in data or data.get('status')=='completed':
        complete=bool(data.get('completed',data.get('status')=='completed')); obj.completed_at=timezone.now() if complete else None; obj.status='completed' if complete else ('todo' if obj.status=='completed' else obj.status)
    obj.full_clean(); obj.save(); append_audit(request=request,action='candidates.task.updated',object_type='CandidateTask',object_id=obj.id); return JsonResponse(_serialize(obj))

@require_http_methods(['GET','POST','DELETE'])
@permission_required('candidates.view')
def candidate_mentors(request,pk):
    candidate=_candidate_for_request(request,pk)
    if not candidate:return JsonResponse({'error':'not_found'},status=404)
    if request.method=='GET':return JsonResponse({'results':[_serialize(x) for x in candidate.mentor_assignments.filter(is_active=True).select_related('mentor')]})
    if not has_permission(request.tenant_membership.role,'candidates.edit',effective_custom_permissions(request.tenant_membership)):return JsonResponse({'error':'forbidden'},status=403)
    data=_body(request); mentor=Member.objects.filter(pk=data.get('mentor_id'),tenant=request.tenant).first()
    if not mentor:return JsonResponse({'error':'invalid_or_cross_tenant_mentor'},status=400)
    role=str(data.get('role') or 'mentor')[:80]
    if request.method=='DELETE':
        CandidateMentorAssignment.objects.filter(candidate=candidate,mentor=mentor,role=role,is_active=True).update(is_active=False,ended_at=timezone.now()); append_audit(request=request,action='candidates.mentor.removed',object_type='Candidate',object_id=candidate.id,metadata={'mentor_id':str(mentor.id),'role':role}); return JsonResponse({'ok':True})
    obj,_=CandidateMentorAssignment.objects.update_or_create(tenant=request.tenant,candidate=candidate,mentor=mentor,role=role,defaults={'is_active':True,'ended_at':None})
    if not candidate.mentor_id: candidate.mentor=mentor; candidate.save(update_fields=['mentor','updated_at'])
    append_audit(request=request,action='candidates.mentor.assigned',object_type='CandidateMentorAssignment',object_id=obj.id); return JsonResponse(_serialize(obj),status=201)

@require_http_methods(['GET','POST'])
@permission_required('settings.view')
def custom_roles(request):
    if request.method=='GET': return JsonResponse({'results':[_serialize(x) for x in CustomRoleDefinition.objects.filter(tenant=request.tenant).order_by('name')]})
    if not has_permission(request.tenant_membership.role,'settings.admin',effective_custom_permissions(request.tenant_membership)): return JsonResponse({'error':'forbidden'},status=403)
    data=_body(request)
    try: permissions=ensure_grantable_permissions(request.tenant_membership.role,effective_custom_permissions(request.tenant_membership),data.get('permissions',[]))
    except ValueError as exc: return JsonResponse({'error':'invalid_permissions','detail':str(exc)},status=400)
    except PermissionError as exc: return JsonResponse({'error':'privilege_escalation_blocked','detail':str(exc)},status=403)
    name=str(data.get('name') or '').strip()
    if not name or len(name)>120:return JsonResponse({'error':'valid_role_name_required'},status=400)
    approvals=data.get('approval_requirements',{})
    if not isinstance(approvals,dict):return JsonResponse({'error':'invalid_approval_requirements'},status=400)
    obj,created=CustomRoleDefinition.objects.update_or_create(tenant=request.tenant,name=name,defaults={'permissions':permissions,'approval_requirements':approvals})
    append_audit(request=request,action='settings.custom_role.saved',object_type='CustomRoleDefinition',object_id=obj.id,after={'name':obj.name,'permissions':permissions,'approval_requirements':approvals})
    return JsonResponse(_serialize(obj),status=201 if created else 200)

@require_http_methods(['PATCH','DELETE'])
@permission_required('settings.admin')
def custom_role_detail(request,pk):
    try: obj=CustomRoleDefinition.objects.get(pk=pk,tenant=request.tenant)
    except CustomRoleDefinition.DoesNotExist:return JsonResponse({'error':'not_found'},status=404)
    before={'name':obj.name,'permissions':list(obj.permissions or []),'approval_requirements':obj.approval_requirements}
    if request.method=='DELETE':
        if obj.memberships.filter(is_active=True).exists():
            return JsonResponse({'error':'custom_role_in_use','assigned_memberships':obj.memberships.filter(is_active=True).count()},status=409)
        obj.delete(); append_audit(request=request,action='settings.custom_role.deleted',object_type='CustomRoleDefinition',object_id=pk,before=before); return HttpResponse(status=204)
    data=_body(request)
    try: permissions=ensure_grantable_permissions(request.tenant_membership.role,effective_custom_permissions(request.tenant_membership),data.get('permissions',obj.permissions))
    except ValueError as exc:return JsonResponse({'error':'invalid_permissions','detail':str(exc)},status=400)
    except PermissionError as exc:return JsonResponse({'error':'privilege_escalation_blocked','detail':str(exc)},status=403)
    name=str(data.get('name',obj.name)).strip()
    if not name or len(name)>120:return JsonResponse({'error':'valid_role_name_required'},status=400)
    if CustomRoleDefinition.objects.filter(tenant=request.tenant,name=name).exclude(pk=obj.pk).exists():return JsonResponse({'error':'role_name_already_exists'},status=409)
    approvals=data.get('approval_requirements',obj.approval_requirements)
    if not isinstance(approvals,dict):return JsonResponse({'error':'invalid_approval_requirements'},status=400)
    obj.name=name; obj.permissions=permissions; obj.approval_requirements=approvals; obj.save(update_fields=['name','permissions','approval_requirements','updated_at'])
    after={'name':obj.name,'permissions':list(obj.permissions or []),'approval_requirements':obj.approval_requirements}
    append_audit(request=request,action='settings.custom_role.updated',object_type='CustomRoleDefinition',object_id=obj.id,before=before,after=after,metadata={'authorization_refresh':'dynamic'})
    return JsonResponse(_serialize(obj))


def _tenant_credential(tenant,provider,mode='live'):
    obj=_tenant_or_global_credential(tenant,provider,mode)
    if obj is None:raise EncryptedCredential.DoesNotExist
    return obj,decrypt_json(obj.ciphertext)

@require_http_methods(['POST'])
@permission_required('dues.edit')
def assessment_payment_link(request,pk,provider):
    try: a=DuesAssessment.objects.select_related('member').get(pk=pk,tenant=request.tenant)
    except DuesAssessment.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
    provider=provider.lower(); mode=request.GET.get('mode','live').lower()
    if provider not in ('paymongo','xendit'): return JsonResponse({'error':'provider_not_supported_for_local_dues'},status=400)
    if mode not in ('test','live'): return JsonResponse({'error':'invalid_provider_mode'},status=400)
    enabled=((request.tenant.settings or {}).get('payments') or {}).get(provider,{}).get('enabled',True)
    if not enabled: return JsonResponse({'error':'provider_disabled'},status=409)
    due=amount_due(a.amount,a.paid_amount,a.waived_amount,a.late_fee,a.adjustment_amount)
    if due<=0: return JsonResponse({'error':'assessment_has_no_balance'},status=409)
    try: _cred,secrets=_tenant_credential(request.tenant,provider,mode)
    except EncryptedCredential.DoesNotExist: return JsonResponse({'error':'provider_not_configured'},status=503)
    currency=a.currency.upper()
    configured_currency=str(secrets.get('currency') or currency).upper()[:3]
    if configured_currency!=currency:return JsonResponse({'error':'provider_currency_configuration_mismatch','assessment_currency':currency,'provider_currency':configured_currency},status=409)
    internal_ref=f'lfpr-{uuid.uuid4()}'
    payment_request=PaymentRequest.objects.create(tenant=request.tenant,assessment=a,member=a.member,provider=provider,mode=mode,internal_reference=internal_ref,amount=due,currency=currency,status='created',request_payload={'assessment_id':str(a.id),'amount':str(due),'currency':currency})
    try:
        reference=f'lfpr:{payment_request.id}'
        if provider=='paymongo':
            result=PayMongoAdapter(secrets['secret_key']).create_link(int(due*100),f'{request.tenant.name} dues',reference,currency=currency)
            provider_data=result.get('data') or {};url=provider_data.get('url');ref=provider_data.get('id','')
        else:
            result=XenditAdapter(secrets['secret_key']).create_payment_request(reference_id=reference,amount=due,currency=currency,country=secrets.get('country','PH'))
            actions=result.get('actions') or []; url=next((x.get('url') for x in actions if isinstance(x,dict) and x.get('url')),None); ref=result.get('payment_request_id') or result.get('id','')
        payment_request.provider_reference=str(ref)[:180]; payment_request.response_payload=result; payment_request.status='pending'; payment_request.save(update_fields=['provider_reference','response_payload','status','updated_at'])
    except Exception as exc:
        payment_request.status='failed'; payment_request.response_payload={'error':type(exc).__name__}; payment_request.save(update_fields=['status','response_payload','updated_at'])
        append_audit(request=request,action='dues.payment_link.failed',object_type='PaymentRequest',object_id=payment_request.id,metadata={'provider':provider,'error':type(exc).__name__}); return JsonResponse({'error':'provider_request_failed','payment_request_id':str(payment_request.id)},status=502)
    append_audit(request=request,action='dues.payment_link.created',object_type='PaymentRequest',object_id=payment_request.id,metadata={'provider':provider,'provider_reference':ref,'mode':mode})
    return JsonResponse({'provider':provider,'assessment_id':str(a.id),'amount':str(due),'currency':currency,'checkout_url':url,'provider_reference':ref,'payment_request_id':str(payment_request.id),'mode':mode})


@require_http_methods(['POST'])
@login_required
def replay_webhook(request,event_id):
    if not (user_has_platform_permission(request.user,'billing.edit') or user_has_platform_permission(request.user,'logs.view')):
        return JsonResponse({'error':'platform_forbidden','permission':'billing.edit_or_logs.view'},status=403)
    if not has_recent_step_up(request):
        return JsonResponse({'error':'step_up_required','reauthenticate':True},status=403)
    try: event=WebhookEvent.objects.get(pk=event_id)
    except WebhookEvent.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
    from .services.webhook_processing import process_verified_webhook,WebhookQuarantined
    try: result=process_verified_webhook(event)
    except WebhookQuarantined as exc:
        event.refresh_from_db();append_audit(request=request,actor=request.user,tenant=event.tenant,action='platform.webhook.replay_quarantined',object_type='WebhookEvent',object_id=event.id,metadata={'provider':event.provider,'mode':event.mode,'reason':str(exc)});return JsonResponse({'id':str(event.id),'status':'quarantined','reason':str(exc)},status=409)
    except Exception as exc:
        append_audit(request=request,actor=request.user,tenant=event.tenant,action='platform.webhook.replay_failed',object_type='WebhookEvent',object_id=event.id,metadata={'provider':event.provider,'mode':event.mode,'error':type(exc).__name__});return JsonResponse({'id':str(event.id),'status':'failed','error':'replay_failed'},status=502)
    append_audit(request=request,actor=request.user,tenant=event.tenant,action='platform.webhook.replayed',object_type='WebhookEvent',object_id=event.id,metadata={'provider':event.provider,'attempt':event.attempts,'result':result}); return JsonResponse({'id':str(event.id),'status':event.status,'attempts':event.attempts,'result':result})

@require_http_methods(['POST'])
@login_required
def credential_health(request,credential_id):
    try: obj=EncryptedCredential.objects.get(pk=credential_id)
    except EncryptedCredential.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
    required='credentials.billing' if obj.provider in ('stripe','paymongo','xendit') else 'credentials.infrastructure'
    if not user_has_platform_permission(request.user,required) and not user_has_platform_permission(request.user,'credentials.infrastructure'): return JsonResponse({'error':'platform_forbidden','permission':required},status=403)
    try:
        values=decrypt_json(obj.ciphertext);required=PROVIDER_REQUIRED_FIELDS.get(obj.provider,());healthy=bool(values) and all(str(values.get(k) or '').strip() for k in required) if required else bool(values) and all(v not in (None,'') for v in values.values());obj.health_status='healthy' if healthy else 'invalid'
    except Exception: obj.health_status='decrypt_failed'
    obj.last_checked_at=timezone.now(); obj.save(update_fields=['health_status','last_checked_at','updated_at']); append_audit(request=request,actor=request.user,tenant=obj.tenant,action='platform.credential.health_checked',object_type='EncryptedCredential',object_id=obj.id,metadata={'provider':obj.provider,'status':obj.health_status}); return JsonResponse({'id':str(obj.id),'health_status':obj.health_status,'last_checked_at':obj.last_checked_at.isoformat()})

@require_http_methods(['POST'])
@login_required
def credential_toggle(request,credential_id):
    try: obj=EncryptedCredential.objects.get(pk=credential_id)
    except EncryptedCredential.DoesNotExist:return JsonResponse({'error':'not_found'},status=404)
    required='credentials.billing' if obj.provider in ('stripe','paymongo','xendit') else 'credentials.infrastructure'
    if not user_has_platform_permission(request.user,required) and not user_has_platform_permission(request.user,'credentials.infrastructure'):return JsonResponse({'error':'platform_forbidden','permission':required},status=403)
    if not has_recent_step_up(request):return JsonResponse({'error':'step_up_required','reauthenticate':True},status=403)
    data=_body(request);enabled=bool(data.get('is_enabled',not obj.is_enabled));before={'is_enabled':obj.is_enabled,'health_status':obj.health_status}
    obj.is_enabled=enabled;obj.health_status='unchecked' if enabled else 'disabled';obj.last_checked_at=None;obj.save(update_fields=['is_enabled','health_status','last_checked_at','updated_at'])
    action='platform.credential.enabled' if enabled else 'platform.credential.disabled'
    security_alert(kind='credential_enabled' if enabled else 'credential_disabled',message=f'{obj.provider} {obj.mode} credential {"enabled" if enabled else "disabled"}',tenant=obj.tenant,actor=request.user,severity='high',metadata={'provider':obj.provider,'mode':obj.mode,'key_version':obj.key_version})
    append_audit(request=request,actor=request.user,tenant=obj.tenant,action=action,object_type='EncryptedCredential',object_id=obj.id,before=before,after={'is_enabled':obj.is_enabled,'health_status':obj.health_status},metadata={'provider':obj.provider,'mode':obj.mode})
    return JsonResponse({'id':str(obj.id),'provider':obj.provider,'mode':obj.mode,'is_enabled':obj.is_enabled,'health_status':obj.health_status})


@require_http_methods(['POST'])
@login_required
def credential_rotate(request,credential_id):
    if not has_recent_step_up(request): return JsonResponse({'error':'step_up_required','reauthenticate':True},status=403)
    try: obj=EncryptedCredential.objects.get(pk=credential_id)
    except EncryptedCredential.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
    required='credentials.billing' if obj.provider in ('stripe','paymongo','xendit') else 'credentials.infrastructure'
    if not user_has_platform_permission(request.user,required) and not user_has_platform_permission(request.user,'credentials.infrastructure'): return JsonResponse({'error':'platform_forbidden','permission':required},status=403)
    data=_body(request);values=data.get('credentials')
    try:values=_validate_provider_credentials(obj.provider,values)
    except ValueError as exc:return JsonResponse({'error':str(exc)},status=400)
    obj.ciphertext=encrypt_json(values); obj.key_version+=1; obj.health_status='unchecked'; obj.last_checked_at=None; obj.save(update_fields=['ciphertext','key_version','health_status','last_checked_at','updated_at']); security_alert(kind='credential_rotated',message=f'{obj.provider} credential rotated',tenant=obj.tenant,actor=request.user,severity='high',metadata={'provider':obj.provider,'mode':obj.mode,'key_version':obj.key_version}); append_audit(request=request,actor=request.user,tenant=obj.tenant,action='platform.credential.rotated',object_type='EncryptedCredential',object_id=obj.id,metadata={'provider':obj.provider,'key_version':obj.key_version}); return JsonResponse({'id':str(obj.id),'provider':obj.provider,'key_version':obj.key_version})

@require_http_methods(['GET','POST'])
@platform_permission_required('users.view')
def platform_users(request):
    User=get_user_model()
    if request.method=='GET':
        qs=User.objects.all().order_by('username')[:500]; return JsonResponse({'results':[{'id':u.id,'username':u.get_username(),'email':u.email,'is_active':u.is_active,'is_staff':u.is_staff,'platform_roles':list(u.platform_roles.filter(is_active=True).values_list('role',flat=True))} for u in qs]})
    if not user_has_platform_permission(request.user,'users.edit'): return JsonResponse({'error':'platform_forbidden'},status=403)
    data=_body(request); user=User.objects.get(pk=data['user_id']); action=data.get('action')
    if action in {'set_unusable_password','reset_mfa','set_platform_role'} and not has_recent_step_up(request): return JsonResponse({'error':'step_up_required','reauthenticate':True},status=403)
    if action=='lock': user.is_active=False; user.save(update_fields=['is_active'])
    elif action=='unlock': user.is_active=True; user.save(update_fields=['is_active'])
    elif action=='set_unusable_password': user.set_unusable_password(); user.save(update_fields=['password'])
    elif action=='reset_mfa': MFADevice.objects.filter(user=user).delete(); security_alert(kind='mfa_reset',message=f'MFA reset for {user.get_username()}',actor=request.user,severity='high',metadata={'target_user_id':user.id})
    elif action=='set_platform_role':
        if not request.user.is_superuser: return JsonResponse({'error':'super_admin_required_for_platform_roles'},status=403)
        if data.get('role') not in dict(PlatformRoleAssignment.ROLE): return JsonResponse({'error':'invalid_platform_role'},status=400)
        PlatformRoleAssignment.objects.update_or_create(user=user,role=data['role'],defaults={'is_active':bool(data.get('enabled',True))})
    else: return JsonResponse({'error':'unsupported_action'},status=400)
    append_audit(request=request,actor=request.user,action='platform.user.updated',object_type='User',object_id=user.id,metadata={'action':action}); return JsonResponse({'id':user.id,'is_active':user.is_active})

# CMS/SEO operational resources
@require_http_methods(['GET','POST'])
@platform_permission_required('tenants.view')
def platform_announcements(request):
    if not request.user.is_superuser: return JsonResponse({'error':'super_admin_required'},status=403)
    if request.method=='GET': return JsonResponse({'results':[_serialize(x) for x in PlatformAnnouncement.objects.order_by('-created_at')[:250]]})
    if not request.user.is_superuser: return JsonResponse({'error':'platform_write_requires_super_admin'},status=403)
    data=_body(request); obj=PlatformAnnouncement.objects.create(**{k:data[k] for k in ['kind','title','body','starts_at','ends_at','is_active'] if k in data}); append_audit(request=request,actor=request.user,action='platform.announcement.created',object_type='PlatformAnnouncement',object_id=obj.id); return JsonResponse(_serialize(obj),status=201)

@require_http_methods(['PATCH','DELETE'])
@platform_permission_required('tenants.view')
def platform_announcement_detail(request,pk):
    if not request.user.is_superuser: return JsonResponse({'error':'super_admin_required'},status=403)
    try: obj=PlatformAnnouncement.objects.get(pk=pk)
    except PlatformAnnouncement.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
    before=_serialize(obj)
    if request.method=='DELETE':
        obj.delete(); append_audit(request=request,actor=request.user,action='platform.announcement.deleted',object_type='PlatformAnnouncement',object_id=pk,before=before); return HttpResponse(status=204)
    data=_body(request)
    for key in ['kind','title','body','starts_at','ends_at','is_active']:
        if key in data: setattr(obj,key,data[key] or None if key in ('starts_at','ends_at') else data[key])
    obj.full_clean(); obj.save(); append_audit(request=request,actor=request.user,action='platform.announcement.updated',object_type='PlatformAnnouncement',object_id=obj.id,before=before,after=_serialize(obj)); return JsonResponse(_serialize(obj))

@require_http_methods(['GET','POST'])
@platform_permission_required('tenants.view')
def platform_redirects(request):
    if not request.user.is_superuser: return JsonResponse({'error':'super_admin_required'},status=403)
    if request.method=='GET': return JsonResponse({'results':[_serialize(x) for x in RedirectRule.objects.order_by('source')]})
    if not request.user.is_superuser: return JsonResponse({'error':'platform_write_requires_super_admin'},status=403)
    data=_body(request)
    try: source,destination,status_code=validate_redirect_rule(data.get('source'),data.get('destination'),data.get('status_code',301))
    except ValidationError as exc: return JsonResponse({'error':'redirect_validation','detail':exc.message_dict},status=400)
    obj,created=RedirectRule.objects.update_or_create(source=source,defaults={'destination':destination,'status_code':status_code,'is_active':bool(data.get('is_active',True))}); append_audit(request=request,actor=request.user,action='platform.redirect.saved',object_type='RedirectRule',object_id=obj.id,after=_serialize(obj)); return JsonResponse(_serialize(obj),status=201 if created else 200)

@require_http_methods(['PATCH','DELETE'])
@platform_permission_required('tenants.view')
def platform_redirect_detail(request,pk):
    if not request.user.is_superuser: return JsonResponse({'error':'super_admin_required'},status=403)
    try: obj=RedirectRule.objects.get(pk=pk)
    except RedirectRule.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
    before=_serialize(obj)
    if request.method=='DELETE':
        obj.delete(); append_audit(request=request,actor=request.user,action='platform.redirect.deleted',object_type='RedirectRule',object_id=pk,before=before); return HttpResponse(status=204)
    data=_body(request)
    source=data.get('source',obj.source); destination=data.get('destination',obj.destination); status_code=data.get('status_code',obj.status_code)
    try: source,destination,status_code=validate_redirect_rule(source,destination,status_code)
    except ValidationError as exc: return JsonResponse({'error':'redirect_validation','detail':exc.message_dict},status=400)
    obj.source=source; obj.destination=destination; obj.status_code=status_code
    if 'is_active' in data: obj.is_active=bool(data.get('is_active'))
    obj.full_clean(); obj.save(); append_audit(request=request,actor=request.user,action='platform.redirect.updated',object_type='RedirectRule',object_id=obj.id,before=before,after=_serialize(obj)); return JsonResponse(_serialize(obj))

@require_http_methods(['GET','POST'])
@platform_permission_required('tenants.view')
def platform_lead_forms(request):
    if not request.user.is_superuser: return JsonResponse({'error':'super_admin_required'},status=403)
    if request.method=='GET': return JsonResponse({'results':[_serialize(x) for x in LeadForm.objects.order_by('name')]})
    if not request.user.is_superuser: return JsonResponse({'error':'platform_write_requires_super_admin'},status=403)
    data=_body(request)
    try: fields=validate_lead_form_fields(data.get('fields',[]))
    except ValidationError as exc:return JsonResponse({'error':'lead_form_validation','fields':exc.message_dict},status=400)
    routing=data.get('routing',{})
    if not isinstance(routing,dict):return JsonResponse({'error':'invalid_routing'},status=400)
    obj,_=LeadForm.objects.update_or_create(key=data['key'],defaults={'name':data['name'],'fields':fields,'routing':routing,'is_active':bool(data.get('is_active',True))}); append_audit(request=request,actor=request.user,action='platform.lead_form.saved',object_type='LeadForm',object_id=obj.id,after=_serialize(obj)); return JsonResponse(_serialize(obj),status=201)

@require_http_methods(['GET'])
def public_redirect(request):
    source=str(request.GET.get('source') or '')
    if not source.startswith('/'): return JsonResponse({'error':'invalid_source'},status=400)
    obj=RedirectRule.objects.filter(source=source,is_active=True).first()
    if not obj: return JsonResponse({'error':'not_found'},status=404)
    if obj.destination==source: return JsonResponse({'error':'redirect_loop'},status=409)
    return JsonResponse({'destination':obj.destination,'status_code':obj.status_code})

@require_http_methods(['PATCH','DELETE'])
@platform_permission_required('tenants.view')
def platform_lead_form_detail(request,pk):
    if not request.user.is_superuser: return JsonResponse({'error':'super_admin_required'},status=403)
    try: obj=LeadForm.objects.get(pk=pk)
    except LeadForm.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
    before=_serialize(obj)
    if request.method=='DELETE':
        if obj.submissions.exists(): return JsonResponse({'error':'lead_form_has_submissions_disable_instead'},status=409)
        obj.delete(); append_audit(request=request,actor=request.user,action='platform.lead_form.deleted',object_type='LeadForm',object_id=pk,before=before); return HttpResponse(status=204)
    data=_body(request)
    try:
        if 'fields' in data:data['fields']=validate_lead_form_fields(data.get('fields'))
        if 'routing' in data and not isinstance(data.get('routing'),dict):raise ValidationError({'routing':'Routing must be an object.'})
        for key in ['name','fields','routing','is_active']:
            if key in data: setattr(obj,key,data[key])
        obj.full_clean(); obj.save()
    except (ValidationError,ValueError,TypeError) as exc:return JsonResponse({'error':'lead_form_validation','detail':getattr(exc,'message_dict',str(exc))},status=400)
    append_audit(request=request,actor=request.user,action='platform.lead_form.updated',object_type='LeadForm',object_id=obj.id,before=before,after=_serialize(obj)); return JsonResponse(_serialize(obj))

@require_http_methods(['GET'])
@platform_permission_required('tenants.view')
def platform_lead_submissions(request):
    if not request.user.is_superuser: return JsonResponse({'error':'super_admin_required'},status=403)
    qs=LeadSubmission.objects.select_related('form').order_by('-created_at'); form_id=request.GET.get('form_id'); status=request.GET.get('status')
    if form_id: qs=qs.filter(form_id=form_id)
    if status: qs=qs.filter(status=status)
    return JsonResponse({'results':[{'id':str(x.id),'form_id':str(x.form_id),'form':x.form.name,'status':x.status,'data':x.data,'source':x.source,'created_at':x.created_at.isoformat()} for x in qs[:500]]})

@require_http_methods(['PATCH'])
@platform_permission_required('tenants.view')
def platform_lead_submission_detail(request,pk):
    if not request.user.is_superuser: return JsonResponse({'error':'super_admin_required'},status=403)
    try: obj=LeadSubmission.objects.get(pk=pk)
    except LeadSubmission.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
    data=_body(request); status=str(data.get('status') or '')
    if status not in {'new','contacted','qualified','closed','spam'}: return JsonResponse({'error':'invalid_status'},status=400)
    before=_serialize(obj); obj.status=status; obj.save(update_fields=['status','updated_at']); append_audit(request=request,actor=request.user,action='platform.lead_submission.updated',object_type='LeadSubmission',object_id=obj.id,before=before,after={'status':status}); return JsonResponse(_serialize(obj))

@require_http_methods(['GET'])
def robots_txt(request):
    setting=SiteSetting.objects.filter(key='robots').first(); body=(setting.value or {}).get('content') if setting else None
    return HttpResponse(body or 'User-agent: *\nAllow: /\nSitemap: /api/public/sitemap.xml\n',content_type='text/plain')

@require_http_methods(['GET'])
def sitemap_xml(request):
    base=(SiteSetting.objects.filter(key='site').first().value or {}).get('base_url','') if SiteSetting.objects.filter(key='site').exists() else ''
    pages=CMSPage.objects.filter(status='published',robots_index=True).values_list('slug','updated_at')
    urls=''.join(f'<url><loc>{base}/{slug if slug != "home" else ""}</loc><lastmod>{updated.date().isoformat()}</lastmod></url>' for slug,updated in pages)
    return HttpResponse('<?xml version="1.0" encoding="UTF-8"?><urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">'+urls+'</urlset>',content_type='application/xml')

def _meeting_attachment_json(link):
    return {
        'id':str(link.id),'meeting_id':str(link.meeting_id),'document_id':str(link.document_id),
        'label':link.label or link.document.title,'sort_order':link.sort_order,'include_in_pack':link.include_in_pack,
        'document_title':link.document.title,'mime_type':link.document.mime_type,'file_size':link.document.file_size,
        'scan_status':link.document.scan_status,'updated_at':link.updated_at.isoformat(),
    }


def _meeting_pack_document_error(document):
    if not document.file_path:return 'document_has_no_file'
    if str(document.mime_type or '').lower()!='application/pdf':return 'meeting_pack_attachment_must_be_pdf'
    if document.scan_status not in {'clean','not_configured','unavailable_optional'}:return 'meeting_pack_attachment_scan_not_approved'
    limit=int(getattr(django_settings,'MEETING_PACK_ATTACHMENT_MAX_BYTES',15*1024*1024))
    if int(document.file_size or 0)>limit:return 'meeting_pack_attachment_too_large'
    return ''


@require_http_methods(['GET','POST'])
def meeting_attachments(request,meeting_id):
    permission='meetings.view' if request.method=='GET' else 'meetings.edit'
    if not has_permission(request.tenant_membership.role,permission,effective_custom_permissions(request.tenant_membership)):return JsonResponse({'error':'forbidden'},status=403)
    if not has_permission(request.tenant_membership.role,'documents.view',effective_custom_permissions(request.tenant_membership)):return JsonResponse({'error':'forbidden'},status=403)
    try:meeting=Meeting.objects.get(pk=meeting_id,tenant=request.tenant)
    except Meeting.DoesNotExist:return JsonResponse({'error':'not_found'},status=404)
    if request.method=='GET':
        links=MeetingAttachment.objects.filter(tenant=request.tenant,meeting=meeting).select_related('document').order_by('sort_order','created_at')
        visible=[_meeting_attachment_json(x) for x in links if _document_allowed(request,x.document)]
        return JsonResponse({'count':len(visible),'results':visible})
    data=_body(request);document_id=data.get('document_id') or data.get('document')
    try:document=Document.objects.get(pk=document_id,tenant=request.tenant)
    except (Document.DoesNotExist,ValueError,TypeError):return JsonResponse({'error':'document_not_found'},status=404)
    if not _document_allowed(request,document):return JsonResponse({'error':'document_not_found'},status=404)
    error=_meeting_pack_document_error(document)
    if error:return JsonResponse({'error':error},status=409)
    if MeetingAttachment.objects.filter(tenant=request.tenant,meeting=meeting,document=document).exists():return JsonResponse({'error':'attachment_already_configured'},status=409)
    if MeetingAttachment.objects.filter(tenant=request.tenant,meeting=meeting).count()>=20:return JsonResponse({'error':'meeting_pack_attachment_limit'},status=409)
    try:sort_order=max(0,min(1000,int(data.get('sort_order',0) or 0)))
    except (TypeError,ValueError):return JsonResponse({'error':'invalid_sort_order'},status=400)
    label=str(data.get('label') or document.title).strip()[:180]
    link=MeetingAttachment.objects.create(tenant=request.tenant,meeting=meeting,document=document,label=label,sort_order=sort_order,include_in_pack=bool(data.get('include_in_pack',True)),created_by=request.user)
    append_audit(request=request,action='meetings.pack_attachment.created',object_type='MeetingAttachment',object_id=link.id,after=_meeting_attachment_json(link));return JsonResponse(_meeting_attachment_json(link),status=201)


@require_http_methods(['PATCH','DELETE'])
@permission_required('meetings.edit')
def meeting_attachment_detail(request,meeting_id,pk):
    try:link=MeetingAttachment.objects.select_related('document').get(pk=pk,meeting_id=meeting_id,tenant=request.tenant)
    except MeetingAttachment.DoesNotExist:return JsonResponse({'error':'not_found'},status=404)
    if not _document_allowed(request,link.document):return JsonResponse({'error':'not_found'},status=404)
    if request.method=='DELETE':
        before=_meeting_attachment_json(link);link.delete();append_audit(request=request,action='meetings.pack_attachment.deleted',object_type='MeetingAttachment',object_id=pk,before=before);return HttpResponse(status=204)
    data=_body(request);before=_meeting_attachment_json(link)
    if 'label' in data:link.label=str(data.get('label') or link.document.title).strip()[:180]
    if 'sort_order' in data:
        try:link.sort_order=max(0,min(1000,int(data.get('sort_order') or 0)))
        except (TypeError,ValueError):return JsonResponse({'error':'invalid_sort_order'},status=400)
    if 'include_in_pack' in data:
        if bool(data.get('include_in_pack')):
            error=_meeting_pack_document_error(link.document)
            if error:return JsonResponse({'error':error},status=409)
        link.include_in_pack=bool(data.get('include_in_pack'))
    link.save();after=_meeting_attachment_json(link);append_audit(request=request,action='meetings.pack_attachment.updated',object_type='MeetingAttachment',object_id=link.id,before=before,after=after);return JsonResponse(after)


@require_http_methods(['GET'])
@permission_required('officers.view')
def officer_roster(request):
    from .services.pdfs import officer_roster_pdf
    terms=OfficerTerm.objects.filter(tenant=request.tenant,is_current=True).select_related('member').order_by('office','member__last_name');data=officer_roster_pdf(terms,request.tenant.name,request.GET.get('effective_date') or timezone.now().date());doc=_persist_generated_document(request,'officer_roster','Tenant',request.tenant.id,data,'officer-roster-v1');response=HttpResponse(data,content_type='application/pdf');response['Content-Disposition']='attachment; filename="officer-roster.pdf"';response['X-Document-SHA256']=doc.sha256;append_audit(request=request,action='officers.roster.generated',object_type='Tenant',object_id=request.tenant.id);return response

@require_http_methods(['GET'])
@permission_required('candidates.view')
def candidate_progress(request,pk):
    candidate=_candidate_for_request(request,pk)
    if not candidate:return JsonResponse({'error':'not_found'},status=404)
    from .services.pdfs import candidate_progress_pdf
    data=candidate_progress_pdf(candidate,request.tenant.name);doc=_persist_generated_document(request,'candidate_progress','Candidate',candidate.id,data,'candidate-progress-v1');response=HttpResponse(data,content_type='application/pdf');response['Content-Disposition']=f'attachment; filename="candidate-{candidate.id}-progress.pdf"';response['X-Document-SHA256']=doc.sha256;append_audit(request=request,action='candidates.progress.generated',object_type='Candidate',object_id=candidate.id);return response

@require_http_methods(['GET'])
@permission_required('meetings.view')
def meeting_pack(request,meeting_id,kind='pack'):
    if kind not in ('pack','summons','attendance'): return JsonResponse({'error':'invalid_pack_type'},status=400)
    try: meeting=Meeting.objects.get(pk=meeting_id,tenant=request.tenant)
    except Meeting.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
    from .services.pdfs import meeting_pack_pdf
    minute=MinuteVersion.objects.filter(tenant=request.tenant,meeting=meeting).order_by('-version').first()
    attachment_labels=[];attachment_payload=[]
    if kind=='pack':
        previous=Meeting.objects.filter(tenant=request.tenant,starts_at__lt=meeting.starts_at).order_by('-starts_at').first()
        prior=MinuteVersion.objects.filter(tenant=request.tenant,meeting=previous,state__in=['approved','locked']).order_by('-version').first() if previous else None
        minute=prior or minute
        links=MeetingAttachment.objects.filter(tenant=request.tenant,meeting=meeting,include_in_pack=True).select_related('document').order_by('sort_order','created_at')
        per_file_limit=int(getattr(django_settings,'MEETING_PACK_ATTACHMENT_MAX_BYTES',15*1024*1024))
        for link in links:
            document=link.document
            if not _document_allowed(request,document):
                append_audit(request=request,action='meetings.pack_attachment.denied',object_type='Document',object_id=document.id,metadata={'meeting_id':str(meeting.id)})
                return JsonResponse({'error':'meeting_pack_attachment_forbidden','document_id':str(document.id)},status=403)
            error=_meeting_pack_document_error(document)
            if error:return JsonResponse({'error':error,'document_id':str(document.id)},status=409)
            try:
                with default_storage.open(document.file_path,'rb') as handle:payload=handle.read(per_file_limit+1)
            except Exception as exc:
                append_audit(request=request,action='meetings.pack_attachment.read_failed',object_type='Document',object_id=document.id,metadata={'meeting_id':str(meeting.id),'error':type(exc).__name__})
                return JsonResponse({'error':'meeting_pack_attachment_storage_unavailable','document_id':str(document.id)},status=502)
            if len(payload)>per_file_limit:return JsonResponse({'error':'meeting_pack_attachment_too_large','document_id':str(document.id)},status=413)
            label=link.label or document.title;attachment_labels.append(label);attachment_payload.append((label,payload))
    data=meeting_pack_pdf(meeting,minute,meeting.attendance.select_related('member').all(),meeting.visitors.all(),lodge_name=request.tenant.name,kind=kind,attachment_labels=attachment_labels)
    if kind=='pack' and attachment_payload:
        from .services.pdf_merge import merge_pdf_documents,PDFMergeError
        try:data=merge_pdf_documents(data,attachment_payload,max_attachment_bytes=int(getattr(django_settings,'MEETING_PACK_ATTACHMENT_MAX_BYTES',15*1024*1024)),max_total_bytes=int(getattr(django_settings,'MEETING_PACK_MAX_TOTAL_BYTES',40*1024*1024)),max_total_pages=int(getattr(django_settings,'MEETING_PACK_MAX_PAGES',250)))
        except PDFMergeError as exc:
            append_audit(request=request,action='meetings.pack_merge_failed',object_type='Meeting',object_id=meeting.id,metadata={'code':exc.code,'label':exc.label})
            return JsonResponse({'error':exc.code,'attachment':exc.label},status=413 if exc.code in {'pdf_attachment_too_large','pdf_bundle_too_large','pdf_page_limit_exceeded'} else 422)
    doc=_persist_generated_document(request,f'meeting_{kind}','Meeting',meeting.id,data,f'meeting-{kind}-v2')
    response=HttpResponse(data,content_type='application/pdf'); response['Content-Disposition']=f'attachment; filename="meeting-{meeting.id}-{kind}.pdf"'; response['X-Document-SHA256']=doc.sha256
    append_audit(request=request,action=f'meetings.{kind}.pdf_generated',object_type='Meeting',object_id=meeting.id,metadata={'attachment_count':len(attachment_payload)}); return response

@require_http_methods(['POST'])
@permission_required('documents.view')
def render_document_template(request,pk):
    try: template=Document.objects.get(pk=pk,tenant=request.tenant)
    except Document.DoesNotExist: return JsonResponse({'error':'not_found'},status=404)
    if not _document_allowed(request,template): return JsonResponse({'error':'not_found'},status=404)
    if not template.content: return JsonResponse({'error':'template_content_empty'},status=409)
    data=_body(request); context=dict(data.get('context') or {})
    member_id=data.get('member_id')
    if member_id:
        try: m=Member.objects.get(pk=member_id,tenant=request.tenant)
        except Member.DoesNotExist: return JsonResponse({'error':'member_not_found'},status=404)
        context={**{'member.first_name':m.first_name,'member.last_name':m.last_name,'member.email':m.email,'member.phone':m.phone,'member.address':m.address,'member.status':m.status,'member.degree':m.degree},**context}
    context.setdefault('lodge.name',request.tenant.name); context.setdefault('lodge.number',request.tenant.lodge_number); context.setdefault('lodge.jurisdiction',request.tenant.jurisdiction)
    context.setdefault('lodge.locale',request.tenant.locale); context.setdefault('document.generated_date',timezone.now().date().isoformat()); context.setdefault('document.generated_at',timezone.now().isoformat())
    from .services.template_rendering import render_template,TemplateRenderError
    try: rendered=render_template(template.content,context)
    except TemplateRenderError as exc: return JsonResponse({'error':exc.code,'fields':exc.fields},status=400)
    fmt=str(data.get('format','text')).lower(); append_audit(request=request,action='documents.template.rendered',object_type='Document',object_id=template.id,metadata={'format':fmt,'member_id':member_id})
    if fmt=='pdf':
        from .services.pdfs import text_document_pdf
        import hashlib
        profile=LodgeProfile.objects.filter(tenant=request.tenant).first(); pdf=text_document_pdf(template.title,rendered,header=(profile.letterhead if profile else request.tenant.name)); template_version=hashlib.sha256((template.content or '').encode()).hexdigest()[:12]; doc=_persist_generated_document(request,'template_document','Document',template.id,pdf,template.template_key or 'document-template-v1',template_version); response=HttpResponse(pdf,content_type='application/pdf'); response['Content-Disposition']=f'attachment; filename="{template.template_key or template.id}.pdf"'; response['X-Document-SHA256']=doc.sha256; return response
    return JsonResponse({'title':template.title,'content':rendered})

def _validated_billing_return_url(request,value,fallback):
    from urllib.parse import urlsplit
    raw=str(value or fallback or '').strip()
    if not raw:return '/settings'
    if raw.startswith('//'):raise ValueError('protocol_relative_return_url_forbidden')
    parsed=urlsplit(raw)
    if not parsed.scheme and not parsed.netloc:
        if not raw.startswith('/'):raise ValueError('relative_return_url_must_start_with_slash')
        return raw
    if parsed.scheme not in ({'https'} if not django_settings.DEBUG else {'http','https'}):raise ValueError('return_url_requires_https')
    origin=f'{parsed.scheme}://{parsed.netloc}'.rstrip('/')
    tenant_payments=((request.tenant.settings or {}).get('payments') or {})
    site=(SiteSetting.objects.filter(key='billing_plans').first().value if SiteSetting.objects.filter(key='billing_plans').exists() else {}) or {}
    allowed={str(x).rstrip('/') for x in tenant_payments.get('approved_return_origins',[]) if x}
    allowed.update(str(x).rstrip('/') for x in site.get('approved_return_origins',[]) if x)
    allowed.update(str(x).rstrip('/') for x in getattr(django_settings,'CSRF_TRUSTED_ORIGINS',[]) if x)
    if origin not in allowed:raise ValueError('return_url_origin_not_approved')
    return raw


@require_http_methods(['POST'])
@permission_required('settings.admin')
def billing_checkout(request):
    from urllib.parse import urlencode
    data=_body(request);plan=str(data.get('plan','annual'));mode=str(data.get('mode','live')).lower()
    if mode not in {'test','live'}:return JsonResponse({'error':'invalid_provider_mode'},status=400)
    plans=(SiteSetting.objects.filter(key='billing_plans').first().value if SiteSetting.objects.filter(key='billing_plans').exists() else {}) or {};plan_config=plans.get(plan) or {};price_id=plan_config.get('stripe_price_id')
    if not price_id:return JsonResponse({'error':'plan_not_configured'},status=503)
    cred=_tenant_or_global_credential(request.tenant,'stripe',mode)
    if not cred:return JsonResponse({'error':'stripe_not_configured'},status=503)
    try:
        secrets=_validate_provider_credentials('stripe',decrypt_json(cred.ciphertext));success_url=_validated_billing_return_url(request,data.get('success_url'),plan_config.get('success_url') or plans.get('success_url') or '/settings?billing=success');cancel_url=_validated_billing_return_url(request,data.get('cancel_url'),plan_config.get('cancel_url') or plans.get('cancel_url') or '/settings?billing=cancelled')
    except ValueError as exc:return JsonResponse({'error':str(exc)},status=400)
    fields=[('mode','subscription'),('line_items[0][price]',price_id),('line_items[0][quantity]','1'),('client_reference_id',str(request.tenant.id)),('metadata[tenant_id]',str(request.tenant.id)),('metadata[plan]',plan),('metadata[mode]',mode),('subscription_data[metadata][tenant_id]',str(request.tenant.id)),('subscription_data[metadata][plan]',plan),('subscription_data[metadata][mode]',mode),('success_url',success_url),('cancel_url',cancel_url)]
    if data.get('trial_days'):
        try:trial=max(1,min(int(data['trial_days']),90));fields.append(('subscription_data[trial_period_days]',str(trial)))
        except (TypeError,ValueError):return JsonResponse({'error':'invalid_trial_days'},status=400)
    try:result=StripeAdapter(secrets['secret_key']).create_checkout_session(urlencode(fields).encode(),f'sub-{request.tenant.id}-{mode}-{plan}-{uuid.uuid4()}')
    except Exception as exc:append_audit(request=request,action='billing.checkout.failed',metadata={'plan':plan,'mode':mode,'error':type(exc).__name__});return JsonResponse({'error':'stripe_request_failed'},status=502)
    BillingEvent.objects.create(tenant=request.tenant,provider='stripe',event_type='checkout.session.created',provider_reference=str(result.get('id') or '')[:180],mode=mode,status='created',payload={'plan':plan})
    append_audit(request=request,action='billing.checkout.created',metadata={'plan':plan,'mode':mode,'session_id':result.get('id')});return JsonResponse({'id':result.get('id'),'url':result.get('url'),'plan':plan,'mode':mode})


@require_http_methods(['POST'])
@permission_required('settings.admin')
def billing_portal(request):
    from urllib.parse import urlencode
    sub=Subscription.objects.filter(tenant=request.tenant,provider='stripe').first()
    if not sub or not sub.provider_customer_id:return JsonResponse({'error':'stripe_customer_not_found'},status=404)
    data=_body(request);mode=str(data.get('mode','live')).lower()
    if mode not in {'test','live'}:return JsonResponse({'error':'invalid_provider_mode'},status=400)
    cred=_tenant_or_global_credential(request.tenant,'stripe',mode)
    if not cred:return JsonResponse({'error':'stripe_not_configured'},status=503)
    site=(SiteSetting.objects.filter(key='billing_plans').first().value if SiteSetting.objects.filter(key='billing_plans').exists() else {}) or {}
    try:return_url=_validated_billing_return_url(request,data.get('return_url'),site.get('portal_return_url') or '/settings');secrets=_validate_provider_credentials('stripe',decrypt_json(cred.ciphertext))
    except ValueError as exc:return JsonResponse({'error':str(exc)},status=400)
    try:result=StripeAdapter(secrets['secret_key']).create_customer_portal_session(urlencode({'customer':sub.provider_customer_id,'return_url':return_url}).encode(),f'portal-{request.tenant.id}-{mode}-{uuid.uuid4()}')
    except Exception as exc:append_audit(request=request,action='billing.portal.failed',metadata={'mode':mode,'error':type(exc).__name__});return JsonResponse({'error':'stripe_request_failed'},status=502)
    BillingEvent.objects.create(tenant=request.tenant,provider='stripe',event_type='billing_portal.session.created',provider_reference=str(result.get('id') or '')[:180],mode=mode,status='created',payload={})
    append_audit(request=request,action='billing.portal.created',metadata={'mode':mode,'session_id':result.get('id')});return JsonResponse({'id':result.get('id'),'url':result.get('url'),'mode':mode})
